import os
import html
import matplotlib.pyplot as plt
import math

import pandas as pd
def avg_ignore_zeros_cols(row, cols):
    """
    Average of columns in row ignoring zeros/NaN, and ignoring the earliest week column.
    """
    use_cols = _week_cols_excluding_first(row.to_frame().T, cols)
    vals = []
    for c in use_cols:
        v = row.get(c, np.nan)
        if pd.isna(v):
            continue
        try:
            fv = float(v)
        except Exception:
            continue
        if fv == 0:
            continue
        vals.append(fv)
    return float(np.mean(vals)) if vals else 0.0


def _week_cols_excluding_first(df, week_cols):
    """
    Remove the earliest week column from week_cols (to ignore partial first week).
    Uses parsed week start date from the column name when possible.
    """
    if not week_cols:
        return week_cols
    parsed = [pd.to_datetime(c, errors="coerce") for c in week_cols]
    if all(pd.isna(p) for p in parsed):
        return week_cols[1:] if len(week_cols) > 1 else week_cols
    pairs = [(c, p) for c, p in zip(week_cols, parsed) if pd.notna(p)]
    if not pairs:
        return week_cols[1:] if len(week_cols) > 1 else week_cols
    earliest = min(pairs, key=lambda x: x[1])[0]
    return [c for c in week_cols if c != earliest]


import re
from pathlib import Path
from datetime import date, timedelta

import io
import numpy as np
import pandas as pd
import streamlit as st

from .quarter_utils import add_quarter_columns



def style_numeric_posneg(df: pd.DataFrame, cols: list[str]):
    def _s(v):
        try:
            if pd.isna(v):
                return ""
            x = float(v)
        except Exception:
            return ""
        if x > 0:
            return "color: #1a7f37; font-weight:700;"
        if x < 0:
            return "color: #c62828; font-weight:700;"
        return ""
    return df.style.applymap(_s, subset=[c for c in cols if c in df.columns])
# -----------------------------
# Data Coverage + Insights + One-pager Export
# -----------------------------
from datetime import datetime


def make_comparison_pdf_bytes(
    title: str,
    subtitle: str,
    kpi: dict,
    retailers: pd.DataFrame,
    drivers: pd.DataFrame,
    top_increase: pd.DataFrame,
    top_decrease: pd.DataFrame,
    momentum: pd.DataFrame,
    logo_path: str | None = None,
) -> bytes:
    """Executive-style comparison PDF.

    Styled to match the Executive Summary (Weekly Summary) PDF tables/boxes.
    """
    try:
        from io import BytesIO
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate,
            Paragraph,
            Spacer,
            Table,
            TableStyle,
            Image,
            KeepTogether,
            KeepInFrame,
            PageBreak,
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        import re as _re
        import os as _os
    except Exception:
        return b""


def make_multi_year_pdf_bytes(
    title: str,
    subtitle: str,
    table_df: pd.DataFrame,
    years: list[int],
    metric: str = "Sales",
    logo_path: str | None = None,
) -> bytes:
    """Executive-style multi-year PDF.

    This is a separate export from A-vs-B so we can faithfully include *all* selected years.
    """
    try:
        from io import BytesIO
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate,
            Paragraph,
            Spacer,
            Table,
            TableStyle,
            Image,
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        import os as _os
    except Exception:
        return b""

    styles = getSampleStyleSheet()
    if "H1" not in styles:
        styles.add(ParagraphStyle(name="H1", parent=styles["Heading1"], fontSize=16, leading=18, spaceAfter=8))
    if "H2" not in styles:
        styles.add(ParagraphStyle(name="H2", parent=styles["Heading2"], fontSize=12, leading=14, spaceAfter=6))
    if "Body" not in styles:
        styles.add(ParagraphStyle(name="Body", parent=styles["Normal"], fontSize=9.5, leading=12))

    def _make_table(df: pd.DataFrame):
        if df is None or getattr(df, "empty", True):
            return Paragraph("No data.", styles["Body"])
        disp = df.copy()
        # Format columns similar to the executive PDF
        for c in disp.columns:
            cn = str(c).lower()
            num = pd.to_numeric(disp[c], errors="coerce")
            if not num.notna().any():
                disp[c] = disp[c].astype(str)
                continue
            if "sales" in cn:
                disp[c] = num.apply(lambda x: f"${x:,.2f}")
            elif "unit" in cn or "qty" in cn:
                disp[c] = num.apply(lambda x: f"{int(round(x)):,}")
            else:
                disp[c] = num.apply(lambda x: f"{x:,.2f}")

        data = [list(disp.columns)] + disp.astype(str).values.tolist()
        tbl = Table(data)
        ts = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e5e7eb")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
            ("FONTSIZE", (0, 1), (-1, -1), 8.5),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]
        for j, cn in enumerate(disp.columns):
            cn_l = str(cn).lower()
            if any(k in cn_l for k in ["sales", "unit", "%", "diff", "Δ"]):
                ts.append(("ALIGN", (j, 0), (j, -1), "RIGHT"))
        tbl.setStyle(TableStyle(ts))
        return tbl

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, leftMargin=0.75 * inch, rightMargin=0.75 * inch, topMargin=0.75 * inch, bottomMargin=0.75 * inch)
    story = []

    # Header
    header_cells = []
    if logo_path and _os.path.exists(logo_path):
        try:
            header_cells.append(Image(logo_path, width=1.55 * inch, height=0.55 * inch))
        except Exception:
            header_cells.append(Paragraph("", styles["Body"]))
    header_cells.append(Paragraph(f"<b>{html.escape(title)}</b><br/><font size=9 color='#6b7280'>{html.escape(subtitle)}</font>", styles["Body"]))
    hdr = Table([header_cells], colWidths=[1.75 * inch, doc.width - 1.75 * inch])
    hdr.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("BOTTOMPADDING", (0, 0), (-1, -1), 6)]))
    story.append(hdr)
    story.append(Spacer(1, 0.12 * inch))

    yrs = ", ".join([str(y) for y in years])
    story.append(Paragraph(f"Multi-year view ({html.escape(yrs)}) — Highlight metric: {html.escape(metric)}", styles["H2"]))
    story.append(_make_table(table_df))

    try:
        doc.build(story)
        return buf.getvalue()
    except Exception:
        return b""

    styles = getSampleStyleSheet()
    if "H1" not in styles:
        styles.add(ParagraphStyle(name="H1", parent=styles["Heading1"], fontSize=16, leading=18, spaceAfter=8))
    if "H2" not in styles:
        styles.add(ParagraphStyle(name="H2", parent=styles["Heading2"], fontSize=12, leading=14, spaceAfter=6))
    if "Body" not in styles:
        styles.add(ParagraphStyle(name="Body", parent=styles["Normal"], fontSize=9.5, leading=12))

    def _num(v):
        try:
            if v is None:
                return None
            s2 = _re.sub(r"[^0-9\-\.]", "", str(v))
            if s2 in ("", "-", ".", "-."):
                return None
            return float(s2)
        except Exception:
            return None

    def _delta_color(v):
        x = _num(v)
        if x is None:
            return colors.HexColor("#111827")
        if x > 0:
            return colors.HexColor("#1a7f37")
        if x < 0:
            return colors.HexColor("#c62828")
        return colors.HexColor("#111827")

    def _make_table(df: pd.DataFrame, wow_col_name: str | None = None):
        if df is None or getattr(df, "empty", True):
            return Paragraph("No data.", styles["Body"])
        disp = df.copy()
        # format numerics
        for c in disp.columns:
            cn = str(c)
            cn_l = cn.lower()
            num = pd.to_numeric(disp[c], errors="coerce")
            if not num.notna().any():
                disp[c] = disp[c].astype(str)
                continue
            is_delta = ("Δ" in cn) or ("delta" in cn_l) or ("diff" in cn_l) or ("change" in cn_l)
            if "sales" in cn_l:
                disp[c] = num.apply(lambda x: (f"+${x:,.2f}" if x > 0 else f"${x:,.2f}") if is_delta else f"${x:,.2f}")
            elif "unit" in cn_l or "qty" in cn_l:
                disp[c] = num.apply(lambda x: (f"+{int(round(x)):,}" if x > 0 else f"{int(round(x)):,}") if is_delta else f"{int(round(x)):,}")
            else:
                disp[c] = num.apply(lambda x: f"{x:,.2f}")

        data = [list(disp.columns)] + disp.astype(str).values.tolist()
        tbl = Table(data)
        ts = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e5e7eb")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
            ("FONTSIZE", (0, 1), (-1, -1), 8.5),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]
        # align right for numeric-ish columns
        for j, cn in enumerate(disp.columns):
            cn_l = str(cn).lower()
            if any(k in cn_l for k in ["sales", "unit", "Δ", "delta", "diff", "%", "momentum"]):
                ts.append(("ALIGN", (j, 0), (j, -1), "RIGHT"))

        if wow_col_name and wow_col_name in disp.columns:
            j = list(disp.columns).index(wow_col_name)
            for i in range(1, len(data)):
                ts.append(("TEXTCOLOR", (j, i), (j, i), _delta_color(df.iloc[i-1][wow_col_name])))
                ts.append(("FONTNAME", (j, i), (j, i), "Helvetica-Bold"))

        tbl.setStyle(TableStyle(ts))
        return tbl

    def _on_page(c, d):
        c.saveState()
        c.setFont("Helvetica", 8)
        c.setFillColor(colors.HexColor("#6b7280"))
        c.drawRightString(letter[0] - 0.75 * inch, 0.55 * inch, f"Page {d.page}")
        c.restoreState()

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, leftMargin=0.75 * inch, rightMargin=0.75 * inch, topMargin=0.75 * inch, bottomMargin=0.75 * inch)
    story = []

    # Header row
    header_cells = []
    if logo_path and _os.path.exists(logo_path):
        try:
            header_cells.append(Image(logo_path, width=1.55 * inch, height=0.55 * inch))
        except Exception:
            header_cells.append(Paragraph("", styles["Body"]))
    header_cells.append(Paragraph(f"<b>{html.escape(title)}</b><br/><font size=9 color='#6b7280'>{html.escape(subtitle)}</font>", styles["Body"]))
    hdr = Table([header_cells], colWidths=[1.75 * inch, doc.width - 1.75 * inch])
    hdr.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("BOTTOMPADDING", (0, 0), (-1, -1), 6)]))
    story.append(hdr)
    story.append(Spacer(1, 0.12 * inch))

    # KPI hero boxes
    # Expected structure:
    #   kpi = {
    #     "Sales": {"A": "...", "B": "...", "Δ": "..."},
    #     "Units": {"A": "...", "B": "...", "Δ": "..."},
    #     "LabelA": "2025",
    #     "LabelB": "2024",
    #   }
    # Backward-compat: if a flat dict is provided, fall back gracefully.
    sales = (kpi or {}).get("Sales", {}) if isinstance((kpi or {}).get("Sales", {}), dict) else {}
    units = (kpi or {}).get("Units", {}) if isinstance((kpi or {}).get("Units", {}), dict) else {}
    label_a = (kpi or {}).get("LabelA", "A")
    label_b = (kpi or {}).get("LabelB", "B")

    if not sales and isinstance(kpi, dict):
        # Try to reconstruct from legacy keys
        sa = None
        sb = None
        ua = None
        ub = None
        for k, v in (kpi or {}).items():
            if not isinstance(k, str):
                continue
            kl = k.lower()
            if "total sales" in kl and sa is None:
                sa = v
            elif "total sales" in kl and sb is None:
                sb = v
            if "total units" in kl and ua is None:
                ua = v
            elif "total units" in kl and ub is None:
                ub = v
        sales = {"A": sa or "—", "B": sb or "—", "Δ": (kpi or {}).get("Sales Δ (A−B)", "")}
        units = {"A": ua or "—", "B": ub or "—", "Δ": (kpi or {}).get("Units Δ (A−B)", "")}
    hero = Table(
        [[
            Paragraph(
                f"<font size=9 color='#6b7280'>Sales</font><br/>"
                f"<font size=16><b>{html.escape(str(sales.get('A','—')))}</b></font>"
                f"<br/><font size=9 color='#6b7280'>{html.escape(str(label_a))} vs {html.escape(str(label_b))}: {html.escape(str(sales.get('B','—')))}</font>"
                f"<br/><font color='{_delta_color(sales.get('Δ','')).hexval()}'><b>{html.escape(str(sales.get('Δ','')))}</b></font>",
                styles["Body"],
            ),
            Paragraph(
                f"<font size=9 color='#6b7280'>Units</font><br/>"
                f"<font size=16><b>{html.escape(str(units.get('A','—')))}</b></font>"
                f"<br/><font size=9 color='#6b7280'>{html.escape(str(label_a))} vs {html.escape(str(label_b))}: {html.escape(str(units.get('B','—')))}</font>"
                f"<br/><font color='{_delta_color(units.get('Δ','')).hexval()}'><b>{html.escape(str(units.get('Δ','')))}</b></font>",
                styles["Body"],
            ),
        ]],
        colWidths=[doc.width / 2.0, doc.width / 2.0],
    )
    hero.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f3f4f6")), ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")), ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e5e7eb")), ("LEFTPADDING", (0, 0), (-1, -1), 10), ("RIGHTPADDING", (0, 0), (-1, -1), 10), ("TOPPADDING", (0, 0), (-1, -1), 8), ("BOTTOMPADDING", (0, 0), (-1, -1), 8)]))
    story.append(hero)
    story.append(Spacer(1, 0.18 * inch))

    story.append(Paragraph("Retailers (All)", styles["H2"]))
    story.append(_make_table(retailers, wow_col_name="Sales Δ" if retailers is not None and "Sales Δ" in retailers.columns else None))
    story.append(PageBreak())

    story.append(Paragraph("Operational Movement", styles["H1"]))
    if drivers is not None and not getattr(drivers, "empty", True):
        story.append(KeepTogether([Paragraph("Top 5 Drivers (Largest |Sales Δ|)", styles["H2"]), _make_table(drivers, wow_col_name="Sales Δ")]))
        story.append(Spacer(1, 0.14 * inch))
    story.append(KeepTogether([Paragraph("Top Increase SKUs", styles["H2"]), _make_table(top_increase, wow_col_name="Sales Δ")]))
    story.append(Spacer(1, 0.14 * inch))
    story.append(KeepTogether([Paragraph("Top Decrease SKUs", styles["H2"]), _make_table(top_decrease, wow_col_name="Sales Δ")]))
    # No momentum page for now (keep PDF focused on up/down changes)

    try:
        doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)
        return buf.getvalue()
    except Exception:
        return b""


def _consecutive_positive_wow(values):
    """Count consecutive week-over-week increases ending at the most recent week.
    values must be ordered chronologically (oldest -> newest).
    Non-numeric / missing values are treated as 0.0.
    """
    if values is None:
        return 0
    vals = []
    for v in values:
        try:
            if v is None or (isinstance(v, float) and (math.isnan(v) or math.isinf(v))):
                vals.append(0.0)
            else:
                vals.append(float(v))
        except Exception:
            vals.append(0.0)
    if len(vals) < 2:
        return 0
    cnt = 0
    for i in range(len(vals) - 1, 0, -1):
        if (vals[i] - vals[i - 1]) > 0:
            cnt += 1
        else:
            break
    return cnt

def build_data_coverage(df_all: pd.DataFrame) -> dict:
    if df_all is None or df_all.empty or "StartDate" not in df_all.columns:
        return {"ok": False, "msg": "No sales data loaded."}
    d = df_all.copy()
    d["StartDate"] = pd.to_datetime(d["StartDate"], errors="coerce")
    d = d[d["StartDate"].notna()].copy()
    if d.empty:
        return {"ok": False, "msg": "No valid dates in data."}
    d["Year"] = d["StartDate"].dt.year.astype(int)
    years = sorted(d["Year"].unique().tolist())
    overall = {
        "years": years,
        "min_date": d["StartDate"].min().date(),
        "max_date": d["StartDate"].max().date(),
        "rows": int(len(d)),
    }
    by_year = d.groupby("Year", as_index=False).agg(
        Weeks=("StartDate", "nunique"),
        Units=("Units", "sum"),
        Sales=("Sales", "sum"),
        MinDate=("StartDate", "min"),
        MaxDate=("StartDate", "max"),
    ).sort_values("Year")
    by_year["MinDate"] = by_year["MinDate"].dt.date.astype(str)
    by_year["MaxDate"] = by_year["MaxDate"].dt.date.astype(str)

    by_retailer = None
    if "Retailer" in d.columns:
        by_retailer = d.groupby("Retailer", as_index=False).agg(
            Weeks=("StartDate", "nunique"),
            LastWeek=("StartDate", "max"),
            Units=("Units", "sum"),
            Sales=("Sales", "sum"),
        ).sort_values("Sales", ascending=False)
        by_retailer["LastWeek"] = pd.to_datetime(by_retailer["LastWeek"]).dt.date.astype(str)

    by_vendor = None
    if "Vendor" in d.columns:
        by_vendor = d.groupby("Vendor", as_index=False).agg(
            Weeks=("StartDate", "nunique"),
            LastWeek=("StartDate", "max"),
            Units=("Units", "sum"),
            Sales=("Sales", "sum"),
        ).sort_values("Sales", ascending=False)
        by_vendor["LastWeek"] = pd.to_datetime(by_vendor["LastWeek"]).dt.date.astype(str)

    return {"ok": True, "overall": overall, "by_year": by_year, "by_retailer": by_retailer, "by_vendor": by_vendor}

def render_data_coverage_panel(df_all: pd.DataFrame):
    st.markdown("### Data coverage")
    cov = build_data_coverage(df_all)
    if not cov.get("ok"):
        st.info(cov.get("msg", "No data."))
        return
    o = cov["overall"]
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Years loaded", str(len(o["years"])), delta=" / ".join(map(str, o["years"])) if o["years"] else "—")
    c2.metric("Rows", fmt_int(o["rows"]))
    c3.metric("First week", str(o["min_date"]))
    c4.metric("Last week", str(o["max_date"]))

    st.markdown("#### By year")
    by_year_disp = cov["by_year"].copy()
    by_year_disp["Units"] = by_year_disp["Units"].apply(fmt_int)
    by_year_disp["Sales"] = by_year_disp["Sales"].apply(fmt_currency)
    st.dataframe(by_year_disp, use_container_width=True, hide_index=True)

    with st.expander("Coverage by retailer", expanded=False):
        if cov["by_retailer"] is None or cov["by_retailer"].empty:
            st.write("—")
        else:
            br = cov["by_retailer"].copy()
            br["Units"] = br["Units"].apply(fmt_int)
            br["Sales"] = br["Sales"].apply(fmt_currency)
            st.dataframe(br, use_container_width=True, hide_index=True, height=_table_height(br, max_px=650))

    with st.expander("Coverage by vendor", expanded=False):
        if cov["by_vendor"] is None or cov["by_vendor"].empty:
            st.write("—")
        else:
            bv = cov["by_vendor"].copy()
            bv["Units"] = bv["Units"].apply(fmt_int)
            bv["Sales"] = bv["Sales"].apply(fmt_currency)
            st.dataframe(bv, use_container_width=True, hide_index=True, height=_table_height(bv, max_px=650))

def generate_change_insights(a: pd.DataFrame, b: pd.DataFrame, label_a: str, label_b: str, value_col: str) -> list[str]:
    insights = []
    if a is None: a = pd.DataFrame()
    if b is None: b = pd.DataFrame()
    col = value_col




def render_comparison_extras(ctx: dict):
    """Top movers + Explain + One-page PDF export (Comparisons tab only)."""
    if not isinstance(ctx, dict):
        return

    a = ctx.get("a", pd.DataFrame())
    b = ctx.get("b", pd.DataFrame())
    label_a = ctx.get("label_a", "A")
    label_b = ctx.get("label_b", "B")
    value_col = ctx.get("value_col", "Sales")

    if a is None:
        a = pd.DataFrame()
    if b is None:
        b = pd.DataFrame()

    if a.empty and b.empty:
        return

    # -----------------------------
    # Top SKU movers (based on selected metric)
    # -----------------------------
    col = value_col if value_col in ("Units", "Sales") else "Sales"
    st.markdown("---")
    st.markdown("### Top SKU movers")

    if "SKU" not in a.columns and "SKU" not in b.columns:
        st.info("No SKU column available in the comparison data.")
    else:
        ga = a.groupby("SKU", as_index=False).agg(A=(col, "sum")) if (not a.empty and "SKU" in a.columns and col in a.columns) else pd.DataFrame(columns=["SKU", "A"])
        gb = b.groupby("SKU", as_index=False).agg(B=(col, "sum")) if (not b.empty and "SKU" in b.columns and col in b.columns) else pd.DataFrame(columns=["SKU", "B"])
        g = ga.merge(gb, on="SKU", how="outer").fillna(0.0)
        g["Delta"] = g["A"] - g["B"]

        up = g.sort_values("Delta", ascending=False).head(15).copy()
        down = g.sort_values("Delta", ascending=True).head(15).copy()

        fmt = fmt_currency if col == "Sales" else fmt_int
        fmt_s = fmt_currency_signed if col == "Sales" else fmt_int_signed

        def _prep(df_):
            out = df_.copy()
            out["A"] = out["A"].apply(fmt)
            out["B"] = out["B"].apply(fmt)
            out["Delta"] = out["Delta"].apply(fmt_s)
            out.rename(columns={"A": f"{label_a}", "B": f"{label_b}"}, inplace=True)
            return out[["SKU", f"{label_a}", f"{label_b}", "Delta"]]

        c1, c2 = st.columns(2)
        with c1:
            st.caption("Largest increases")
            st.dataframe(_prep(up), use_container_width=True, hide_index=True)
        with c2:
            st.caption("Largest decreases")
            st.dataframe(_prep(down), use_container_width=True, hide_index=True)

    # -----------------------------
    # Explain this change (always)
    # -----------------------------
    insights = []
    try:
        base = generate_change_insights(a, b, label_a, label_b, col)
        if isinstance(base, str):
            insights = [base] if base.strip() else []
        elif isinstance(base, list):
            insights = base
        elif base is None:
            insights = []
        else:
            try:
                insights = list(base)
            except Exception:
                insights = []
    except Exception:
        insights = []

    def _driver_lines(dim: str):
        lines = []
        if dim not in b.columns and dim not in a.columns:
            return lines

        def _dim_table(metric_col: str):
            da = a.groupby(dim, as_index=False).agg(A=(metric_col, "sum")) if (not a.empty and dim in a.columns and metric_col in a.columns) else pd.DataFrame(columns=[dim, "A"])
            db = b.groupby(dim, as_index=False).agg(B=(metric_col, "sum")) if (not b.empty and dim in b.columns and metric_col in b.columns) else pd.DataFrame(columns=[dim, "B"])
            t = da.merge(db, on=dim, how="outer").fillna(0.0)
            t["Delta"] = t["A"] - t["B"]
            up = t.sort_values("Delta", ascending=False).head(3)
            dn = t.sort_values("Delta", ascending=True).head(3)
            return up, dn

        for metric_col, fmt_signed, label in [
            ("Units", fmt_int_signed, "Units"),
            ("Sales", fmt_currency_signed, "Sales"),
        ]:
            up, dn = _dim_table(metric_col)

            up_names = ", ".join([f"{row[dim]} ({fmt_signed(row['Delta'])})" for _, row in up.iterrows() if row["Delta"] != 0])
            dn_names = ", ".join([f"{row[dim]} ({fmt_signed(row['Delta'])})" for _, row in dn.iterrows() if row["Delta"] != 0])

            if up_names:
                lines.append(f"{dim} drivers up ({label}): {up_names}.")
            if dn_names:
                lines.append(f"{dim} drivers down ({label}): {dn_names}.")

        return lines

    driver_lines = []
    driver_lines += _driver_lines("Retailer")
    driver_lines += _driver_lines("Vendor")

    with st.expander("Explain this change", expanded=True):
        for it in (insights[:10] + driver_lines)[:18]:
            st.write(f"- {it}" if it else "")
        if not insights and not driver_lines:
            st.write("—")

    # -----------------------------
    # -----------------------------
    # Executive-style Comparison PDF (matches Weekly Summary styling)
    # -----------------------------
    st.markdown("### Comparison PDF (Executive style)")

    # If the user is in Multi-year mode (2..5 years), export a dedicated PDF that includes ALL selected years.
    if ctx.get("mode") == "multi_year" and isinstance(ctx.get("multi_year_table"), pd.DataFrame):
        try:
            years_sel = ctx.get("multi_year_years") or []
            years_sel = [int(y) for y in years_sel if str(y).strip().isdigit()]
            tbl_my = ctx.get("multi_year_table")
            if years_sel and tbl_my is not None and not tbl_my.empty:
                subtitle_my = f"Years: {', '.join(map(str, years_sel))} • Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                pdf_my = make_multi_year_pdf_bytes(
                    title="Multi-year Comparison",
                    subtitle=subtitle_my,
                    table_df=tbl_my,
                    years=years_sel,
                    metric=value_col if value_col in ("Sales", "Units") else "Sales",
                    logo_path=LOGO_PATH if "LOGO_PATH" in globals() else None,
                )
                if pdf_my:
                    st.download_button(
                        "Download Multi-year PDF (Executive Style)",
                        data=pdf_my,
                        file_name=f"MultiYear_{'_'.join(map(str, years_sel))}.pdf",
                        mime="application/pdf",
                        use_container_width=True,
                        key=f"dl_multi_year_pdf_{'_'.join(map(str, years_sel))}",
                    )
                else:
                    st.caption("Multi-year PDF export unavailable.")
        except Exception:
            st.caption("Multi-year PDF export unavailable.")
    try:
        # Totals
        uA = float(a["Units"].sum()) if (not a.empty and "Units" in a.columns) else 0.0
        uB = float(b["Units"].sum()) if (not b.empty and "Units" in b.columns) else 0.0
        sA = float(a["Sales"].sum()) if (not a.empty and "Sales" in a.columns) else 0.0
        sB = float(b["Sales"].sum()) if (not b.empty and "Sales" in b.columns) else 0.0

        # Δ is A − B so that "A vs B" reads naturally (e.g., 2025 vs 2024)
        du = uA - uB
        ds = sA - sB

        # KPI structure expected by the executive-style PDF
        kpi = {
            "LabelA": label_a,
            "LabelB": label_b,
            "Sales": {"A": fmt_currency(sA), "B": fmt_currency(sB), "Δ": fmt_currency_signed(ds)},
            "Units": {"A": fmt_int(uA), "B": fmt_int(uB), "Δ": fmt_int_signed(du)},
        }

        # Retailers (ALL) table
        retailers_tbl = pd.DataFrame()
        if "Retailer" in a.columns or "Retailer" in b.columns:
            ra = a.groupby("Retailer", as_index=False).agg(SalesA=("Sales","sum"), UnitsA=("Units","sum")) if (not a.empty and "Retailer" in a.columns) else pd.DataFrame(columns=["Retailer","SalesA","UnitsA"])
            rb = b.groupby("Retailer", as_index=False).agg(SalesB=("Sales","sum"), UnitsB=("Units","sum")) if (not b.empty and "Retailer" in b.columns) else pd.DataFrame(columns=["Retailer","SalesB","UnitsB"])
            retailers_tbl = ra.merge(rb, on="Retailer", how="outer").fillna(0.0)
            retailers_tbl["Sales Δ"] = retailers_tbl["SalesA"] - retailers_tbl["SalesB"]
            retailers_tbl["Units Δ"] = retailers_tbl["UnitsA"] - retailers_tbl["UnitsB"]
            retailers_tbl = retailers_tbl.rename(columns={"SalesA": f"Sales {label_a}", "SalesB": f"Sales {label_b}", "UnitsA": f"Units {label_a}", "UnitsB": f"Units {label_b}"})
            retailers_tbl = retailers_tbl.sort_values(f"Sales {label_a}", ascending=False)

        # SKU movers (Top up/down by Sales Δ)
        sku_tbl = pd.DataFrame()
        if "SKU" in a.columns or "SKU" in b.columns:
            sa = a.groupby("SKU", as_index=False).agg(SalesA=("Sales","sum"), UnitsA=("Units","sum")) if (not a.empty and "SKU" in a.columns) else pd.DataFrame(columns=["SKU","SalesA","UnitsA"])
            sb = b.groupby("SKU", as_index=False).agg(SalesB=("Sales","sum"), UnitsB=("Units","sum")) if (not b.empty and "SKU" in b.columns) else pd.DataFrame(columns=["SKU","SalesB","UnitsB"])
            sku_tbl = sa.merge(sb, on="SKU", how="outer").fillna(0.0)
            sku_tbl["Sales Δ"] = sku_tbl["SalesA"] - sku_tbl["SalesB"]
            sku_tbl["Units Δ"] = sku_tbl["UnitsA"] - sku_tbl["UnitsB"]
            sku_tbl = sku_tbl.rename(columns={"SalesA": f"Sales {label_a}", "SalesB": f"Sales {label_b}", "UnitsA": f"Units {label_a}", "UnitsB": f"Units {label_b}"})

        # Keep page 2 compact: fewer rows so Drivers + Up + Down fit on one page
        top_up = sku_tbl.sort_values("Sales Δ", ascending=False).head(8) if not sku_tbl.empty else pd.DataFrame()
        top_dn = sku_tbl.sort_values("Sales Δ", ascending=True).head(8) if not sku_tbl.empty else pd.DataFrame()

        # Top 5 drivers (largest absolute Sales change)
        drivers = pd.DataFrame()
        if not sku_tbl.empty:
            drivers = sku_tbl.assign(_abs=sku_tbl["Sales Δ"].abs()).sort_values("_abs", ascending=False).drop(columns=["_abs"]).head(5)

        # Momentum (within period A): last4 vs prev4 weeks by SKU
        # Momentum removed from PDF for now (keep it focused on up/down changes)
        momentum = pd.DataFrame()

        subtitle = f"{label_a} vs {label_b} • Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        pdf_bytes = make_comparison_pdf_bytes(
            title="Comparison Summary",
            subtitle=subtitle,
            kpi=kpi,
            retailers=retailers_tbl,
            drivers=drivers,
            top_increase=top_up,
            top_decrease=top_dn,
            momentum=momentum,
            logo_path=LOGO_PATH if "LOGO_PATH" in globals() else None,
        )

        if pdf_bytes:
            st.download_button(
                "Download Comparison PDF (Executive Style)",
                data=pdf_bytes,
                file_name=f"Comparison_{label_a}_vs_{label_b}.pdf".replace(" ", "_"),
                mime="application/pdf",
                use_container_width=True,
                key=f"dl_cmp_pdf_{label_a}_{label_b}",
            )
        else:
            st.caption("PDF export unavailable (ReportLab missing or no data).")
    except Exception:
        st.caption("PDF export unavailable.")

def make_one_pager_pdf(title: str, subtitle: str, kpis: list, bullets: list[str], table_df: pd.DataFrame|None) -> bytes:
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import inch
        from reportlab.lib import colors
    except Exception:
        return b""

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)
    w, h = letter

    x = 0.75 * inch
    y = h - 0.75 * inch

    c.setFont("Helvetica-Bold", 16)
    c.drawString(x, y, title)
    y -= 0.25 * inch
    c.setFont("Helvetica", 10)
    c.setFillColorRGB(0.2, 0.2, 0.2)
    c.drawString(x, y, subtitle)
    c.setFillColorRGB(0,0,0)
    y -= 0.35 * inch

    c.setFont("Helvetica-Bold", 9)
    box_w = (w - 1.5*inch) / 3.0
    box_h = 0.55 * inch
    cols = 3
    for i, (label, value, delta) in enumerate(kpis[:6]):
        bx = x + (i % cols) * box_w
        by = y - (i // cols) * (box_h + 0.1*inch)
        c.setStrokeColor(colors.lightgrey)
        c.rect(bx, by - box_h, box_w-6, box_h, stroke=1, fill=0)
        c.setFillColor(colors.black)
        c.drawString(bx+6, by-14, str(label)[:45])
        c.setFont("Helvetica-Bold", 12)
        c.drawString(bx+6, by-32, str(value)[:45])
        c.setFont("Helvetica", 9)
        if delta:
            c.setFillColor(colors.green if str(delta).strip().startswith("+") else colors.red if str(delta).strip().startswith("-") else colors.black)
            c.drawString(bx+6, by-46, str(delta)[:60])
            c.setFillColor(colors.black)
        c.setFont("Helvetica-Bold", 9)
    y -= 2 * (box_h + 0.1*inch) + 0.1*inch

    c.setFont("Helvetica-Bold", 11)
    c.drawString(x, y, "Key insights")
    y -= 0.18*inch
    c.setFont("Helvetica", 9)
    for b in bullets[:8]:
        if y < 1.5*inch:
            break
        c.drawString(x+10, y, f"• {b}"[:110])
        y -= 0.16*inch

    if table_df is not None and not table_df.empty:
        y -= 0.10*inch
        c.setFont("Helvetica-Bold", 11)
        c.drawString(x, y, "Top rows")
        y -= 0.20*inch
        c.setFont("Helvetica", 8)
        df = table_df.head(10).copy()
        cols_show = df.columns.tolist()[:6]
        df = df[cols_show]
        col_w = (w - 1.5*inch) / len(cols_show)
        for j, cn in enumerate(cols_show):
            c.setFont("Helvetica-Bold", 8)
            c.drawString(x + j*col_w, y, str(cn)[:18])
        y -= 0.15*inch
        c.setFont("Helvetica", 8)
        for _, row in df.iterrows():
            if y < 0.9*inch:
                break
            for j, cn in enumerate(cols_show):
                c.drawString(x + j*col_w, y, str(row[cn])[:18])
            y -= 0.13*inch

    c.showPage()
    c.save()
    return buf.getvalue()
MONTH_NAME_TO_NUM = {
    "January": 1, "February": 2, "March": 3, "April": 4,
    "May": 5, "June": 6, "July": 7, "August": 8,
    "September": 9, "October": 10, "November": 11, "December": 12
}

AVG_WINDOW_OPTIONS = ["4 weeks","5 weeks","6 weeks","7 weeks","8 weeks","9 weeks","10 weeks","11 weeks","12 weeks",
                      "January","February","March","April","May","June","July","August","September","October","November","December"]


def _build_month_year_labels(dates: pd.Series) -> list[str]:
    """Return sorted unique month-year labels like 'January 2026' present in dates."""
    try:
        ts = pd.to_datetime(dates, errors="coerce").dropna()
        if ts.empty:
            return []
        per = ts.dt.to_period("M")
        months = sorted(per.unique().tolist())
        return [m.to_timestamp().strftime("%B %Y") for m in months]
    except Exception:
        return []

def resolve_avg_use(avg_window, use_cols, current_year):
    """Return which week columns (week start dates) to use for averaging.

    Supported avg_window values:
      - rolling windows like '8 weeks'
      - month name like 'January' (uses current_year)
      - month-year label like 'January 2026'
    """
    if not use_cols:
        return []

    # Normalize to datetime series for safe .dt ops
    dates = pd.to_datetime(pd.Series(list(use_cols)), errors="coerce")

    # Month-Year label (e.g., 'January 2026')
    if isinstance(avg_window, str):
        m = re.match(r"^([A-Za-z]+)\s+(\d{4})$", avg_window.strip())
        if m:
            mon_name, yr = m.group(1), m.group(2)
            if mon_name in MONTH_NAME_TO_NUM:
                mnum = MONTH_NAME_TO_NUM[mon_name]
                mask = (dates.dt.year == int(yr)) & (dates.dt.month == int(mnum))
                return [c for c, ok in zip(use_cols, mask.fillna(False).tolist()) if ok]

    # Month-only (within current_year)
    if isinstance(avg_window, str) and avg_window in MONTH_NAME_TO_NUM:
        mnum = MONTH_NAME_TO_NUM[avg_window]
        mask = (dates.dt.year == int(current_year)) & (dates.dt.month == int(mnum))
        return [c for c, ok in zip(use_cols, mask.fillna(False).tolist()) if ok]

    # Rolling weeks like '8 weeks'
    if isinstance(avg_window, str) and "week" in avg_window:
        try:
            n = int(avg_window.split()[0])
        except Exception:
            n = 4
        return use_cols[-n:] if len(use_cols) >= n else use_cols

    return list(use_cols)



APP_TITLE = "Sales Dashboard (Vendor Map + Weekly Sheets)"
DATA_DIR = Path("data")
DATA_DIR.mkdir(exist_ok=True)

DEFAULT_VENDOR_MAP = DATA_DIR / "vendor_map.xlsx"
DEFAULT_SALES_STORE = DATA_DIR / "sales_store.csv"
DEFAULT_PRICE_HISTORY = DATA_DIR / "price_history.csv"


# Year locks (prevent accidental edits to closed years)
DEFAULT_YEAR_LOCKS = DATA_DIR / "year_locks.json"

def load_year_locks() -> set[int]:
    try:
        if DEFAULT_YEAR_LOCKS.exists():
            obj = json.loads(DEFAULT_YEAR_LOCKS.read_text(encoding="utf-8"))
            years = obj.get("locked_years", [])
            return set(int(y) for y in years)
    except Exception:
        pass
    return set()

def save_year_locks(locked_years: set[int]) -> None:
    try:
        DEFAULT_YEAR_LOCKS.write_text(json.dumps({"locked_years": sorted(list(locked_years))}, indent=2), encoding="utf-8")
    except Exception:
        return

def overwrite_sales_rows(target_year: int, retailers: set[str]) -> None:
    """Remove rows from sales_store.csv for the given year + retailers."""
    if not DEFAULT_SALES_STORE.exists():
        return
    try:
        cur = pd.read_csv(DEFAULT_SALES_STORE)
        cur["StartDate"] = pd.to_datetime(cur.get("StartDate"), errors="coerce")
        cur["Retailer"] = cur.get("Retailer", "").map(_normalize_retailer)
        retailers_n = {_normalize_retailer(r) for r in retailers}
        keep = ~((cur["StartDate"].dt.year == int(target_year)) & (cur["Retailer"].isin(retailers_n)))
        cur2 = cur[keep].copy()
        cur2.to_csv(DEFAULT_SALES_STORE, index=False)
    except Exception:
        return

# -------------------------
# Normalization
# -------------------------
def _normalize_retailer(x: str) -> str:
    if x is None:
        return ""
    x = str(x).strip()
    aliases = {
        "home depot": "Depot",
        "depot": "Depot",
        "the home depot": "Depot",
        "lowes": "Lowe's",
        "lowe's": "Lowe's",
        "tractor supply": "Tractor Supply",
        "tsc": "Tractor Supply",
        "amazon": "Amazon",
    }
    key = re.sub(r"\s+", " ", x.lower()).strip()
    return aliases.get(key, x)

def _normalize_sku(x: str) -> str:
    if pd.isna(x):
        return ""
    return str(x).strip()

# -------------------------
# Formatting
# -------------------------
def fmt_currency(x) -> str:
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return ""
    try:
        v = float(x)
    except Exception:
        return ""
    s = f"${abs(v):,.2f}"
    return f"({s})" if v < 0 else s

def fmt_int(x) -> str:
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return ""
    try:
        v = float(x)
    except Exception:
        return ""
    return f"{int(round(v)):,.0f}"





def fmt_currency_signed(x) -> str:
    try:
        x = float(x)
    except Exception:
        return str(x)
    return f"-${abs(x):,.0f}" if x < 0 else f"${x:,.0f}"



def _diff_color(v):
    """Return CSS color style for signed deltas (green positive, red negative)."""
    try:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return ""
        x = float(v)
    except Exception:
        return ""
    if x > 0:
        return "color: #0a7d00; font-weight: 600;"
    if x < 0:
        return "color: #b00020; font-weight: 600;"
    return ""



def create_app_backup_zip() -> bytes:
    """Create an in-memory ZIP backup of app data directories/files."""
    import io, zipfile, os
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        paths = []
        for name in ["DATA_DIR", "SAVED_VIEWS_DIR", "BACKUP_DIR", "UPLOAD_DIR", "REPORTS_DIR"]:
            p = globals().get(name)
            if isinstance(p, str) and p:
                paths.append(p)
        for name in ["VENDOR_MAP_PATH", "VENDOR_MAP_FILE", "VENDOR_MAP_DEFAULT_PATH"]:
            p = globals().get(name)
            if isinstance(p, str) and p:
                paths.append(p)

        def add_path(path: str):
            if not path or not os.path.exists(path):
                return
            if os.path.isfile(path):
                arc = os.path.relpath(path, start=os.getcwd())
                z.write(path, arcname=arc)
            else:
                for root, _, files in os.walk(path):
                    for fn in files:
                        full = os.path.join(root, fn)
                        arc = os.path.relpath(full, start=os.getcwd())
                        z.write(full, arcname=arc)

        seen=set()
        for p in paths:
            if p in seen:
                continue
            seen.add(p)
            add_path(p)

    buf.seek(0)
    return buf.getvalue()


def restore_app_backup_zip(zip_bytes: bytes) -> tuple[bool, str]:
    """Restore a ZIP created by create_app_backup_zip into the current working directory."""
    import io, zipfile, os
    try:
        with zipfile.ZipFile(io.BytesIO(zip_bytes), "r") as z:
            for member in z.namelist():
                if member.startswith("/") or ".." in member.split("/"):
                    return (False, f"Unsafe path in zip: {member}")
            z.extractall(path=os.getcwd())
        return (True, "Restore complete. Please rerun the app.")
    except Exception as e:
        return (False, f"Restore failed: {e}")

def fmt_int_signed(x) -> str:
    try:
        x = float(x)
    except Exception:
        return str(x)
    x = int(round(x))
    return f"-{abs(x):,}" if x < 0 else f"{x:,}"

def make_unique_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Ensure dataframe has unique column names (pyarrow requirement)."""
    cols = list(df.columns)
    seen = {}
    new_cols = []
    for c in cols:
        if c not in seen:
            seen[c] = 0
            new_cols.append(c)
        else:
            seen[c] += 1
            new_cols.append(f"{c}.{seen[c]}")
    df2 = df.copy()
    df2.columns = new_cols
    return df2

def fmt_2(x) -> str:
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return ""
    try:
        v = float(x)
    except Exception:
        return ""
    return f"{v:,.2f}"

def _color(v) -> str:
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return "inherit"
    try:
        v = float(v)
    except Exception:
        return "inherit"
    if v > 0:
        return "green"
    if v < 0:
        return "red"
    return "inherit"

def _table_height(df: pd.DataFrame, row_px: int = 32, header_px: int = 38, max_px: int = 1100) -> int:
    if df is None:
        return 220
    n = int(df.shape[0])
    h = header_px + (n + 1) * row_px
    return int(min(max(h, 220), max_px))

def style_currency_cols(df: pd.DataFrame, diff_cols=None):
    diff_cols = diff_cols or []
    sty = df.style
    # format all non-first columns as currency
    first = df.columns[0]
    fmt = {c: (lambda v: fmt_currency(v)) for c in df.columns if c != first}
    sty = sty.format(fmt)
    for c in diff_cols:
        if c in df.columns:
            sty = sty.applymap(lambda v: f"color: {_color(v)};", subset=[c])
    return sty

# -------------------------
# Vendor map
# -------------------------
def load_vendor_map(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path)
    cols = {c.lower().strip(): c for c in df.columns}

    def pick(name, fallbacks):
        for k in [name] + fallbacks:
            if k in cols:
                return cols[k]
        return None

    c_retail = pick("retailer", [])
    c_sku = pick("sku", ["item", "item sku"])
    c_vendor = pick("vendor", ["supplier"])
    c_price = pick("price", ["unit price", "cost"])

    out = pd.DataFrame({
        "Retailer": df[c_retail] if c_retail else "",
        "SKU": df[c_sku] if c_sku else "",
        "Vendor": df[c_vendor] if c_vendor else "",
        "Price": df[c_price] if c_price else np.nan,
    })

    out["Retailer"] = out["Retailer"].map(_normalize_retailer)
    out["SKU"] = out["SKU"].map(_normalize_sku)
    out["Vendor"] = out["Vendor"].astype(str).str.strip()
    out["Price"] = pd.to_numeric(out["Price"], errors="coerce")

    # preserve order per retailer
    out["MapOrder"] = 0
    for r, grp in out.groupby("Retailer", sort=False):
        for j, ix in enumerate(grp.index.tolist()):
            out.loc[ix, "MapOrder"] = j

    return out

# -------------------------
# Sales store
# -------------------------
def load_sales_store() -> pd.DataFrame:
    if DEFAULT_SALES_STORE.exists():
        df = pd.read_csv(DEFAULT_SALES_STORE)
        for c in ["StartDate", "EndDate"]:
            if c in df.columns:
                df[c] = pd.to_datetime(df[c], errors="coerce")
        df["Retailer"] = df["Retailer"].map(_normalize_retailer)
        df["SKU"] = df["SKU"].map(_normalize_sku)
        df["Units"] = pd.to_numeric(df["Units"], errors="coerce").fillna(0.0)
        if "UnitPrice" in df.columns:
            df["UnitPrice"] = pd.to_numeric(df["UnitPrice"], errors="coerce")
        else:
            df["UnitPrice"] = np.nan
        return df
    return pd.DataFrame(columns=["Retailer","SKU","Units","UnitPrice","StartDate","EndDate","SourceFile"])



# -------------------------
# Price history (effective dating)
# -------------------------
def _normalize_price_retailer(x):
    x = "" if x is None else str(x).strip()
    if x == "" or x.lower() in {"all","*", "any"}:
        return "*"
    return _normalize_retailer(x)

def load_price_history() -> pd.DataFrame:
    """
    Returns columns: Retailer, SKU, Price, StartDate (datetime64)
    Retailer="*" means applies to all retailers for that SKU.
    """
    if DEFAULT_PRICE_HISTORY.exists():
        ph = pd.read_csv(DEFAULT_PRICE_HISTORY)
        # flexible column names
        colmap = {c.lower(): c for c in ph.columns}
        sku_col = colmap.get("sku") or colmap.get("sku#") or colmap.get("skunumber") or colmap.get("skuid")
        price_col = colmap.get("price") or colmap.get("unitprice") or colmap.get("unit_price")
        date_col = colmap.get("startdate") or colmap.get("start_date") or colmap.get("effective_date") or colmap.get("date")
        ret_col = colmap.get("retailer")

        if sku_col:
            ph["SKU"] = ph[sku_col].map(_normalize_sku)
        else:
            ph["SKU"] = ""
        if price_col:
            ph["Price"] = pd.to_numeric(ph[price_col], errors="coerce")
        else:
            ph["Price"] = np.nan
        if date_col:
            ph["StartDate"] = pd.to_datetime(ph[date_col], errors="coerce")
        else:
            ph["StartDate"] = pd.NaT
        if ret_col:
            ph["Retailer"] = ph[ret_col].map(_normalize_price_retailer)
        else:
            ph["Retailer"] = "*"

        ph = ph[["Retailer","SKU","Price","StartDate"]].dropna(subset=["SKU","Price","StartDate"])
        ph = ph.sort_values(["Retailer","SKU","StartDate"]).reset_index(drop=True)
        return ph
    return pd.DataFrame(columns=["Retailer","SKU","Price","StartDate"])

def save_price_history(ph: pd.DataFrame) -> None:
    ph2 = ph.copy()
    ph2["StartDate"] = pd.to_datetime(ph2["StartDate"], errors="coerce")
    ph2 = ph2.dropna(subset=["Retailer","SKU","Price","StartDate"])
    ph2["Retailer"] = ph2["Retailer"].map(_normalize_price_retailer)
    ph2["SKU"] = ph2["SKU"].map(_normalize_sku)
    ph2["Price"] = pd.to_numeric(ph2["Price"], errors="coerce")
    ph2 = ph2.sort_values(["Retailer","SKU","StartDate"]).reset_index(drop=True)
    ph2.to_csv(DEFAULT_PRICE_HISTORY, index=False)





# -------------------------
# Caching helpers (performance)
# -------------------------
def _file_mtime(p: Path) -> float:
    try:
        return float(p.stat().st_mtime)
    except Exception:
        return 0.0

@st.cache_data(show_spinner=False)
def cached_vendor_map(path_str: str, mtime: float) -> pd.DataFrame:
    # mtime is included to invalidate cache when the file changes
    return load_vendor_map(Path(path_str))

@st.cache_data(show_spinner=False)
def cached_sales_store(mtime: float) -> pd.DataFrame:
    return load_sales_store()

@st.cache_data(show_spinner=False)
def cached_price_history(mtime: float) -> pd.DataFrame:
    return load_price_history()
def _prepare_price_history_upload(new_rows: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Normalize a price history upload. Returns:
      - normalized rows to consider (SKU, Retailer, StartDate, Price)
      - rows ignored (for reporting)

    Rules:
      - Price blank/NaN => ignored
      - Price <= 0 => ignored (treated as blank)
      - Missing SKU or StartDate => ignored

    Column name matching is forgiving (spaces/underscores/case).
    Accepts: SKU, Retailer, Price, StartDate (or "Start Date", "Effective Date", etc.)
    """
    n = new_rows.copy()

    def norm_key(s: str) -> str:
        s = str(s).strip().lower()
        # keep only alphanumerics to make "Start Date" == "start_date" == "StartDate"
        return re.sub(r"[^a-z0-9]+", "", s)

    cols = {norm_key(c): c for c in n.columns}

    def pick(*keys):
        for k in keys:
            if k in cols:
                return cols[k]
        return None

    sku_col = pick("sku", "sku#", "skunumber", "skuid", "itemsku")
    price_col = pick("price", "unitprice", "unit_price")
    date_col = pick("startdate", "start_date", "startdateeffective", "effectivedate", "effective_date", "start", "date", "startdate1", "startdate2", "startdate3", "startdate4", "startdate5", "startdate6", "startdate7", "startdate8", "startdate9", "startdate10", "startdate11", "startdate12", "startdate13", "startdate14", "startdate15", "startdate16", "startdate17", "startdate18", "startdate19", "startdate20", "startdate21", "startdate22", "startdate23", "startdate24", "startdate25", "startdate26", "startdate27", "startdate28", "startdate29", "startdate30", "startdate31", "startdate32", "startdate33", "startdate34", "startdate35", "startdate36", "startdate37", "startdate38", "startdate39", "startdate40", "startdate41", "startdate42", "startdate43", "startdate44", "startdate45", "startdate46", "startdate47", "startdate48", "startdate49", "startdate50", "startdate51", "startdate52", "startdate53", "startdate54", "startdate55", "startdate56", "startdate57", "startdate58", "startdate59", "startdate60")
    # Common: "start date"
    if date_col is None:
        date_col = pick("startdate", "startdate", "startdate")  # no-op, just for clarity
        date_col = cols.get("startdate") or cols.get("startdate")  # no-op

    # Explicitly support "Start Date" / "Effective Date"
    if date_col is None:
        date_col = pick("startdate", "startdate")  # still none
    if date_col is None:
        date_col = pick("startdate")  # still none

    # Final fallback: try any column that normalizes to "startdate"
    if date_col is None and "startdate" in cols:
        date_col = cols["startdate"]

    ret_col = pick("retailer", "store", "channel")

    if not sku_col or not price_col or not date_col:
        raise ValueError("Price history upload must include columns for SKU, Price, and StartDate (e.g., 'Start Date').")

    norm = pd.DataFrame({
        "SKU": n[sku_col].map(_normalize_sku),
        "Price": pd.to_numeric(n[price_col], errors="coerce"),
        "StartDate": pd.to_datetime(n[date_col], errors="coerce"),
        "Retailer": n[ret_col].map(_normalize_price_retailer) if ret_col else "*",
    })

    ignored = norm.copy()
    ignored["IgnoreReason"] = ""

    mask = ignored["SKU"].isna() | (ignored["SKU"].astype(str).str.strip() == "")
    ignored.loc[mask, "IgnoreReason"] = "Missing SKU"

    mask = ignored["StartDate"].isna()
    ignored.loc[mask, "IgnoreReason"] = np.where(
        ignored.loc[mask, "IgnoreReason"].eq(""),
        "Missing StartDate",
        ignored.loc[mask, "IgnoreReason"]
    )

    mask = ignored["Price"].isna()
    ignored.loc[mask, "IgnoreReason"] = np.where(
        ignored.loc[mask, "IgnoreReason"].eq(""),
        "Blank Price",
        ignored.loc[mask, "IgnoreReason"]
    )

    mask = (ignored["Price"].notna()) & (ignored["Price"] <= 0)
    ignored.loc[mask, "IgnoreReason"] = np.where(
        ignored.loc[mask, "IgnoreReason"].eq(""),
        "Price <= 0",
        ignored.loc[mask, "IgnoreReason"]
    )

    keep = norm.dropna(subset=["SKU","StartDate","Price"]).copy()
    keep = keep[keep["Price"] > 0].copy()

    ignored = ignored[ignored["IgnoreReason"] != ""].copy()
    keep = keep.reset_index(drop=True)
    ignored = ignored.reset_index(drop=True)
    return keep, ignored


def _price_history_diff(cur: pd.DataFrame, incoming: pd.DataFrame) -> pd.DataFrame:
    """
    Build a diff table for (Retailer, SKU, StartDate).
    Actions: insert/update/noop
    """
    if cur is None or cur.empty:
        base = incoming.copy()
        base["OldPrice"] = np.nan
        base["Action"] = "insert"
        base["PriceDiff"] = np.nan
        return base[["Retailer","SKU","StartDate","OldPrice","Price","PriceDiff","Action"]].sort_values(["Retailer","SKU","StartDate"])

    cur2 = cur.copy()
    cur2["StartDate"] = pd.to_datetime(cur2["StartDate"], errors="coerce")
    inc = incoming.copy()
    inc["StartDate"] = pd.to_datetime(inc["StartDate"], errors="coerce")

    key = ["Retailer","SKU","StartDate"]
    merged = inc.merge(cur2[key + ["Price"]].rename(columns={"Price":"OldPrice"}), on=key, how="left")
    merged["PriceDiff"] = merged["Price"] - merged["OldPrice"]
    merged["Action"] = np.where(merged["OldPrice"].isna(), "insert",
                        np.where(np.isclose(merged["Price"], merged["OldPrice"], equal_nan=True), "noop", "update"))
    return merged[key + ["OldPrice","Price","PriceDiff","Action"]].sort_values(key)

def upsert_price_history(new_rows: pd.DataFrame) -> tuple[int, int, int]:
    """
    Upsert price history with effective dates.
    Returns (inserted, updated, ignored_noop) counts for reporting.
    """
    cur = load_price_history()
    incoming, _ignored = _prepare_price_history_upload(new_rows)

    if incoming.empty:
        return (0, 0, 0)

    diff = _price_history_diff(cur, incoming)
    to_apply = diff[diff["Action"].isin(["insert","update"])].copy()
    noop = int((diff["Action"] == "noop").sum())

    if to_apply.empty:
        return (0, 0, noop)

    apply_rows = to_apply[["Retailer","SKU","StartDate","Price"]].copy()

    merged = pd.concat([cur, apply_rows], ignore_index=True) if (cur is not None and not cur.empty) else apply_rows.copy()
    merged["StartDate"] = pd.to_datetime(merged["StartDate"], errors="coerce")
    merged = merged.dropna(subset=["SKU","Price","StartDate"])
    merged = merged.drop_duplicates(subset=["Retailer","SKU","StartDate"], keep="last")
    merged = merged.sort_values(["Retailer","SKU","StartDate"]).reset_index(drop=True)
    save_price_history(merged)

    inserted = int((diff["Action"] == "insert").sum())
    updated = int((diff["Action"] == "update").sum())
    return (inserted, updated, noop)



def apply_effective_prices(base: pd.DataFrame, vmap: pd.DataFrame, ph: pd.DataFrame) -> pd.DataFrame:
    """
    Hybrid pricing:
      1) If UnitPrice is provided on the weekly sheet, ALWAYS use it (locks history).
      2) Else, use effective-date price history (retailer-specific first, then wildcard '*' retailer for all).
      3) Else, fall back to vendor map Price.

    Notes:
      - merge_asof requires non-null, sorted datetime keys.
    """
    base = base.copy()

    # Ensure expected columns exist
    if "Price" not in base.columns:
        base["Price"] = np.nan
    if "UnitPrice" not in base.columns:
        base["UnitPrice"] = np.nan

    base["StartDate"] = pd.to_datetime(base["StartDate"], errors="coerce")

    # Start with vendor-map price, then let weekly UnitPrice override
    base["PriceEffective"] = base["Price"]
    base["PriceEffective"] = base["UnitPrice"].combine_first(base["PriceEffective"])

    # If no price history, finish
    if ph is None or ph.empty:
        return base

    ph = ph.copy()
    ph["StartDate"] = pd.to_datetime(ph["StartDate"], errors="coerce")
    ph = ph.dropna(subset=["SKU", "StartDate", "Price"]).copy()
    if ph.empty:
        return base

    # Normalize keys
    if "Retailer" not in ph.columns:
        ph["Retailer"] = "*"
    ph["Retailer"] = ph["Retailer"].fillna("*").astype(str).str.strip()
    ph["SKU"] = ph["SKU"].map(_normalize_sku)
    base["SKU"] = base["SKU"].map(_normalize_sku)

    # merge_asof cannot handle NaT in the 'on' key
    base_valid = base[base["StartDate"].notna()].copy()
    base_invalid = base[base["StartDate"].isna()].copy()

    # Retailer-specific history (not '*')
    ph_exact = ph[ph["Retailer"] != "*"].copy()
    ph_star = ph[ph["Retailer"] == "*"].copy()

    # Apply retailer-specific prices
    if not ph_exact.empty and not base_valid.empty:
        b1 = base_valid.sort_values(["StartDate","Retailer","SKU"], kind="mergesort").reset_index(drop=True)
        p1 = ph_exact.sort_values(["StartDate","Retailer","SKU"], kind="mergesort").reset_index(drop=True)

        exact = pd.merge_asof(
            b1,
            p1[["Retailer", "SKU", "StartDate", "Price"]].rename(columns={"Price": "PH_Price"}),
            by=["Retailer", "SKU"],
            on="StartDate",
            direction="backward",
            allow_exact_matches=True,
        )

        # Only use PH_Price when UnitPrice is missing
        exact["PriceEffective"] = exact["UnitPrice"].combine_first(exact["PH_Price"]).combine_first(exact["PriceEffective"])
        exact = exact.drop(columns=["PH_Price"], errors="ignore")
        base_valid = exact

    # Apply wildcard prices to rows still missing PriceEffective (and no UnitPrice)
    if not ph_star.empty and not base_valid.empty:
        missing = base_valid["UnitPrice"].isna() & base_valid["PriceEffective"].isna()
        if missing.any():
            b2 = base_valid.loc[missing].copy()
            b2 = b2.sort_values(["StartDate","SKU"], kind="mergesort").reset_index(drop=True)
            p2 = ph_star.sort_values(["StartDate","SKU"], kind="mergesort").reset_index(drop=True)

            star = pd.merge_asof(
                b2,
                p2[["SKU", "StartDate", "Price"]].rename(columns={"Price": "PH_PriceStar"}),
                by=["SKU"],
                on="StartDate",
                direction="backward",
                allow_exact_matches=True,
            )
            base_valid.loc[missing, "PriceEffective"] = star["PH_PriceStar"].values

    # Final: ensure UnitPrice still wins
    if not base_valid.empty:
        base_valid["PriceEffective"] = base_valid["UnitPrice"].combine_first(base_valid["PriceEffective"])

    # Recombine
    base_out = pd.concat([base_valid, base_invalid], ignore_index=True)
    return base_out

    ph = ph.copy()
    ph["StartDate"] = pd.to_datetime(ph["StartDate"], errors="coerce")
    ph = ph.dropna(subset=["SKU", "StartDate", "Price"]).copy()

    if ph.empty:
        return base

    # Normalize retailer field
    if "Retailer" not in ph.columns:
        ph["Retailer"] = "*"
    ph["Retailer"] = ph["Retailer"].fillna("*").astype(str).str.strip()
    ph["SKU"] = ph["SKU"].map(_normalize_sku)
    base["SKU"] = base["SKU"].map(_normalize_sku)

    # Retailer-specific history (not '*')
    ph_exact = ph[ph["Retailer"] != "*"].copy()
    # Wildcard history applies to all retailers
    ph_star = ph[ph["Retailer"] == "*"].copy()

    # Apply retailer-specific prices using merge_asof
    if not ph_exact.empty:
        b1 = base.sort_values(["Retailer", "SKU", "StartDate"]).reset_index(drop=True)
        p1 = ph_exact.sort_values(["Retailer", "SKU", "StartDate"]).reset_index(drop=True)

        # merge_asof requires both sides sorted by by-keys then on-key
        exact = pd.merge_asof(
            b1,
            p1[["Retailer", "SKU", "StartDate", "Price"]].rename(columns={"Price": "PH_Price"}),
            by=["Retailer", "SKU"],
            on="StartDate",
            direction="backward",
            allow_exact_matches=True,
        )
        base = exact

        # Only use PH_Price when UnitPrice is missing
        base["PriceEffective"] = base["UnitPrice"].combine_first(base["PH_Price"]).combine_first(base["PriceEffective"])
        base = base.drop(columns=["PH_Price"], errors="ignore")

    # Apply wildcard prices for any rows still missing an effective price (and no UnitPrice)
    if not ph_star.empty:
        missing = base["UnitPrice"].isna() & base["PriceEffective"].isna()
        if missing.any():
            b2 = base.loc[missing].copy()
            b2 = b2.sort_values(["SKU", "StartDate"]).reset_index(drop=True)
            p2 = ph_star.sort_values(["SKU", "StartDate"]).reset_index(drop=True)

            star = pd.merge_asof(
                b2,
                p2[["SKU", "StartDate", "Price"]].rename(columns={"Price": "PH_PriceStar"}),
                by=["SKU"],
                on="StartDate",
                direction="backward",
                allow_exact_matches=True,
            )
            base.loc[missing, "PriceEffective"] = star["PH_PriceStar"].values

    # Final: ensure UnitPrice still wins
    base["PriceEffective"] = base["UnitPrice"].combine_first(base["PriceEffective"])

    return base

def upsert_sales(existing: pd.DataFrame, new_rows: pd.DataFrame) -> pd.DataFrame:
    if existing is None or existing.empty:
        return new_rows.copy()
    if new_rows is None or new_rows.empty:
        return existing.copy()

    for c in ["StartDate","EndDate"]:
        if c in existing.columns:
            existing[c] = pd.to_datetime(existing[c], errors="coerce")
        if c in new_rows.columns:
            new_rows[c] = pd.to_datetime(new_rows[c], errors="coerce")

    key_cols = ["Retailer","SKU","StartDate","EndDate","SourceFile"]
    combined = pd.concat([existing, new_rows], ignore_index=True)
    combined = combined.drop_duplicates(subset=key_cols, keep="last")
    return combined

def append_sales_to_store(new_rows: pd.DataFrame) -> None:
    if new_rows is None or new_rows.empty:
        return
    existing = load_sales_store()
    combined = upsert_sales(existing, new_rows)
    combined.to_csv(DEFAULT_SALES_STORE, index=False)

# -------------------------
# Weekly workbook ingestion
# -------------------------
def parse_date_range_from_filename(name: str, year_hint: int):
    n = name.lower()

    m = re.search(r"(\d{4})[-_/](\d{1,2})[-_/](\d{1,2}).*?(?:thru|through|to|–|-).*?(\d{4})[-_/](\d{1,2})[-_/](\d{1,2})", n)
    if m:
        y1, mo1, d1, y2, mo2, d2 = map(int, m.groups())
        return pd.Timestamp(date(y1, mo1, d1)), pd.Timestamp(date(y2, mo2, d2))

    m = re.search(r"(\d{1,2})[-_/](\d{1,2}).*?(?:thru|through|to|–|-).*?(\d{1,2})[-_/](\d{1,2})", n)
    if m:
        mo1, d1, mo2, d2 = map(int, m.groups())
        y = int(year_hint)
        return pd.Timestamp(date(y, mo1, d1)), pd.Timestamp(date(y, mo2, d2))

    return None, None


def read_weekly_workbook(uploaded_file, year: int) -> pd.DataFrame:
    """Read a weekly sales workbook where each sheet is a retailer.
    Expected layout per sheet:
      - Column A: SKU (no header required)
      - Column B: Units
      - Optional Column C: UnitPrice
    NOTE: Some retailers (e.g. Zoro/HomeSelects) may have only a single data row.
    Pandas can sometimes interpret that as header-only depending on the engine,
    so we include an openpyxl fallback to reliably read the first rows.
    """
    import openpyxl
    from io import BytesIO

    # Prefer openpyxl engine for consistency on Streamlit Cloud
    try:
        xls = pd.ExcelFile(uploaded_file, engine="openpyxl")
    except Exception:
        xls = pd.ExcelFile(uploaded_file)

    fname = getattr(uploaded_file, "name", "upload.xlsx")
    sdt, edt = parse_date_range_from_filename(fname, year_hint=year)
    if sdt is None:
        sdt = pd.Timestamp(date.today() - timedelta(days=7))
        edt = pd.Timestamp(date.today())

    # Build an openpyxl workbook once for fallback reads
    wb = None
    try:
        data = uploaded_file.getvalue() if hasattr(uploaded_file, "getvalue") else uploaded_file.read()
        wb = openpyxl.load_workbook(BytesIO(data), data_only=True, read_only=True, keep_links=False)
    except Exception:
        wb = None

    dfs = []
    for sh in xls.sheet_names:
        retailer = _normalize_retailer(sh)

        # Primary read (no headers)
        try:
            raw = pd.read_excel(xls, sheet_name=sh, header=None, engine="openpyxl")
        except Exception:
            raw = pd.read_excel(xls, sheet_name=sh, header=None)

        # Fallback: if pandas returns empty but the sheet has a single row (common for new retailers)
        if (raw is None) or (raw.shape[0] == 0) or (raw.shape[1] < 2):
            if wb is not None and sh in wb.sheetnames:
                ws = wb[sh]
                vals = []
                # read first 500 rows, up to 3 cols, stopping after a run of blanks
                blank_run = 0
                for r in range(1, 501):
                    sku = ws.cell(row=r, column=1).value
                    units = ws.cell(row=r, column=2).value
                    price = ws.cell(row=r, column=3).value
                    if (sku is None or str(sku).strip() == "") and (units is None or str(units).strip() == "") and (price is None or str(price).strip() == ""):
                        blank_run += 1
                        if blank_run >= 20:
                            break
                        continue
                    blank_run = 0
                    vals.append([sku, units, price])
                if vals:
                    raw = pd.DataFrame(vals)
                else:
                    continue
            else:
                continue

        # Keep only first 3 columns
        raw = raw.iloc[:, :3].copy() if raw.shape[1] >= 3 else raw.iloc[:, :2].copy()
        raw.columns = ["SKU", "Units", "UnitPrice"] if raw.shape[1] == 3 else ["SKU", "Units"]

        raw["SKU"] = raw["SKU"].map(_normalize_sku)
        raw["Units"] = pd.to_numeric(raw["Units"], errors="coerce").fillna(0.0)

        if "UnitPrice" in raw.columns:
            raw["UnitPrice"] = pd.to_numeric(raw["UnitPrice"], errors="coerce")
        else:
            raw["UnitPrice"] = np.nan

        raw = raw[raw["SKU"].astype(str).str.strip().ne("")]

        raw["Retailer"] = retailer
        raw["StartDate"] = pd.to_datetime(sdt)
        raw["EndDate"] = pd.to_datetime(edt)
        raw["SourceFile"] = fname
        if "UnitPrice" not in raw.columns:
            raw["UnitPrice"] = np.nan
        dfs.append(raw[["Retailer","SKU","Units","UnitPrice","StartDate","EndDate","SourceFile"]])

    out = pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame(columns=["Retailer","SKU","Units","UnitPrice","StartDate","EndDate","SourceFile"])
    if not out.empty:
        out["Retailer"] = out["Retailer"].map(_normalize_retailer)
        out["SKU"] = out["SKU"].map(_normalize_sku)
        out["StartDate"] = pd.to_datetime(out["StartDate"], errors="coerce")
        out["EndDate"] = pd.to_datetime(out["EndDate"], errors="coerce")
        out = add_quarter_columns(out, week_column="StartDate")
    return out


# -------------------------
# Year-Overview (YOW) workbook ingestion
# -------------------------
def parse_week_range_header(val, year: int):
    """Parse headers like '1-1 / 1-3' into (StartDate, EndDate) timestamps.
    Accepts a few common variants and handles year-crossing weeks (e.g. '12-29 / 1-2').
    """
    if val is None:
        return (None, None)
    s = str(val).strip()
    if s == "":
        return (None, None)

    # Common: 'M-D / M-D' or 'M/D - M/D'
    m = re.search(r"(\d{1,2})\s*[-/]\s*(\d{1,2})\s*(?:/|to|–|-)+\s*(\d{1,2})\s*[-/]\s*(\d{1,2})", s)
    if not m:
        # Variant with explicit months: '1-1 / 1-3' (same as above but stricter)
        m = re.search(r"(\d{1,2})-(\d{1,2})\s*/\s*(\d{1,2})-(\d{1,2})", s)
    if not m:
        return (None, None)

    mo1, d1, mo2, d2 = map(int, m.groups())
    y1 = int(year)
    y2 = int(year)
    # If the end month is earlier than start month, assume it crosses into next year
    if mo2 < mo1:
        y2 = y1 + 1

    try:
        sdt = pd.Timestamp(y1, mo1, d1)
        edt = pd.Timestamp(y2, mo2, d2)
        return (sdt, edt)
    except Exception:
        return (None, None)

def read_yow_workbook(uploaded_file, year: int) -> pd.DataFrame:
    """Read a Year Overview workbook:
    - One sheet per retailer OR a single sheet where A1 is the retailer name.
    - Row 1 contains week ranges across the top (starting in column B).
    - Column A contains SKUs (starting row 2).
    - Cells contain Units.
    """
    import openpyxl

    fname = getattr(uploaded_file, "name", "yow.xlsx")

    # openpyxl is fastest/most tolerant for wide sheets
    wb = openpyxl.load_workbook(uploaded_file, data_only=True, read_only=True, keep_links=False)

    rows_out = []

    for sh in wb.sheetnames:
        ws = wb[sh]

        # Retailer name: A1 (preferred). If blank, fall back to sheet name.
        retailer_name = ws["A1"].value
        retailer = _normalize_retailer(retailer_name if retailer_name not in [None, ""] else sh)

        # Header row: week ranges from B1 onward until blank
        week_cols = []
        col = 2  # B
        while True:
            v = ws.cell(row=1, column=col).value
            if v is None or str(v).strip() == "":
                break
            sdt, edt = parse_week_range_header(v, year=year)
            if sdt is None:
                # Try interpreting as a date (week start) if someone uses real date headers
                dt = pd.to_datetime(v, errors="coerce")
                if pd.notna(dt):
                    sdt = pd.Timestamp(dt).normalize()
                    edt = sdt + pd.Timedelta(days=6)
                else:
                    # stop if header isn't parseable
                    break
            if edt is None:
                edt = sdt + pd.Timedelta(days=6)
            week_cols.append((col, pd.Timestamp(sdt), pd.Timestamp(edt), str(v).strip()))
            col += 1

        if not week_cols:
            continue

        # Data rows: SKUs down column A from row 2 until blank
        row = 2
        while True:
            sku = ws.cell(row=row, column=1).value
            if sku is None or str(sku).strip() == "":
                break
            sku = _normalize_sku(sku)

            for (cidx, sdt, edt, hdr) in week_cols:
                units = ws.cell(row=row, column=cidx).value
                if units is None or (isinstance(units, str) and units.strip() == ""):
                    continue
                try:
                    u = float(units)
                except Exception:
                    continue
                if np.isnan(u) or u == 0:
                    continue

                rows_out.append({
                    "Retailer": retailer,
                    "SKU": sku,
                    "Units": float(u),
                    "UnitPrice": np.nan,          # use current pricing (vendor map / price history)
                    "StartDate": pd.to_datetime(sdt),
                    "EndDate": pd.to_datetime(edt),
                    "SourceFile": f"{fname}::{sh}",
                })

            row += 1

    out = pd.DataFrame(rows_out)
    if not out.empty:
        out["Retailer"] = out["Retailer"].map(_normalize_retailer)
        out["SKU"] = out["SKU"].map(_normalize_sku)
        out["StartDate"] = pd.to_datetime(out["StartDate"], errors="coerce")
        out["EndDate"] = pd.to_datetime(out["EndDate"], errors="coerce")
        out["Units"] = pd.to_numeric(out["Units"], errors="coerce").fillna(0.0)
        out["UnitPrice"] = pd.to_numeric(out["UnitPrice"], errors="coerce")
        out = add_quarter_columns(out, week_column="StartDate")
    return out

# -------------------------
# Enrichment / metrics
# -------------------------
def enrich_sales(sales: pd.DataFrame, vmap: pd.DataFrame, price_hist: pd.DataFrame | None = None) -> pd.DataFrame:
    s = sales.copy()
    s["Retailer"] = s["Retailer"].map(_normalize_retailer)
    s["SKU"] = s["SKU"].map(_normalize_sku)
    s["Units"] = pd.to_numeric(s["Units"], errors="coerce").fillna(0.0).astype(float)
    s["StartDate"] = pd.to_datetime(s["StartDate"], errors="coerce")
    s["EndDate"] = pd.to_datetime(s["EndDate"], errors="coerce")
    s = add_quarter_columns(s, week_column="StartDate")

    m = vmap[["Retailer","SKU","Vendor","Price","MapOrder"]].copy()
    # Normalize keys; treat blank Retailer in vendor map as wildcard ("*")
    m["Retailer"] = m["Retailer"].fillna("*").map(_normalize_retailer)
    m["Retailer"] = m["Retailer"].replace({"": "*"})
    m["SKU"] = m["SKU"].map(_normalize_sku)
    m["Price"] = pd.to_numeric(m["Price"], errors="coerce")

    # 1) Exact match on Retailer + SKU
    out = s.merge(m[m["Retailer"] != "*"], on=["Retailer","SKU"], how="left")

    # 2) Wildcard match on SKU only for rows still missing pricing/vendor
    wild = m[m["Retailer"] == "*"][["SKU","Vendor","Price","MapOrder"]].drop_duplicates()
    if not wild.empty:
        miss = out["Price"].isna()
        if miss.any():
            out2 = out.loc[miss].merge(wild, on=["SKU"], how="left", suffixes=("", "_w"))
            if "Vendor_w" in out2.columns:
                out.loc[miss, "Vendor"] = out2["Vendor"].combine_first(out2["Vendor_w"]).values
            if "Price_w" in out2.columns:
                out.loc[miss, "Price"] = out2["Price"].combine_first(out2["Price_w"]).values
            if "MapOrder_w" in out2.columns:
                out.loc[miss, "MapOrder"] = out2["MapOrder"].combine_first(out2["MapOrder_w"]).values

    # Apply effective-dated pricing (if provided), otherwise fallback to vendor map price
    ph = price_hist if price_hist is not None else load_price_history()
    out = apply_effective_prices(out, vmap, ph)


    # Compute Sales from Units and effective price (Units-only weekly uploads)
    out["Units"] = pd.to_numeric(out.get("Units", 0), errors="coerce").fillna(0.0)
    out["PriceEffective"] = pd.to_numeric(out.get("PriceEffective", np.nan), errors="coerce")
    out["Sales"] = (out["Units"] * out["PriceEffective"]).fillna(0.0)
    # Drop rows with no activity to keep 0/0 Retailers & Vendors out of all tables
    out = out[(pd.to_numeric(out.get("Units"), errors="coerce").fillna(0) > 0) | (pd.to_numeric(out.get("Sales"), errors="coerce").fillna(0) > 0)].copy()
    return out



@st.cache_data(show_spinner=False)
def cached_enrich_sales(store_mtime: float, vmap_path: str, vmap_mtime: float, ph_mtime: float) -> pd.DataFrame:
    """
    Cache the expensive enrichment step (merges + effective-dated pricing).
    Cache is invalidated automatically when:
      - sales_store.csv changes (store_mtime)
      - vendor_map.xlsx changes (vmap_mtime)
      - price_history.csv changes (ph_mtime)
    """
    sales_store = load_sales_store()
    vmap = load_vendor_map(Path(vmap_path))
    price_hist = load_price_history()
    return enrich_sales(sales_store, vmap, price_hist)
def wow_mom_metrics(df: pd.DataFrame) -> dict:
    out = {"total_units":0.0,"total_sales":0.0,"wow_units":None,"wow_sales":None,"mom_units":None,"mom_sales":None}
    if df is None or df.empty:
        return out
    d = df.copy()
    d["StartDate"] = pd.to_datetime(d["StartDate"], errors="coerce")
    out["total_units"] = float(d["Units"].sum())
    out["total_sales"] = float(d["Sales"].fillna(0).sum())

    periods = sorted(d["StartDate"].dropna().dt.date.unique().tolist())
    if len(periods) >= 1:
        cur_p = periods[-1]
        cur = d[d["StartDate"].dt.date == cur_p]
        cur_u = cur["Units"].sum()
        cur_s = cur["Sales"].fillna(0).sum()
        if len(periods) >= 2:
            prev_p = periods[-2]
            prev = d[d["StartDate"].dt.date == prev_p]
            prev_u = prev["Units"].sum()
            prev_s = prev["Sales"].fillna(0).sum()
        else:
            prev_u = 0.0
            prev_s = 0.0
        out["wow_units"] = float(cur_u - prev_u)
        out["wow_sales"] = float(cur_s - prev_s)

    d["MonthP"] = d["StartDate"].dt.to_period("M")
    months = sorted(d["MonthP"].dropna().unique().tolist())
    if len(months) >= 1:
        cur_m = months[-1]
        cur = d[d["MonthP"] == cur_m]
        cur_u = cur["Units"].sum()
        cur_s = cur["Sales"].fillna(0).sum()
        if len(months) >= 2:
            prev_m = months[-2]
            prev = d[d["MonthP"] == prev_m]
            prev_u = prev["Units"].sum()
            prev_s = prev["Sales"].fillna(0).sum()
        else:
            prev_u = 0.0
            prev_s = 0.0
        out["mom_units"] = float(cur_u - prev_u)
        out["mom_sales"] = float(cur_s - prev_s)

    return out

def month_label(p: pd.Period) -> str:
    return p.to_timestamp().strftime("%B %Y")

# -------------------------
# App UI
# -------------------------
BUILD_ID = "BULLETPROOF1"



def run_app():
    st.set_page_config(page_title=APP_TITLE, layout="wide")

    # --- Company Logo ---
    try:
        st.image("cornerstone_logo.jpg", width=350)
    except Exception:
        pass

    with st.sidebar:
        st.header("Data Inputs")
        edit_mode = st.checkbox("Enable Edit Mode (edit Vendor/Price)", value=False)

        this_year = date.today().year
        year = st.selectbox("Year (for filename date parsing)", options=list(range(this_year-3, this_year+2)), index=3)
        view_year = st.selectbox("View Year (dashboard)", options=list(range(this_year-3, this_year+2)), index=3, key="view_year")

        st.subheader("Vendor Map")
        vm_upload = st.file_uploader("Upload Vendor Map (.xlsx)", type=["xlsx"], key="vm_up")

        # Backup / Restore (sidebar)
        st.markdown("---")
        with st.expander("Backup / Restore", expanded=False):
            st.caption("Download your sales data as a CSV, or restore by uploading the CSV back in.")
            # Download sales store CSV (no zip)
            if DEFAULT_SALES_STORE.exists():
                dl_name = "sales_store.csv"
                st.download_button(
                    "Download Sales Data (CSV)",
                    data=DEFAULT_SALES_STORE.read_bytes(),
                    file_name=dl_name,
                    mime="text/csv",
                    key="sb_download_sales_store_csv",
                )
            else:
                st.warning("No sales_store.csv found yet in the data folder.")

            st.markdown("---")

            # Restore sales store CSV (no zip)
            # Upload first, then click Restore. This avoids Streamlit "reloading…" hangs.
            if "sb_restore_uploader_key" not in st.session_state:
                st.session_state.sb_restore_uploader_key = 0

            up_csv = st.file_uploader(
                "Restore: upload sales_store.csv",
                type=["csv"],
                key=f"sb_restore_sales_store_csv_{st.session_state.sb_restore_uploader_key}",
            )

            do_restore = st.button(
                "Restore sales_store.csv",
                disabled=(up_csv is None),
                use_container_width=True,
                key="sb_btn_restore_sales_store_csv",
            )

            if do_restore and up_csv is not None:
                try:
                    DEFAULT_SALES_STORE.write_bytes(up_csv.getvalue())
                    # Clear cached data so the app immediately reflects the restored file
                    st.cache_data.clear()
                    # Reset uploader so you don't need to click the X
                    st.session_state.sb_restore_uploader_key += 1
                    st.success("Sales data restored (sales_store.csv). Reloading…")
                    st.rerun()
                except Exception as e:
                    st.error(f"Restore failed: {e}")
        a, b = st.columns(2)
        with a:
            if st.button("Use uploaded as default", disabled=vm_upload is None):
                DEFAULT_VENDOR_MAP.write_bytes(vm_upload.getbuffer())
                st.success("Saved as default vendor map.")
                st.rerun()
        with b:
            if st.button("Reload"):
                st.rerun()

        st.subheader("Weekly Sales Workbooks")
        wk_uploads = st.file_uploader("Upload weekly sales workbook(s) (.xlsx)", type=["xlsx"], accept_multiple_files=True, key="wk_up")
        if st.button("Ingest uploads", disabled=not wk_uploads):
            all_new = []
            prog = st.progress(0, text="Reading workbooks…")
            total = len(wk_uploads)
            for i, f in enumerate(wk_uploads, start=1):
                try:
                    new_rows = read_weekly_workbook(f, year=year)
                    if new_rows is not None and not new_rows.empty:
                        all_new.append(new_rows)
                except Exception as e:
                    st.error(f"Failed to read {getattr(f, 'name', 'upload')}: {e}")
                prog.progress(i / max(total, 1), text=f"Reading workbooks… ({i}/{total})")
            prog.empty()

            if all_new:
                combined_new = pd.concat(all_new, ignore_index=True)
                append_sales_to_store(combined_new)
                st.success(f"Ingested {len(all_new)} workbook(s) into the sales store.")
            else:
                st.info("No rows found in the uploaded workbooks.")
            st.rerun()

        st.divider()
        if st.button("Clear ALL stored sales data"):
            if DEFAULT_SALES_STORE.exists():
                DEFAULT_SALES_STORE.unlink()
            st.warning("Sales store cleared.")
            st.rerun()


    # Ensure view_year exists for downstream tabs
    view_year = st.session_state.get('view_year', year)



    # Load vendor map (persistent)
    BUNDLED_VENDOR_MAP = Path(__file__).parent / "vendor_map.xlsx"

    # If a default vendor map hasn't been set yet, seed it from the bundled file in the repo.
    try:
        if (not DEFAULT_VENDOR_MAP.exists()) and BUNDLED_VENDOR_MAP.exists():
            DEFAULT_VENDOR_MAP.write_bytes(BUNDLED_VENDOR_MAP.read_bytes())
    except Exception:
        pass

    if vm_upload is not None:
        tmp = DATA_DIR / "_session_vendor_map.xlsx"
        tmp.write_bytes(vm_upload.getbuffer())
        vmap_path_used = str(tmp)
        vmap = cached_vendor_map(vmap_path_used, _file_mtime(tmp))
    elif DEFAULT_VENDOR_MAP.exists():
        vmap_path_used = str(DEFAULT_VENDOR_MAP)
        vmap = cached_vendor_map(vmap_path_used, _file_mtime(DEFAULT_VENDOR_MAP))
    else:
        st.info("Upload a vendor map to begin.")
        st.stop()

    sales_store = cached_sales_store(_file_mtime(DEFAULT_SALES_STORE))
    price_hist = cached_price_history(_file_mtime(DEFAULT_PRICE_HISTORY))
    df_all_raw = cached_enrich_sales(_file_mtime(DEFAULT_SALES_STORE), vmap_path_used, _file_mtime(Path(vmap_path_used)), _file_mtime(DEFAULT_PRICE_HISTORY))

    def _apply_overview_filters(df: pd.DataFrame) -> pd.DataFrame:
        """(Disabled) Overview drill-down filters removed."""
        return df


    # df_all is the filtered view used by most tabs; df_all_raw is always the full dataset
    df_all = df_all_raw


    def _compute_wow_insights(df_full: pd.DataFrame):
        """Return (executive_takeaway, drivers_df, opportunities_df) for the latest week."""
        if df_full is None or df_full.empty:
            return ("", pd.DataFrame(), pd.DataFrame())

        cur, prev, cur_start, cur_end = _get_current_and_prev_week(df_full)
        if cur is None or cur.empty:
            return ("", pd.DataFrame(), pd.DataFrame())

        prev = prev if (prev is not None) else pd.DataFrame(columns=cur.columns)

        cur_s = cur.groupby("SKU", as_index=False)[["Units","Sales"]].sum()
        prev_s = prev.groupby("SKU", as_index=False)[["Units","Sales"]].sum() if (prev is not None and not prev.empty) else pd.DataFrame(columns=["SKU","Units","Sales"])

        wow = cur_s.merge(prev_s, on="SKU", how="left", suffixes=("_cur","_prev"))
        wow["Sales_prev"] = wow["Sales_prev"].fillna(0.0)
        wow["Units_prev"] = wow["Units_prev"].fillna(0.0)
        wow["WoW Sales"] = wow["Sales_cur"] - wow["Sales_prev"]
        wow["WoW Units"] = wow["Units_cur"] - wow["Units_prev"]

        drivers = wow.assign(_abs=wow["WoW Sales"].abs()).sort_values("_abs", ascending=False).head(3).drop(columns=["_abs"])
        drivers_df = drivers.rename(columns={
            "SKU":"SKU",
            "Sales_cur":"Sales (This Week)",
            "Units_cur":"Units (This Week)",
            "WoW Sales":"WoW Sales ($)",
            "WoW Units":"WoW Units"
        })[["SKU","WoW Sales ($)","Sales (This Week)","Units (This Week)","WoW Units"]]

        opportunities = wow[wow["WoW Sales"] > 0].sort_values("WoW Sales", ascending=False).head(10)
        opp_df = opportunities.rename(columns={
            "SKU":"SKU",
            "Sales_cur":"Sales (This Week)",
            "Units_cur":"Units (This Week)",
            "WoW Sales":"WoW Sales ($)",
            "WoW Units":"WoW Units"
        })[["SKU","WoW Sales ($)","Sales (This Week)","Units (This Week)","WoW Units"]]

        total_sales_cur = float(cur["Sales"].sum())
        total_sales_prev = float(prev["Sales"].sum()) if (prev is not None and not prev.empty) else 0.0
        wow_sales_total = total_sales_cur - total_sales_prev

        # Build a one-line takeaway
        def _money(v):
            try:
                v = float(v)
            except Exception:
                return "—"
            return f"-${abs(v):,.0f}" if v < 0 else f"${v:,.0f}"

        top_neg = wow.sort_values("WoW Sales").head(1)
        top_pos = wow.sort_values("WoW Sales", ascending=False).head(1)
        neg_txt = ""
        pos_txt = ""
        if not top_neg.empty and float(top_neg["WoW Sales"].iloc[0]) < 0:
            neg_txt = f"biggest headwind was SKU {str(top_neg['SKU'].iloc[0])} ({_money(top_neg['WoW Sales'].iloc[0])})."
        if not top_pos.empty and float(top_pos["WoW Sales"].iloc[0]) > 0:
            pos_txt = f"Top offset/opportunity was SKU {str(top_pos['SKU'].iloc[0])} (+{_money(top_pos['WoW Sales'].iloc[0]).lstrip('-')})."

        direction = "up" if wow_sales_total > 0 else ("down" if wow_sales_total < 0 else "flat")
        takeaway = f"WoW sales were {direction} {_money(wow_sales_total)}. " + (neg_txt + " " if neg_txt else "") + (pos_txt if pos_txt else "")
        takeaway = takeaway.strip()

        return (takeaway, drivers_df, opp_df)
    # ------------
    # Sidebar: PDF Export (Weekly Summary)
    # ----------------------
    st.sidebar.markdown("---")
    with st.sidebar.expander("📄 PDF Export", expanded=False):
        st.caption("Generate the Weekly Summary PDF without navigating to another tab.")

        use_exec_scope = st.checkbox("Export using Executive Summary selection (Scope/Vendor/Retailer)", value=False, key="sb_pdf_use_exec_scope")
        if use_exec_scope:
            _sc = st.session_state.get("ex_scope", "All")
            if _sc == "Vendor":
                st.caption(f"Scope: Vendor — {st.session_state.get('ex_pick_v', '(not selected)')}")
            elif _sc == "Retailer":
                st.caption(f"Scope: Retailer — {st.session_state.get('ex_pick_r', '(not selected)')}")
            else:
                st.caption("Scope: All")

        if "sb_weekly_pdf_bytes" not in st.session_state:
            st.session_state.sb_weekly_pdf_bytes = None
            st.session_state.sb_weekly_pdf_name = None

        if st.button("Build Weekly Summary PDF", use_container_width=True, key="sb_btn_build_weekly_pdf"):
            try:
                if df_all is None or df_all.empty:
                    st.warning("No sales data loaded yet.")
                else:
                    # Use filtered dataset for export
                    df_src = df_all.copy() if ('df_all' in globals() and df_all is not None) else df_all_raw
                    if use_exec_scope and df_src is not None and not df_src.empty:
                        _sc = st.session_state.get("ex_scope", "All")
                        if _sc == "Vendor":
                            _pick = st.session_state.get("ex_pick_v", None)
                            if _pick:
                                df_src = df_src[df_src["Vendor"] == _pick].copy()
                        elif _sc == "Retailer":
                            _pick = st.session_state.get("ex_pick_r", None)
                            if _pick:
                                df_src = df_src[df_src["Retailer"] == _pick].copy()

                    cur, prev, cur_start, cur_end = _get_current_and_prev_week(df_src)
                    if cur is None or cur.empty:
                        st.warning("No current-week data found.")
                    else:
                        # Core tables
                        cur_r = cur.groupby("Retailer", as_index=False)[["Units","Sales"]].sum()
                        prev_r = prev.groupby("Retailer", as_index=False)[["Units","Sales"]].sum() if (prev is not None and not prev.empty) else pd.DataFrame(columns=["Retailer","Units","Sales"])
                        cur_v = cur.groupby("Vendor", as_index=False)[["Units","Sales"]].sum()
                        prev_v = prev.groupby("Vendor", as_index=False)[["Units","Sales"]].sum() if (prev is not None and not prev.empty) else pd.DataFrame(columns=["Vendor","Units","Sales"])
                        cur_s = cur.groupby("SKU", as_index=False)[["Units","Sales"]].sum()
                        prev_s = prev.groupby("SKU", as_index=False)[["Units","Sales"]].sum() if (prev is not None and not prev.empty) else pd.DataFrame(columns=["SKU","Units","Sales"])

                        def _add_wow_sales(cur_df: pd.DataFrame, prev_df: pd.DataFrame, key: str) -> pd.DataFrame:
                            cur2 = cur_df.copy()
                            prev2 = prev_df.copy() if prev_df is not None else pd.DataFrame(columns=[key, "Sales", "Units"])
                            if key not in cur2.columns:
                                return cur2
                            if "Sales" not in cur2.columns: cur2["Sales"] = 0.0
                            if "Units" not in cur2.columns: cur2["Units"] = 0.0
                            if key not in prev2.columns: prev2[key] = []
                            if "Sales" not in prev2.columns: prev2["Sales"] = 0.0
                            prev2 = prev2[[key, "Sales"]].rename(columns={"Sales":"Sales_prev"})
                            cur2 = cur2.merge(prev2, on=key, how="left")
                            cur2["Sales_prev"] = cur2["Sales_prev"].fillna(0.0)
                            cur2["WoW $ Diff"] = cur2["Sales"] - cur2["Sales_prev"]
                            cur2["WoW $ %"] = np.where(cur2["Sales_prev"] > 0, cur2["WoW $ Diff"] / cur2["Sales_prev"], np.nan)
                            return cur2

                        top_r = _add_wow_sales(cur_r, prev_r, "Retailer").sort_values("Sales", ascending=False).head(10)
                        top_v = _add_wow_sales(cur_v, prev_v, "Vendor").sort_values("Sales", ascending=False).head(10)
                        top_s = _add_wow_sales(cur_s, prev_s, "SKU").sort_values("Sales", ascending=False).head(10)

                        # Biggest movers (SKU)
                        movers = cur_s.merge(prev_s, on="SKU", how="left", suffixes=("_cur","_prev"))
                        movers["Sales_prev"] = movers["Sales_prev"].fillna(0.0)
                        movers["Units_prev"] = movers["Units_prev"].fillna(0.0)
                        movers["WoW Sales"] = movers["Sales_cur"] - movers["Sales_prev"]
                        movers["WoW Units"] = movers["Units_cur"] - movers["Units_prev"]
                        movers["_abs_wow_sales"] = movers["WoW Sales"].abs()
                        movers = movers.sort_values("_abs_wow_sales", ascending=False).head(25).drop(columns=["_abs_wow_sales"])

                        movers_pdf = movers.rename(columns={
                            "Sales_cur":"Sales (This Week)",
                            "Sales_prev":"Sales (Prev Week)",
                            "Units_cur":"Units (This Week)",
                            "Units_prev":"Units (Prev Week)",
                        }).copy()

                        for c in ["Sales (This Week)","Sales (Prev Week)","WoW Sales"]:
                            if c in movers_pdf.columns: movers_pdf[c] = movers_pdf[c].map(_fmt_pdf_money)
                        for c in ["Units (This Week)","Units (Prev Week)","WoW Units"]:
                            if c in movers_pdf.columns: movers_pdf[c] = movers_pdf[c].map(_fmt_pdf_int)

                        # Momentum table (top 15)
                        try:
                            mom_pdf_export = compute_momentum_table(df_all, window=12).head(15).copy()
                        except Exception:
                            mom_pdf_export = pd.DataFrame(columns=["SKU","Momentum","Up Weeks","Down Weeks","Units_Last","Sales_Last"])
                        if mom_pdf_export is None or mom_pdf_export.empty:
                            mom_pdf_export = pd.DataFrame(columns=["SKU","Momentum","Up Weeks","Down Weeks","Units_Last","Sales_Last"])
                        else:
                            if "Units_Last" in mom_pdf_export.columns: mom_pdf_export["Units_Last"] = mom_pdf_export["Units_Last"].map(_fmt_pdf_int)
                            if "Sales_Last" in mom_pdf_export.columns: mom_pdf_export["Sales_Last"] = mom_pdf_export["Sales_Last"].map(_fmt_pdf_money)

                        # KPI dict + highlights
                        total_sales_cur = float(cur_s["Sales"].sum()) if "Sales" in cur_s.columns else 0.0
                        total_units_cur = float(cur_s["Units"].sum()) if "Units" in cur_s.columns else 0.0
                        total_sales_prev = float(prev_s["Sales"].sum()) if (prev is not None and not prev.empty and "Sales" in prev_s.columns) else 0.0
                        total_units_prev = float(prev_s["Units"].sum()) if (prev is not None and not prev.empty and "Units" in prev_s.columns) else 0.0

                        wow_sales = total_sales_cur - total_sales_prev
                        wow_units = total_units_cur - total_units_prev
                        wow_sales_pct = (wow_sales / total_sales_prev) if total_sales_prev > 0 else None
                        wow_units_pct = (wow_units / total_units_prev) if total_units_prev > 0 else None

                        kpi_dict = {
                            "Sales": _fmt_pdf_money(total_sales_cur),
                            "Units": _fmt_pdf_int(total_units_cur),
                            "WoW Sales": _fmt_pdf_money(wow_sales) + (f" ({wow_sales_pct*100:,.1f}%)" if wow_sales_pct is not None else ""),
                            "WoW Units": _fmt_pdf_int(wow_units) + (f" ({wow_units_pct*100:,.1f}%)" if wow_units_pct is not None else ""),
                        }

                        highlights = [
                            f"Week {pd.to_datetime(cur_start).date()} to {pd.to_datetime(cur_end).date()} total sales {_fmt_pdf_money(total_sales_cur)} on {_fmt_pdf_int(total_units_cur)} units.",
                            f"WoW sales change: {_fmt_pdf_money(wow_sales)}.",
                            f"WoW units change: {_fmt_pdf_int(wow_units)}.",
                        ]

                        # Format top tables for PDF
                        def _format_top(df_in: pd.DataFrame, key_col: str) -> pd.DataFrame:
                            if df_in is None or df_in.empty:
                                return pd.DataFrame(columns=[key_col, "Units", "Sales", "WoW $ Diff"])
                            out = df_in.copy()
                            if "Units" in out.columns: out["Units"] = out["Units"].map(_fmt_pdf_int)
                            if "Sales" in out.columns: out["Sales"] = out["Sales"].map(_fmt_pdf_money)
                            if "WoW $ Diff" in out.columns: out["WoW $ Diff"] = out["WoW $ Diff"].map(_fmt_pdf_money)
                            cols = [c for c in [key_col, "Units", "Sales", "WoW $ Diff"] if c in out.columns]
                            return out[cols]

                        top_r_pdf = _format_top(top_r, "Retailer")
                        top_v_pdf = _format_top(top_v, "Vendor")
                        top_s_pdf = _format_top(top_s, "SKU")

                        # Declining vendors / retailers (WoW Sales < 0 based on current vs previous week)
                        def _declines_simple(cur_df: pd.DataFrame, prev_df: pd.DataFrame, key: str) -> pd.DataFrame:
                            cur2 = cur_df.copy()
                            prev2 = prev_df.copy() if prev_df is not None else pd.DataFrame(columns=[key, "Units", "Sales"])
                            cur2 = cur2.rename(columns={"Units": "Units_cur", "Sales": "Sales_cur"})
                            prev2 = prev2.rename(columns={"Units": "Units_prev", "Sales": "Sales_prev"})
                            if key not in prev2.columns:
                                prev2[key] = []
                            out = cur2.merge(prev2[[key, "Units_prev", "Sales_prev"]], on=key, how="left")
                            out["Units_prev"] = out["Units_prev"].fillna(0.0)
                            out["Sales_prev"] = out["Sales_prev"].fillna(0.0)
                            out["WoW Sales"] = out["Sales_cur"] - out["Sales_prev"]
                            out["WoW Units"] = out["Units_cur"] - out["Units_prev"]
                            out["%∆Sales"] = np.where(out["Sales_prev"] > 0, out["WoW Sales"] / out["Sales_prev"], np.nan)
                            out = out[(out["WoW Sales"] < 0) & (out["Sales_prev"] > 0)].sort_values("WoW Sales").head(10)
                            # order columns
                            cols = [key, "Units_cur", "Sales_cur", "Units_prev", "Sales_prev", "WoW Sales", "WoW Units", "%∆Sales"]
                            return out[cols]

                        vendor_decl_pdf = _declines_simple(cur_v, prev_v, "Vendor")
                        retailer_decl_pdf = _declines_simple(cur_r, prev_r, "Retailer")

                        # Format for PDF
                        def _fmt_pct(x):
                            try:
                                return f"{float(x):+.0%}".replace("+", "")
                            except Exception:
                                return ""
                        for _d in [vendor_decl_pdf, retailer_decl_pdf]:
                            if _d is not None and not _d.empty:
                                if "Units_cur" in _d.columns: _d["Units_cur"] = _d["Units_cur"].map(_fmt_pdf_int)
                                if "Units_prev" in _d.columns: _d["Units_prev"] = _d["Units_prev"].map(_fmt_pdf_int)
                                if "Sales_cur" in _d.columns: _d["Sales_cur"] = _d["Sales_cur"].map(_fmt_pdf_money)
                                if "Sales_prev" in _d.columns: _d["Sales_prev"] = _d["Sales_prev"].map(_fmt_pdf_money)
                                if "WoW Sales" in _d.columns: _d["WoW Sales"] = _d["WoW Sales"].map(_fmt_pdf_money)
                                if "WoW Units" in _d.columns: _d["WoW Units"] = _d["WoW Units"].map(_fmt_pdf_int)
                                if "%∆Sales" in _d.columns: _d["%∆Sales"] = _d["%∆Sales"].map(_fmt_pct)

                        executive_takeaway, drivers_df, opportunities_df = _compute_wow_insights(df_all_raw)

                        pdf_bytes = make_weekly_summary_pdf_bytes(
                            "Weekly Summary",
                            highlights,
                            kpi_dict,
                            top_r_pdf,
                            top_v_pdf,
                            top_s_pdf,
                            movers_pdf,
                            vendor_decl_pdf,
                            retailer_decl_pdf,
                            mom_pdf_export,
                            df_all_raw,
                            logo_path=LOGO_PATH if "LOGO_PATH" in globals() else None,
                            executive_takeaway=executive_takeaway,
                            drivers_df=drivers_df,
                            opportunities_df=opportunities_df,
                        )

                        st.session_state.sb_weekly_pdf_bytes = pdf_bytes
                        st.session_state.sb_weekly_pdf_name = f"Weekly_Summary_{pd.to_datetime(cur_end).date()}.pdf"
                        st.success("PDF is ready below.")

            except Exception as e:
                st.error(f"PDF build failed: {e}")

        if st.session_state.sb_weekly_pdf_bytes:
            st.download_button(
                "⬇️ Download Weekly Summary (PDF)",
                data=st.session_state.sb_weekly_pdf_bytes,
                file_name=st.session_state.sb_weekly_pdf_name or "Weekly_Summary.pdf",
                mime="application/pdf",
                use_container_width=True,
                key="sb_dl_weekly_summary_pdf",
            )

    # KPIs across top (always current calendar year)
    df_kpi = df_all.copy()
    df_kpi["StartDate"] = pd.to_datetime(df_kpi["StartDate"], errors="coerce")
    df_kpi = df_kpi[df_kpi["StartDate"].dt.year == int(this_year)].copy()

    # Apply view-year filter for all reporting tabs
    df = df_all.copy()
    df["StartDate"] = pd.to_datetime(df["StartDate"], errors="coerce")
    df = df[df["StartDate"].dt.year == int(view_year)].copy()

    # KPIs across top
    m_all = wow_mom_metrics(df_kpi)

    st.markdown("## 📊 Overview (All Retailers)")

    # Build richer top metrics: YTD + This Week vs Prev Week + This Month vs Prev Month + Last 8 Weeks vs Prev 8 Weeks
    d_top = df_kpi.copy()
    d_top["StartDate"] = pd.to_datetime(d_top["StartDate"], errors="coerce")
    d_top["EndDate"] = pd.to_datetime(d_top["EndDate"], errors="coerce")

    # YTD
    ytd_units = float(pd.to_numeric(d_top.get("Units"), errors="coerce").fillna(0).sum())
    ytd_sales = float(pd.to_numeric(d_top.get("Sales"), errors="coerce").fillna(0).sum())

    # Weekly totals
    weeks = sorted(d_top["StartDate"].dropna().dt.date.unique().tolist())
    cur_week_total = 0.0
    prev_week_total = 0.0
    if len(weeks) >= 1:
        cur_w = weeks[-1]
        cur_week_total = float(d_top.loc[d_top["StartDate"].dt.date == cur_w, "Sales"].fillna(0).sum())
    if len(weeks) >= 2:
        prev_w = weeks[-2]
        prev_week_total = float(d_top.loc[d_top["StartDate"].dt.date == prev_w, "Sales"].fillna(0).sum())
    wow_diff_sales = cur_week_total - prev_week_total

    # Monthly totals
    d_top["MonthP"] = d_top["StartDate"].dt.to_period("M")
    months = sorted(d_top["MonthP"].dropna().unique().tolist())
    cur_month_total = 0.0
    prev_month_total = 0.0
    if len(months) >= 1:
        cm = months[-1]
        cur_month_total = float(d_top.loc[d_top["MonthP"] == cm, "Sales"].fillna(0).sum())
    if len(months) >= 2:
        pm = months[-2]
        prev_month_total = float(d_top.loc[d_top["MonthP"] == pm, "Sales"].fillna(0).sum())
    mom_diff_sales = cur_month_total - prev_month_total

    # Last 8 weeks vs prior 8 weeks (by EndDate)
    d8 = d_top.dropna(subset=["EndDate"]).copy()
    d8 = d8.sort_values("EndDate")
    end_weeks = sorted(d8["EndDate"].dt.date.unique().tolist())
    last8_total = 0.0
    prev8_total = 0.0
    if len(end_weeks) >= 1:
        last8 = end_weeks[-8:]
        prev8 = end_weeks[-16:-8] if len(end_weeks) >= 16 else []
        last8_total = float(d8.loc[d8["EndDate"].dt.date.isin(last8), "Sales"].fillna(0).sum())
        prev8_total = float(d8.loc[d8["EndDate"].dt.date.isin(prev8), "Sales"].fillna(0).sum()) if prev8 else 0.0
    last8_diff = last8_total - prev8_total

    def _delta_badge(val, is_money=True):
        if val is None:
            return ""
        try:
            v = float(val)
        except Exception:
            return ""
        if is_money:
            s = f"-${abs(v):,.0f}" if v < 0 else f"${v:,.0f}"
        else:
            s = f"{int(v):,}"
        color = "#1f8f4c" if v > 0 else ("#c43d3d" if v < 0 else "#777")
        # small circle-style badge
        return f"<div style='margin-top:4px; font-size:12px; color:{color};'><span style='display:inline-block; padding:2px 8px; border-radius:999px; border:1px solid {color};'>{s}</span></div>"

    def _card(title, value, delta_html=""):
        return f"""<div class='kpi-card'>
            <div class='kpi-label'>{title}</div>
            <div class='kpi-value'>{value}</div>
            {delta_html}
        </div>"""



    # --- Top header KPIs (native Streamlit; theme-aware) ---
    def _fmt_delta_money(v) -> str:
        try:
            x = float(v)
        except Exception:
            return ""
        if x > 0:
            return f"+${x:,.2f}"
        if x < 0:
            return f"-${abs(x):,.2f}"
        return "$0.00"


    # Overview header (auto dark/light)
    st.markdown("""
    <style>
.overview-title { color: var(--text-color); }
</style>
    """, unsafe_allow_html=True)

    st.markdown(
        "<div class='overview-title' style='font-size:22px; font-weight:700; margin-top:8px;'>"
        "Overview (All Retailers)"
        "</div>",
        unsafe_allow_html=True
    )

    c1, c2, c3, c4 = st.columns([1.1, 1.0, 1.0, 1.0])

    with c1:
        st.metric("Total Units (YTD)", fmt_int(ytd_units))
        st.metric("Total Sales (YTD)", fmt_currency(ytd_sales))

    with c2:
        st.metric("This Week Sales", fmt_currency(cur_week_total), delta=_fmt_delta_money(wow_diff_sales))

    with c3:
        st.metric("This Month Sales", fmt_currency(cur_month_total), delta=_fmt_delta_money(mom_diff_sales))

    with c4:
        st.metric("Last 8 Weeks Sales", fmt_currency(last8_total), delta=_fmt_delta_money(last8_diff))

    st.divider()



    # -------------------------
    # Reporting helpers
    # -------------------------
    def week_labels(df_in: pd.DataFrame) -> list[str]:
        if df_in is None or df_in.empty:
            return []
        w = sorted(pd.to_datetime(df_in["StartDate"], errors="coerce").dropna().dt.date.unique().tolist())
        return [pd.Timestamp(x).strftime("%m-%d") for x in w]

    def add_week_col(d: pd.DataFrame) -> pd.DataFrame:
        d = d.copy()
        d["StartDate"] = pd.to_datetime(d["StartDate"], errors="coerce")
        d["Week"] = d["StartDate"].dt.date
        return d

    def nonzero_mean_rowwise(frame: pd.DataFrame) -> pd.Series:
        """Mean across columns, ignoring zeros (treat zeros as missing)."""
        return frame.replace(0, np.nan).mean(axis=1)

    def last_n_weeks(df_in: pd.DataFrame, n: int):
        if df_in is None or df_in.empty:
            return []
        w = sorted(pd.to_datetime(df_in["StartDate"], errors="coerce").dropna().dt.date.unique().tolist())
        return w[-n:] if len(w) >= n else w

    def safe_div(a, b):
        try:
            if b == 0 or pd.isna(b):
                return np.nan
            return a / b
        except Exception:
            return np.nan

    def to_pdf_bytes(title: str, sections: list[tuple[str, list[str]]]) -> bytes:
        """
        Build a simple PDF summary.
        sections: list of (heading, lines[])
        """
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.pdfgen import canvas
            from reportlab.lib.units import inch
        except Exception:
            return b""

        import io
        buf = io.BytesIO()
        c = canvas.Canvas(buf, pagesize=letter)
        width, height = letter
        x = 0.75 * inch
        y = height - 0.75 * inch

        c.setFont("Helvetica-Bold", 16)
        c.drawString(x, y, title)
        y -= 0.35 * inch

        for heading, lines in sections:
            if y < 1.0 * inch:
                c.showPage()
                y = height - 0.75 * inch
            c.setFont("Helvetica-Bold", 11)
            c.drawString(x, y, heading)
            y -= 0.22 * inch
            c.setFont("Helvetica", 10)
            for ln in lines:
                if y < 1.0 * inch:
                    c.showPage()
                    y = height - 0.75 * inch
                    c.setFont("Helvetica", 10)
                c.drawString(x, y, str(ln)[:120])
                y -= 0.18 * inch
            y -= 0.10 * inch

        c.save()
        return buf.getvalue()


    def render_comparison_retailer_vendor():
            st.subheader("Comparison")
            st.session_state["cmp_ctx"] = {}

            if df.empty:
                st.info("No sales data yet.")
                return

            d = df_all.copy()
            d["StartDate"] = pd.to_datetime(d["StartDate"], errors="coerce")
            d = d[d["StartDate"].notna()].copy()

            # Build month + year options across ALL years in the store
            d["MonthP"] = d["StartDate"].dt.to_period("M")
            months = sorted(d["MonthP"].unique().tolist())
            month_labels = [m.to_timestamp().strftime("%B %Y") for m in months]
            label_to_period = dict(zip(month_labels, months))

            d["Year"] = d["StartDate"].dt.year.astype(int)
            years = sorted(d["Year"].dropna().unique().tolist())

            mode = st.radio(
                "Compare type",
                options=["A vs B (Months)", "A vs B (Years)", "Multi-year (high/low highlight)", "Multi-month across years"],
                index=0,
                horizontal=True,
                key="cmp_mode_v2"
            )

            c1, c2, c3 = st.columns([2, 2, 1])
            with c3:
                by = st.selectbox("Compare by", ["Retailer", "Vendor"], key="cmp_by_v2")

            # Optional limiter list
            if by == "Retailer":
                options = sorted(d["Retailer"].dropna().unique().tolist())
            else:
                options = sorted([v for v in d["Vendor"].dropna().unique().tolist() if str(v).strip()])

            sel = st.multiselect(f"Limit to {by}(s) (optional)", options=options, key="cmp_limit_v2")

            # -------------------------
            # Helper: render A vs B table
            # -------------------------
            def _render_a_vs_b(da: pd.DataFrame, db: pd.DataFrame, label_a: str, label_b: str):
                ga = da.groupby(by, as_index=False).agg(Units_A=("Units","sum"), Sales_A=("Sales","sum"))
                gb = db.groupby(by, as_index=False).agg(Units_B=("Units","sum"), Sales_B=("Sales","sum"))

                out = ga.merge(gb, on=by, how="outer").fillna(0.0)
                out["Units_Diff"] = out["Units_A"] - out["Units_B"]
                out["Sales_Diff"] = out["Sales_A"] - out["Sales_B"]
                out["Units_%"] = out["Units_Diff"] / out["Units_B"].replace(0, np.nan)
                out["Sales_%"] = out["Sales_Diff"] / out["Sales_B"].replace(0, np.nan)

                total = {
                    by: "TOTAL",
                    "Units_A": out["Units_A"].sum(),
                    "Sales_A": out["Sales_A"].sum(),
                    "Units_B": out["Units_B"].sum(),
                    "Sales_B": out["Sales_B"].sum(),
                }
                total["Units_Diff"] = total["Units_A"] - total["Units_B"]
                total["Sales_Diff"] = total["Sales_A"] - total["Sales_B"]
                total["Units_%"] = total["Units_Diff"] / total["Units_B"] if total["Units_B"] else np.nan
                total["Sales_%"] = total["Sales_Diff"] / total["Sales_B"] if total["Sales_B"] else np.nan

                out = pd.concat([out, pd.DataFrame([total])], ignore_index=True)

                disp = out[[by,"Units_A","Sales_A","Units_B","Sales_B","Units_Diff","Units_%","Sales_Diff","Sales_%"]].copy()
                disp = disp.rename(columns={
                    "Units_A": f"Units ({label_a})",
                    "Sales_A": f"Sales ({label_a})",
                    "Units_B": f"Units ({label_b})",
                    "Sales_B": f"Sales ({label_b})",
                })

                sty = disp.style.format({
                    f"Units ({label_a})": fmt_int,
                    f"Units ({label_b})": fmt_int,
                    "Units_Diff": fmt_int,
                    "Units_%": lambda v: f"{v*100:.1f}%" if pd.notna(v) else "—",
                    f"Sales ({label_a})": fmt_currency,
                    f"Sales ({label_b})": fmt_currency,
                    "Sales_Diff": fmt_currency,
                    "Sales_%": lambda v: f"{v*100:.1f}%" if pd.notna(v) else "—",
                }).applymap(lambda v: f"color: {_color(v)};", subset=["Units_Diff","Sales_Diff"])

                st.dataframe(sty, use_container_width=True, hide_index=True)

            # -------------------------
            # Mode 1: A vs B (Months)
            # -------------------------
            if mode == "A vs B (Months)":
                with c1:
                    a_pick = st.multiselect(
                        "Selection A (one or more months)",
                        options=month_labels,
                        default=month_labels[-1:] if month_labels else [],
                        key="cmp_a_months_v2"
                    )
                with c2:
                    b_pick = st.multiselect(
                        "Selection B (one or more months)",
                        options=month_labels,
                        default=month_labels[-2:-1] if len(month_labels) >= 2 else [],
                        key="cmp_b_months_v2"
                    )

                a_periods = [label_to_period[x] for x in a_pick if x in label_to_period]
                b_periods = [label_to_period[x] for x in b_pick if x in label_to_period]

                # Store for Top SKU movers table
                st.session_state["movers_a_periods"] = [str(p) for p in a_periods]
                st.session_state["movers_b_periods"] = [str(p) for p in b_periods]

                if not a_periods or not b_periods:
                    st.info("Pick at least one month in Selection A and Selection B.")
                    return

                da = d[d["MonthP"].isin(a_periods)]
                db = d[d["MonthP"].isin(b_periods)]

                if sel:
                    da = da[da[by].isin(sel)]
                    db = db[db[by].isin(sel)]

                label_a = " + ".join(a_pick) if a_pick else "A"
                label_b = " + ".join(b_pick) if b_pick else "B"
                _render_a_vs_b(da, db, label_a, label_b)
                st.session_state["cmp_ctx"] = {"a": da.copy(), "b": db.copy(), "label_a": label_a, "label_b": label_b, "value_col": "Sales"}
                return

            # -------------------------
            # Mode 2: A vs B (Years)
            # Example: compare (2023+2024) vs (2024+2025)
            # -------------------------
            if mode == "A vs B (Years)":
                with c1:
                    years_a = st.multiselect(
                        "Selection A (one or more years)",
                        options=years,
                        default=years[-2:-1] if len(years) >= 2 else years,
                        key="cmp_years_a_v2"
                    )
                with c2:
                    years_b = st.multiselect(
                        "Selection B (one or more years)",
                        options=years,
                        default=years[-1:] if years else [],
                        key="cmp_years_b_v2"
                    )

                if not years_a or not years_b:
                    st.info("Pick at least one year in Selection A and Selection B.")
                    return

                da = d[d["Year"].isin([int(y) for y in years_a])]
                db = d[d["Year"].isin([int(y) for y in years_b])]

                # Store for Top SKU movers table (months within those years)
                try:
                    st.session_state["movers_a_periods"] = [str(p) for p in sorted(da["MonthP"].unique().tolist())] if "MonthP" in da.columns else []
                    st.session_state["movers_b_periods"] = [str(p) for p in sorted(db["MonthP"].unique().tolist())] if "MonthP" in db.columns else []
                except Exception:
                    st.session_state["movers_a_periods"] = []
                    st.session_state["movers_b_periods"] = []

                if sel:
                    da = da[da[by].isin(sel)]
                    db = db[db[by].isin(sel)]

                label_a = " + ".join([str(y) for y in years_a])
                label_b = " + ".join([str(y) for y in years_b])
                _render_a_vs_b(da, db, label_a, label_b)
                st.session_state["cmp_ctx"] = {"a": da.copy(), "b": db.copy(), "label_a": label_a, "label_b": label_b, "value_col": "Sales"}
                return


            # -------------------------
            # Mode 2b: Multi-month across years (pick Month+Year periods)
            # -------------------------
            if mode == "Multi-month across years":
                month_names = ["January","February","March","April","May","June","July","August","September","October","November","December"]
                period_options = []
                for y in years:
                    for mname in month_names:
                        period_options.append(f"{mname} {y}")

                with c1:
                    periods_pick = st.multiselect(
                        "Month + Year periods",
                        options=period_options,
                        default=[period_options[-1]] if period_options else [],
                        key="cmp_mm_periods",
                    )
                with c2:
                    metric = st.selectbox(
                        "Highlight based on",
                        options=["Sales", "Units"],
                        index=0,
                        key="cmp_mm_metric",
                    )
                with c3:
                    topn = st.selectbox("Show", options=[25, 50, 100, 250], index=1, key="cmp_mm_topn")

                if not periods_pick:
                    st.info("Pick at least one Month + Year period.")
                    return

                month_to_num = {m:i+1 for i,m in enumerate(month_names)}
                pairs = []
                for p in periods_pick:
                    try:
                        parts = p.split(" ")
                        mname = " ".join(parts[:-1])
                        yy = int(parts[-1])
                        mn = month_to_num.get(mname, None)
                        if mn is not None:
                            pairs.append((yy, mn, p))
                    except Exception:
                        continue
                if not pairs:
                    st.info("No valid periods selected.")
                    return

                dd = d.copy()
                mask = False
                for (yy, mn, _) in pairs:
                    mask = mask | ((dd["Year"] == int(yy)) & (dd["StartDate"].dt.month == int(mn)))
                dd = dd[mask].copy()
                if sel:
                    dd = dd[dd[by].isin(sel)]

                if dd.empty:
                    st.info("No data found for those periods with current filters.")
                    return

                pieces = []
                for (yy, mn, lab_full) in pairs:
                    lab = lab_full.replace("January","Jan").replace("February","Feb").replace("March","Mar").replace("April","Apr").replace("June","Jun").replace("July","Jul").replace("August","Aug").replace("September","Sep").replace("October","Oct").replace("November","Nov").replace("December","Dec")
                    dyy = dd[(dd["Year"] == int(yy)) & (dd["StartDate"].dt.month == int(mn))].copy()
                    gy = dyy.groupby(by, as_index=False).agg(**{
                        f"Units_{lab}": ("Units", "sum"),
                        f"Sales_{lab}": ("Sales", "sum"),
                    })
                    pieces.append((lab, gy))

                out = pieces[0][1]
                for lab, p in pieces[1:]:
                    out = out.merge(p, on=by, how="outer")
                out = out.fillna(0.0)

                total = {by: "TOTAL"}
                for c in out.columns:
                    if c == by:
                        continue
                    total[c] = float(out[c].sum()) if pd.api.types.is_numeric_dtype(out[c]) else ""
                out = pd.concat([out, pd.DataFrame([total])], ignore_index=True)

                cols = [by]
                for lab, _ in pieces:
                    cols += [f"Units_{lab}", f"Sales_{lab}"]
                disp = out[cols].copy()

                metric_cols = [c for c in disp.columns if c.startswith(metric + "_")]
                if metric_cols:
                    disp = disp.sort_values(metric_cols[-1], ascending=False).head(int(topn)).copy()

                spark_chars = ["▁","▂","▃","▄","▅","▆","▇","█"]
                def _spark(vals):
                    vals = [float(v) if v is not None and not pd.isna(v) else np.nan for v in vals]
                    if len(vals) == 0 or all(pd.isna(v) for v in vals):
                        return ""
                    vmin = np.nanmin(vals); vmax = np.nanmax(vals)
                    if np.isnan(vmin) or np.isnan(vmax) or np.isclose(vmin, vmax):
                        return "▁" * len(vals)
                    out_s=[]
                    for v in vals:
                        if pd.isna(v):
                            out_s.append(" "); continue
                        t=(v-vmin)/(vmax-vmin)
                        idx=int(round(t*(len(spark_chars)-1)))
                        idx=max(0,min(len(spark_chars)-1,idx))
                        out_s.append(spark_chars[idx])
                    return "".join(out_s)

                def _cagr(a,b,periods):
                    try:
                        a=float(a); b=float(b)
                    except Exception:
                        return np.nan
                    if a <= 0 or b <= 0 or periods <= 0:
                        return np.nan
                    return (b/a)**(1.0/periods)-1.0

                if metric_cols:
                    series_vals = disp[metric_cols].copy()
                    disp["Spark"] = series_vals.apply(lambda r: _spark(r.tolist()), axis=1)
                    periods_n = max(1, len(metric_cols)-1)
                    disp["CAGR %"] = series_vals.apply(lambda r: _cagr(r[metric_cols[0]], r[metric_cols[-1]], periods_n), axis=1)

                def _hl(row):
                    styles=[""]*len(row)
                    if not metric_cols:
                        return styles
                    vals=[]
                    for c in metric_cols:
                        try:
                            vals.append(float(row[c]))
                        except Exception:
                            vals.append(np.nan)
                    if len(vals)==0 or all(pd.isna(v) for v in vals):
                        return styles
                    vmax=np.nanmax(vals); vmin=np.nanmin(vals)
                    for i,c in enumerate(row.index):
                        if c in metric_cols:
                            v=float(row[c]) if pd.notna(row[c]) else np.nan
                            if pd.notna(v) and np.isclose(v,vmax):
                                styles[i]="background-color: rgba(0, 200, 0, 0.18); font-weight: 600;"
                            elif pd.notna(v) and np.isclose(v,vmin):
                                styles[i]="background-color: rgba(220, 0, 0, 0.14);"
                    return styles

                fmt_map = {c: fmt_int for c in disp.columns if c.startswith("Units_")}
                fmt_map.update({c: fmt_currency for c in disp.columns if c.startswith("Sales_")})
                if "CAGR %" in disp.columns:
                    fmt_map["CAGR %"] = lambda v: f"{v*100:.1f}%" if pd.notna(v) else "—"

                st.caption("Multi-month across years (selected Month+Year periods)")
                sty = disp.style.format(fmt_map)
                if metric_cols:
                    sty = sty.apply(_hl, axis=1)
                st.dataframe(sty, use_container_width=True, hide_index=True)

                first = pairs[0]; last = pairs[-1]
                a_ctx = dd[(dd["Year"] == int(first[0])) & (dd["StartDate"].dt.month == int(first[1]))].copy()
                b_ctx = dd[(dd["Year"] == int(last[0])) & (dd["StartDate"].dt.month == int(last[1]))].copy()
                st.session_state["cmp_ctx"] = {"a": a_ctx, "b": b_ctx, "label_a": first[2], "label_b": last[2], "value_col": metric}
                return

            # -------------------------
            # Mode 3: Multi-year highlight table
            # - pick 2..5 years
            # - show Units_YYYY and Sales_YYYY columns
            # - highlight highest and lowest per row (for Sales columns)
            # -------------------------
            with c1:
                years_pick = st.multiselect(
                    "Years to view (pick 2 to 5)",
                    options=years,
                    default=years[-3:] if len(years) >= 3 else years,
                    key="cmp_years_pick_multi_v2"
                )
            with c2:
                metric = st.selectbox(
                    "Highlight based on",
                    options=["Sales", "Units"],
                    index=0,
                    key="cmp_multi_metric_v2"
                )

            years_pick = [int(y) for y in years_pick]
            if len(years_pick) < 2:
                st.info("Pick at least two years.")
                return
            years_pick = years_pick[:5]

            dd = d[d["Year"].isin(years_pick)].copy()
            if sel:
                dd = dd[dd[by].isin(sel)]

            # Store for Top SKU movers table: compare first year vs last year in the selection
            try:
                y_first = int(years_pick[0]); y_last = int(years_pick[-1])
                a_df = dd[dd["Year"] == y_last].copy()
                b_df = dd[dd["Year"] == y_first].copy()
                # Context for Explain + One-pager: first vs last year
                a_ctx = dd[dd["Year"] == y_first].copy()
                b_ctx = dd[dd["Year"] == y_last].copy()
                st.session_state["cmp_ctx"] = {
                    "a": a_ctx,
                    "b": b_ctx,
                    "label_a": str(y_first),
                    "label_b": str(y_last),
                    "value_col": metric,
                    "mode": "multi_year",
                    "multi_year_years": years_pick,
                    "multi_year_by": by,
                }
                st.session_state["movers_a_periods"] = [str(p) for p in sorted(a_df["MonthP"].unique().tolist())] if "MonthP" in a_df.columns else []
                st.session_state["movers_b_periods"] = [str(p) for p in sorted(b_df["MonthP"].unique().tolist())] if "MonthP" in b_df.columns else []
            except Exception:
                st.session_state["movers_a_periods"] = []
                st.session_state["movers_b_periods"] = []

            pieces = []
            for y in years_pick:
                gy = dd[dd["Year"] == int(y)].groupby(by, as_index=False).agg(**{
                    f"Units_{y}": ("Units", "sum"),
                    f"Sales_{y}": ("Sales", "sum"),
                })
                pieces.append(gy)

            out = pieces[0]
            for p in pieces[1:]:
                out = out.merge(p, on=by, how="outer")

            out = out.fillna(0.0)

            # Totals row
            total = {by: "TOTAL"}
            for c in out.columns:
                if c == by:
                    continue
                total[c] = float(out[c].sum()) if pd.api.types.is_numeric_dtype(out[c]) else ""
            out = pd.concat([out, pd.DataFrame([total])], ignore_index=True)

            # Column order
            cols = [by]
            for y in years_pick:
                cols += [f"Units_{y}", f"Sales_{y}"]
            disp = out[cols].copy()

            # Store the full multi-year table for PDF export
            try:
                if isinstance(st.session_state.get("cmp_ctx"), dict):
                    st.session_state["cmp_ctx"]["multi_year_table"] = disp.copy()
            except Exception:
                pass

            # Highlight: highest and lowest across selected years for chosen metric
            metric_cols = [f"{metric}_{y}" for y in years_pick if f"{metric}_{y}" in disp.columns]

            # --- Extra insights (trend, CAGR, sparkline) ---
            # Use the selected metric across the chosen years
            spark_chars = ["▁","▂","▃","▄","▅","▆","▇","█"]

            def _sparkline(vals):
                vals = [float(v) if v is not None and not pd.isna(v) else np.nan for v in vals]
                if len(vals) == 0 or all(pd.isna(v) for v in vals):
                    return ""
                vmin = np.nanmin(vals)
                vmax = np.nanmax(vals)
                # all equal -> flat line
                if np.isnan(vmin) or np.isnan(vmax) or np.isclose(vmin, vmax):
                    return "▁" * len(vals)
                out_s = []
                for v in vals:
                    if pd.isna(v):
                        out_s.append(" ")
                        continue
                    t = (v - vmin) / (vmax - vmin)
                    idx = int(round(t * (len(spark_chars) - 1)))
                    idx = max(0, min(len(spark_chars) - 1, idx))
                    out_s.append(spark_chars[idx])
                return "".join(out_s)

            def _pct_change(a, b):
                try:
                    a = float(a); b = float(b)
                except Exception:
                    return np.nan
                if a == 0:
                    return np.nan
                return (b - a) / a

            def _cagr(a, b, periods):
                try:
                    a = float(a); b = float(b)
                except Exception:
                    return np.nan
                if a <= 0 or b <= 0 or periods <= 0:
                    return np.nan
                return (b / a) ** (1.0 / periods) - 1.0

            # Build per-row series for chosen metric (exclude TOTAL row)
            metric_year_cols = [(y, f"{metric}_{y}") for y in years_pick if f"{metric}_{y}" in disp.columns]
            if metric_year_cols:
                series_vals = disp[[c for _, c in metric_year_cols]].copy()

                # Sparkline
                disp["Spark"] = series_vals.apply(lambda r: _sparkline(r.tolist()), axis=1)

                # Trend (first -> last)
                first_col = metric_year_cols[0][1]
                last_col = metric_year_cols[-1][1]
                pct = series_vals.apply(lambda r: _pct_change(r[first_col], r[last_col]), axis=1)
                disp["Trend"] = np.where(
                    pct.isna(),
                    "—",
                    np.where(pct > 0, "↑", np.where(pct < 0, "↓", "→"))
                )
                disp["Trend %"] = pct

                # CAGR across (n_years - 1) intervals
                periods = max(1, len(metric_year_cols) - 1)
                disp["CAGR %"] = series_vals.apply(lambda r: _cagr(r[first_col], r[last_col], periods), axis=1)

                # Clear insight columns on TOTAL row (if present)
                try:
                    is_total = disp[by].astype(str) == "TOTAL"
                    for c in ["Spark", "Trend", "Trend %", "CAGR %"]:
                        if c in disp.columns:
                            disp.loc[is_total, c] = ""
                except Exception:
                    pass

            # Move insight columns next to the year columns
            insight_cols = [c for c in ["Spark", "Trend", "Trend %", "CAGR %"] if c in disp.columns]
            if insight_cols:
                disp = disp[[by] + [c for c in disp.columns if c != by and c not in insight_cols] + insight_cols]
            def _hl_minmax(row):
                styles = [""] * len(row)
                # Don't highlight TOTAL row
                if str(row.iloc[0]) == "TOTAL":
                    return styles
                vals = []
                idxs = []
                for j, c in enumerate(disp.columns):
                    if c in metric_cols:
                        try:
                            v = float(row[c])
                        except Exception:
                            v = np.nan
                        vals.append(v)
                        idxs.append(j)
                if not vals:
                    return styles
                vmin = np.nanmin(vals)
                vmax = np.nanmax(vals)
                # If all equal, no highlight
                if np.isnan(vmin) or np.isnan(vmax) or np.isclose(vmin, vmax):
                    return styles
                for v, j in zip(vals, idxs):
                    if np.isclose(v, vmax):
                        styles[j] = "background-color: rgba(0, 200, 0, 0.18); font-weight: 600;"
                    elif np.isclose(v, vmin):
                        styles[j] = "background-color: rgba(220, 0, 0, 0.14);"
                return styles

            fmt = {}
            for c in disp.columns:
                if c.startswith("Units_"):
                    fmt[c] = fmt_int
                elif c.startswith("Sales_"):
                    fmt[c] = fmt_currency
            if "Trend %" in disp.columns:
                fmt["Trend %"] = lambda v: (f"{float(v)*100:.1f}%" if (v is not None and pd.notna(v) and str(v).strip() not in {"", "—"}) else "—")
            if "CAGR %" in disp.columns:
                fmt["CAGR %"] = lambda v: (f"{float(v)*100:.1f}%" if (v is not None and pd.notna(v) and str(v).strip() not in {"", "—"}) else "—")

            sty = disp.style.format(fmt).apply(_hl_minmax, axis=1)

            st.dataframe(sty, use_container_width=True, hide_index=True)




    def render_comparison_sku():
            st.subheader("SKU Comparison")
            st.session_state["cmp_ctx"] = {}

            if df_all.empty:
                st.info("No sales data yet.")
                return

            d = df_all.copy()
            d["StartDate"] = pd.to_datetime(d["StartDate"], errors="coerce")
            d = d[d["StartDate"].notna()].copy()
            # Optional: select individual SKUs to compare
            if "SKU" in d.columns:
                all_skus = sorted([str(x).strip() for x in d["SKU"].dropna().unique().tolist() if str(x).strip()])
                sku_pick = st.multiselect("Select individual SKUs (optional)", options=all_skus, default=[], key="skucmp_sku_pick")
                if sku_pick:
                    d = d[d["SKU"].astype(str).isin([str(s).strip() for s in sku_pick])].copy()


            # Month options across ALL years
            d["MonthP"] = d["StartDate"].dt.to_period("M")
            months = sorted(d["MonthP"].unique().tolist())
            month_labels = [m.to_timestamp().strftime("%B %Y") for m in months]
            label_to_period = dict(zip(month_labels, months))

            d["Year"] = d["StartDate"].dt.year.astype(int)
            years = sorted(d["Year"].dropna().unique().tolist())

            mode = st.radio(
                "Compare type",
                options=["A vs B (Months)", "A vs B (Years)", "Multi-year (high/low highlight)", "Multi-month across years"],
                index=0,
                horizontal=True,
                key="skucmp_mode_v2"
            )

            c1, c2, c3 = st.columns([2, 2, 1])
            with c3:
                filt_by = st.selectbox("Filter by", ["All", "Retailer", "Vendor"], index=0, key="skucmp_filter_by_v2")

            # Optional limiter list
            sel = []
            if filt_by == "Retailer":
                opts = sorted([x for x in d["Retailer"].dropna().unique().tolist() if str(x).strip()])
                sel = st.multiselect("Limit to retailer(s) (optional)", options=opts, key="skucmp_limit_retailer_v2")
            elif filt_by == "Vendor":
                opts = sorted([x for x in d["Vendor"].dropna().unique().tolist() if str(x).strip()])
                sel = st.multiselect("Limit to vendor(s) (optional)", options=opts, key="skucmp_limit_vendor_v2")

            def _apply_filter(dd: pd.DataFrame) -> pd.DataFrame:
                if filt_by == "Retailer" and sel:
                    return dd[dd["Retailer"].isin(sel)].copy()
                if filt_by == "Vendor" and sel:
                    return dd[dd["Vendor"].isin(sel)].copy()
                return dd

            def _render_a_vs_b(da: pd.DataFrame, db: pd.DataFrame, label_a: str, label_b: str):
                ga = da.groupby("SKU", as_index=False).agg(Units_A=("Units","sum"), Sales_A=("Sales","sum"))
                gb = db.groupby("SKU", as_index=False).agg(Units_B=("Units","sum"), Sales_B=("Sales","sum"))

                out = ga.merge(gb, on="SKU", how="outer").fillna(0.0)
                out["Units_Diff"] = out["Units_A"] - out["Units_B"]
                out["Sales_Diff"] = out["Sales_A"] - out["Sales_B"]
                out["Units_%"] = out["Units_Diff"] / out["Units_B"].replace(0, np.nan)
                out["Sales_%"] = out["Sales_Diff"] / out["Sales_B"].replace(0, np.nan)

                # Add Vendor from map for context
                try:
                    if isinstance(vmap, pd.DataFrame) and "SKU" in vmap.columns and "Vendor" in vmap.columns:
                        sku_vendor = vmap[["SKU","Vendor"]].drop_duplicates()
                        out = out.merge(sku_vendor, on="SKU", how="left")
                except Exception:
                    pass

                total = {
                    "SKU": "TOTAL",
                    "Units_A": out["Units_A"].sum(),
                    "Sales_A": out["Sales_A"].sum(),
                    "Units_B": out["Units_B"].sum(),
                    "Sales_B": out["Sales_B"].sum(),
                }
                total["Units_Diff"] = total["Units_A"] - total["Units_B"]
                total["Sales_Diff"] = total["Sales_A"] - total["Sales_B"]
                total["Units_%"] = total["Units_Diff"] / total["Units_B"] if total["Units_B"] else np.nan
                total["Sales_%"] = total["Sales_Diff"] / total["Sales_B"] if total["Sales_B"] else np.nan
                out = pd.concat([out, pd.DataFrame([total])], ignore_index=True)

                cols = ["SKU"] + (["Vendor"] if "Vendor" in out.columns else []) + ["Units_A","Sales_A","Units_B","Sales_B","Units_Diff","Units_%","Sales_Diff","Sales_%"]
                disp = out[cols].copy()
                disp = disp.rename(columns={
                    "Units_A": f"Units ({label_a})",
                    "Sales_A": f"Sales ({label_a})",
                    "Units_B": f"Units ({label_b})",
                    "Sales_B": f"Sales ({label_b})",
                })

                sty = disp.style.format({
                    f"Units ({label_a})": fmt_int,
                    f"Units ({label_b})": fmt_int,
                    "Units_Diff": fmt_int_signed,
                    "Units_%": lambda v: f"{v*100:.1f}%" if pd.notna(v) else "—",
                    f"Sales ({label_a})": fmt_currency,
                    f"Sales ({label_b})": fmt_currency,
                    "Sales_Diff": fmt_currency_signed,
                    "Sales_%": lambda v: f"{v*100:.1f}%" if pd.notna(v) else "—",
                }).applymap(lambda v: f"color: {_color(v)};", subset=["Units_Diff","Sales_Diff"])

                st.dataframe(sty, use_container_width=True, hide_index=True, height=_table_height(disp, max_px=1200))

            # -------------------------
            # Mode 1: A vs B (Months)
            # -------------------------
            if mode == "A vs B (Months)":
                with c1:
                    a_pick = st.multiselect(
                        "Selection A (one or more months)",
                        options=month_labels,
                        default=month_labels[-1:] if month_labels else [],
                        key="skucmp_a_months_v2"
                    )
                with c2:
                    b_pick = st.multiselect(
                        "Selection B (one or more months)",
                        options=month_labels,
                        default=month_labels[-2:-1] if len(month_labels) >= 2 else [],
                        key="skucmp_b_months_v2"
                    )

                a_periods = [label_to_period[x] for x in a_pick if x in label_to_period]
                b_periods = [label_to_period[x] for x in b_pick if x in label_to_period]

                st.session_state["movers_a_periods"] = [str(p) for p in a_periods]
                st.session_state["movers_b_periods"] = [str(p) for p in b_periods]

                if not a_periods or not b_periods:
                    st.info("Pick at least one month in Selection A and Selection B.")
                    return

                da = _apply_filter(d[d["MonthP"].isin(a_periods)])
                db = _apply_filter(d[d["MonthP"].isin(b_periods)])

                label_a = " + ".join(a_pick) if a_pick else "A"
                label_b = " + ".join(b_pick) if b_pick else "B"
                _render_a_vs_b(da, db, label_a, label_b)
                st.session_state["cmp_ctx"] = {"a": da.copy(), "b": db.copy(), "label_a": label_a, "label_b": label_b, "value_col": "Sales"}
                return

            # -------------------------
            # Mode 2: A vs B (Years)
            # -------------------------
            if mode == "A vs B (Years)":
                with c1:
                    years_a = st.multiselect(
                        "Selection A (one or more years)",
                        options=years,
                        default=years[-2:-1] if len(years) >= 2 else years,
                        key="skucmp_years_a_v2"
                    )
                with c2:
                    years_b = st.multiselect(
                        "Selection B (one or more years)",
                        options=years,
                        default=years[-1:] if years else [],
                        key="skucmp_years_b_v2"
                    )

                if not years_a or not years_b:
                    st.info("Pick at least one year in Selection A and Selection B.")
                    return

                da = _apply_filter(d[d["Year"].isin([int(y) for y in years_a])])
                db = _apply_filter(d[d["Year"].isin([int(y) for y in years_b])])

                try:
                    st.session_state["movers_a_periods"] = [str(p) for p in sorted(da["MonthP"].unique().tolist())]
                    st.session_state["movers_b_periods"] = [str(p) for p in sorted(db["MonthP"].unique().tolist())]
                except Exception:
                    st.session_state["movers_a_periods"] = []
                    st.session_state["movers_b_periods"] = []

                label_a = " + ".join([str(y) for y in years_a])
                label_b = " + ".join([str(y) for y in years_b])
                _render_a_vs_b(da, db, label_a, label_b)
                st.session_state["cmp_ctx"] = {"a": da.copy(), "b": db.copy(), "label_a": label_a, "label_b": label_b, "value_col": "Sales"}
                return


            # -------------------------
            # Mode 2b: Multi-month across years (pick Month+Year periods)
            # -------------------------
            if mode == "Multi-month across years":
                month_names = ["January","February","March","April","May","June","July","August","September","October","November","December"]
                period_options = []
                for y in years:
                    for mname in month_names:
                        period_options.append(f"{mname} {y}")

                with c1:
                    periods_pick = st.multiselect(
                        "Month + Year periods",
                        options=period_options,
                        default=[period_options[-1]] if period_options else [],
                        key="skucmp_mm_periods",
                    )
                with c2:
                    metric = st.selectbox(
                        "Highlight based on",
                        options=["Sales", "Units"],
                        index=0,
                        key="skucmp_mm_metric",
                    )
                with c3:
                    topn = st.selectbox("Show", options=[25, 50, 100, 250], index=1, key="skucmp_mm_topn")

                if not periods_pick:
                    st.info("Pick at least one Month + Year period.")
                    return

                month_to_num = {m:i+1 for i,m in enumerate(month_names)}
                pairs = []
                for p in periods_pick:
                    try:
                        parts = p.split(" ")
                        mname = " ".join(parts[:-1])
                        yy = int(parts[-1])
                        mn = month_to_num.get(mname, None)
                        if mn is not None:
                            pairs.append((yy, mn, p))
                    except Exception:
                        continue
                if not pairs:
                    st.info("No valid periods selected.")
                    return

                dd = d.copy()
                mask = False
                for (yy, mn, _) in pairs:
                    mask = mask | ((dd["Year"] == int(yy)) & (dd["StartDate"].dt.month == int(mn)))
                dd = dd[mask].copy()

                if dd.empty:
                    st.info("No data found for those periods with current filters.")
                    return

                pieces = []
                for (yy, mn, lab_full) in pairs:
                    lab = lab_full.replace("January","Jan").replace("February","Feb").replace("March","Mar").replace("April","Apr").replace("June","Jun").replace("July","Jul").replace("August","Aug").replace("September","Sep").replace("October","Oct").replace("November","Nov").replace("December","Dec")
                    dyy = dd[(dd["Year"] == int(yy)) & (dd["StartDate"].dt.month == int(mn))].copy()
                    gy = dyy.groupby("SKU", as_index=False).agg(**{
                        f"Units_{lab}": ("Units", "sum"),
                        f"Sales_{lab}": ("Sales", "sum"),
                    })
                    pieces.append((lab, gy))

                out = pieces[0][1]
                for lab, p in pieces[1:]:
                    out = out.merge(p, on="SKU", how="outer")
                out = out.fillna(0.0)

                total = {"SKU": "TOTAL"}
                for c in out.columns:
                    if c == "SKU":
                        continue
                    total[c] = float(out[c].sum()) if pd.api.types.is_numeric_dtype(out[c]) else ""
                out = pd.concat([out, pd.DataFrame([total])], ignore_index=True)

                cols = ["SKU"]
                for lab, _ in pieces:
                    cols += [f"Units_{lab}", f"Sales_{lab}"]
                disp = out[cols].copy()

                metric_cols = [c for c in disp.columns if c.startswith(metric + "_")]
                if metric_cols:
                    disp = disp.sort_values(metric_cols[-1], ascending=False).head(int(topn)).copy()

                spark_chars = ["▁","▂","▃","▄","▅","▆","▇","█"]
                def _spark(vals):
                    vals = [float(v) if v is not None and not pd.isna(v) else np.nan for v in vals]
                    if len(vals) == 0 or all(pd.isna(v) for v in vals):
                        return ""
                    vmin = np.nanmin(vals); vmax = np.nanmax(vals)
                    if np.isnan(vmin) or np.isnan(vmax) or np.isclose(vmin, vmax):
                        return "▁" * len(vals)
                    out_s=[]
                    for v in vals:
                        if pd.isna(v):
                            out_s.append(" "); continue
                        t=(v-vmin)/(vmax-vmin)
                        idx=int(round(t*(len(spark_chars)-1)))
                        idx=max(0,min(len(spark_chars)-1,idx))
                        out_s.append(spark_chars[idx])
                    return "".join(out_s)

                def _cagr(a,b,periods):
                    try:
                        a=float(a); b=float(b)
                    except Exception:
                        return np.nan
                    if a <= 0 or b <= 0 or periods <= 0:
                        return np.nan
                    return (b/a)**(1.0/periods)-1.0

                if metric_cols:
                    series_vals = disp[metric_cols].copy()
                    disp["Spark"] = series_vals.apply(lambda r: _spark(r.tolist()), axis=1)
                    periods_n = max(1, len(metric_cols)-1)
                    disp["CAGR %"] = series_vals.apply(lambda r: _cagr(r[metric_cols[0]], r[metric_cols[-1]], periods_n), axis=1)

                def _hl(row):
                    styles=[""]*len(row)
                    if not metric_cols:
                        return styles
                    vals=[]
                    for c in metric_cols:
                        try:
                            vals.append(float(row[c]))
                        except Exception:
                            vals.append(np.nan)
                    if len(vals)==0 or all(pd.isna(v) for v in vals):
                        return styles
                    vmax=np.nanmax(vals); vmin=np.nanmin(vals)
                    for i,c in enumerate(row.index):
                        if c in metric_cols:
                            v=float(row[c]) if pd.notna(row[c]) else np.nan
                            if pd.notna(v) and np.isclose(v,vmax):
                                styles[i]="background-color: rgba(0, 200, 0, 0.18); font-weight: 600;"
                            elif pd.notna(v) and np.isclose(v,vmin):
                                styles[i]="background-color: rgba(220, 0, 0, 0.14);"
                    return styles

                fmt_map = {c: fmt_int for c in disp.columns if c.startswith("Units_")}
                fmt_map.update({c: fmt_currency for c in disp.columns if c.startswith("Sales_")})
                if "CAGR %" in disp.columns:
                    fmt_map["CAGR %"] = lambda v: f"{v*100:.1f}%" if pd.notna(v) else "—"

                st.caption("Multi-month across years (selected Month+Year periods)")
                sty = disp.style.format(fmt_map)
                if metric_cols:
                    sty = sty.apply(_hl, axis=1)
                st.dataframe(sty, use_container_width=True, hide_index=True)

                first = pairs[0]; last = pairs[-1]
                a_ctx = dd[(dd["Year"] == int(first[0])) & (dd["StartDate"].dt.month == int(first[1]))].copy()
                b_ctx = dd[(dd["Year"] == int(last[0])) & (dd["StartDate"].dt.month == int(last[1]))].copy()
                st.session_state["cmp_ctx"] = {"a": a_ctx, "b": b_ctx, "label_a": first[2], "label_b": last[2], "value_col": metric}
                return

            # -------------------------
            # Mode 3: Multi-year highlight table
            # -------------------------
            with c1:
                years_pick = st.multiselect(
                    "Years to view (pick 2 to 5)",
                    options=years,
                    default=years[-3:] if len(years) >= 3 else years,
                    key="skucmp_years_pick_multi_v2"
                )
            with c2:
                metric = st.selectbox(
                    "Highlight based on",
                    options=["Sales", "Units"],
                    index=0,
                    key="skucmp_multi_metric_v2"
                )

            years_pick = [int(y) for y in years_pick]
            if len(years_pick) < 2:
                st.info("Pick at least two years.")
                return
            years_pick = years_pick[:5]

            dd = _apply_filter(d[d["Year"].isin(years_pick)].copy())

            # Movers compare first vs last selected year
            try:
                y_first = int(years_pick[0]); y_last = int(years_pick[-1])
                a_df = dd[dd["Year"] == y_last].copy()
                b_df = dd[dd["Year"] == y_first].copy()
                # Context for Explain + One-pager: first vs last year
                a_ctx = dd[dd["Year"] == y_first].copy()
                b_ctx = dd[dd["Year"] == y_last].copy()
                st.session_state["cmp_ctx"] = {"a": a_ctx, "b": b_ctx, "label_a": str(y_first), "label_b": str(y_last), "value_col": metric}
                st.session_state["movers_a_periods"] = [str(p) for p in sorted(a_df["MonthP"].unique().tolist())]
                st.session_state["movers_b_periods"] = [str(p) for p in sorted(b_df["MonthP"].unique().tolist())]
            except Exception:
                st.session_state["movers_a_periods"] = []
                st.session_state["movers_b_periods"] = []

            pieces = []
            for y in years_pick:
                gy = dd[dd["Year"] == int(y)].groupby("SKU", as_index=False).agg(**{
                    f"Units_{y}": ("Units", "sum"),
                    f"Sales_{y}": ("Sales", "sum"),
                })
                pieces.append(gy)

            out = pieces[0]
            for p in pieces[1:]:
                out = out.merge(p, on="SKU", how="outer")

            out = out.fillna(0.0)

            # Totals row
            total = {"SKU": "TOTAL"}
            for c in out.columns:
                if c == "SKU":
                    continue
                total[c] = float(out[c].sum()) if pd.api.types.is_numeric_dtype(out[c]) else ""
            out = pd.concat([out, pd.DataFrame([total])], ignore_index=True)

            cols = ["SKU"]
            for y in years_pick:
                cols += [f"Units_{y}", f"Sales_{y}"]
            disp = out[cols].copy()

            metric_cols = [f"{metric}_{y}" for y in years_pick if f"{metric}_{y}" in disp.columns]

            def _hl_minmax(row):
                styles = [""] * len(row)
                if str(row.iloc[0]) == "TOTAL":
                    return styles
                vals = []
                idxs = []
                for j, c in enumerate(disp.columns):
                    if c in metric_cols:
                        try:
                            v = float(row[c])
                        except Exception:
                            v = np.nan
                        vals.append(v)
                        idxs.append(j)
                if not vals:
                    return styles
                vmin = np.nanmin(vals)
                vmax = np.nanmax(vals)
                if np.isnan(vmin) or np.isnan(vmax) or np.isclose(vmin, vmax):
                    return styles
                for v, j in zip(vals, idxs):
                    if np.isclose(v, vmax):
                        styles[j] = "background-color: rgba(0, 200, 0, 0.18); font-weight: 600;"
                    elif np.isclose(v, vmin):
                        styles[j] = "background-color: rgba(220, 0, 0, 0.14);"
                return styles

            fmt = {}
            for c in disp.columns:
                if c.startswith("Units_"):
                    fmt[c] = fmt_int
                elif c.startswith("Sales_"):
                    fmt[c] = fmt_currency

            st.dataframe(disp.style.format(fmt).apply(_hl_minmax, axis=1), use_container_width=True, hide_index=True, height=_table_height(disp, max_px=1200))


    def render_sku_health():
            st.subheader("SKU Health Score")

            if df_all.empty:
                st.info("No sales data yet.")
                return

            d = df_all.copy()
            d["StartDate"] = pd.to_datetime(d["StartDate"], errors="coerce")
            d = d[d["StartDate"].notna()].copy()

            d["Year"] = d["StartDate"].dt.year.astype(int)
            d["Month"] = d["StartDate"].dt.month.astype(int)
            d["MonthP"] = d["StartDate"].dt.to_period("M")

            compare_mode = st.selectbox(
                "Compare mode",
                options=["Year vs Year", "Month vs Month (multi-month)"],
                index=0,
                key="sh_compare_mode"
            )

            basis = st.radio("Primary basis", options=["Sales", "Units"], index=0, horizontal=True, key="sh_basis")

            # Shared filters
            f1, f2, f3, f4 = st.columns([2, 2, 1, 1])
            with f1:
                vendor_filter = st.multiselect(
                    "Vendor filter (optional)",
                    options=sorted([x for x in d["Vendor"].dropna().unique().tolist() if str(x).strip()]),
                    key="sh_vendor_filter"
                )
            with f2:
                retailer_filter = st.multiselect(
                    "Retailer filter (optional)",
                    options=sorted([x for x in d["Retailer"].dropna().unique().tolist() if str(x).strip()]),
                    key="sh_retailer_filter"
                )
            with f3:
                top_n = st.number_input("Top N", min_value=20, max_value=2000, value=200, step=20, key="sh_topn")
            with f4:
                status_pick = st.multiselect(
                    "Status",
                    options=["🔥 Strong","📈 Growing","⚠ Watch","❌ At Risk"],
                    default=["🔥 Strong","📈 Growing","⚠ Watch","❌ At Risk"],
                    key="sh_status"
                )

            dd = d.copy()
            if vendor_filter:
                dd = dd[dd["Vendor"].isin(vendor_filter)]
            if retailer_filter:
                dd = dd[dd["Retailer"].isin(retailer_filter)]

            # Build A vs B selections
            if compare_mode == "Year vs Year":
                years = sorted(dd["Year"].unique().tolist())
                c1, c2, c3 = st.columns([1, 1, 2])
                with c1:
                    base_year = st.selectbox("Base year", options=years, index=max(0, len(years)-2), key="sh_base")
                with c2:
                    comp_year = st.selectbox("Compare to", options=years, index=len(years)-1 if years else 0, key="sh_comp")
                with c3:
                    pmode = st.selectbox("Period", options=["Full year", "Specific months"], index=0, key="sh_period_mode")

                sel_months = list(range(1,13))
                if pmode == "Specific months":
                    month_name = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}
                    month_list = [month_name[i] for i in range(1,13)]
                    sel_names = st.multiselect("Months", options=month_list, default=[month_list[0]], key="sh_months_pick")
                    sel_months = [k for k,v in month_name.items() if v in sel_names]

                a = dd[(dd["Year"] == int(base_year)) & (dd["Month"].isin(sel_months))].copy()
                b = dd[(dd["Year"] == int(comp_year)) & (dd["Month"].isin(sel_months))].copy()

                a_label = str(base_year)
                b_label = str(comp_year)

            else:
                # Month vs Month (can be same year or different years)
                months = sorted(dd["MonthP"].unique().tolist())
                month_labels = [m.to_timestamp().strftime("%B %Y") for m in months]
                label_to_period = dict(zip(month_labels, months))

                c1, c2 = st.columns(2)
                with c1:
                    a_pick = st.multiselect(
                        "Selection A (one or more months)",
                        options=month_labels,
                        default=month_labels[-1:] if month_labels else [],
                        key="sh_mm_a"
                    )
                with c2:
                    b_pick = st.multiselect(
                        "Selection B (one or more months)",
                        options=month_labels,
                        default=month_labels[-2:-1] if len(month_labels) >= 2 else [],
                        key="sh_mm_b"
                    )

                a_periods = [label_to_period[x] for x in a_pick if x in label_to_period]
                b_periods = [label_to_period[x] for x in b_pick if x in label_to_period]

                if (not a_periods) or (not b_periods):
                    st.info("Pick at least one month in Selection A and Selection B.")
                    return

                a = dd[dd["MonthP"].isin(a_periods)].copy()
                b = dd[dd["MonthP"].isin(b_periods)].copy()

                a_label = "Selection A"
                b_label = "Selection B"

            ga = a.groupby("SKU", as_index=False).agg(Sales_A=("Sales","sum"), Units_A=("Units","sum"))
            gb = b.groupby("SKU", as_index=False).agg(Sales_B=("Sales","sum"), Units_B=("Units","sum"))
            out = ga.merge(gb, on="SKU", how="outer").fillna(0.0)

            # Coverage context (based on B selection)
            cov = b.groupby("SKU", as_index=False).agg(Retailers=("Retailer","nunique"), ActiveWeeks=("StartDate","nunique"))
            out = out.merge(cov, on="SKU", how="left").fillna({"Retailers": 0, "ActiveWeeks": 0})

            out["Δ Sales"] = out["Sales_B"] - out["Sales_A"]
            out["Δ Units"] = out["Units_B"] - out["Units_A"]
            out["Sales %"] = out["Δ Sales"] / out["Sales_A"].replace(0, np.nan)
            out["Units %"] = out["Δ Units"] / out["Units_A"].replace(0, np.nan)

            out["Score"] = out["Δ Sales"] if basis == "Sales" else out["Δ Units"]

            def _status(row):
                a0 = float(row["Sales_A"] if basis=="Sales" else row["Units_A"])
                b0 = float(row["Sales_B"] if basis=="Sales" else row["Units_B"])
                delta = b0 - a0
                if a0 == 0 and b0 > 0:
                    return "📈 Growing"
                if a0 > 0 and b0 == 0:
                    return "❌ At Risk"
                if delta > 0:
                    return "🔥 Strong"
                if delta < 0:
                    return "⚠ Watch"
                return "⚠ Watch"

            out["Status"] = out.apply(_status, axis=1)
            out = out[out["Status"].isin(status_pick)].copy()
            out = out.sort_values("Score", ascending=False, kind="mergesort").head(int(top_n))

            # Vendor lookup
            try:
                if isinstance(vmap, pd.DataFrame) and "SKU" in vmap.columns and "Vendor" in vmap.columns:
                    out = out.merge(vmap[["SKU","Vendor"]].drop_duplicates(), on="SKU", how="left")
            except Exception:
                pass

            cols = ["SKU"] + (["Vendor"] if "Vendor" in out.columns else []) + ["Status","Sales_A","Sales_B","Δ Sales","Sales %","Units_A","Units_B","Δ Units","Units %","Retailers","ActiveWeeks"]
            disp = out[cols].copy()
            disp = disp.rename(columns={
                "Sales_A": a_label,
                "Sales_B": b_label,
                "Units_A": f"Units {a_label}",
                "Units_B": f"Units {b_label}",
            })
            disp = make_unique_columns(disp)

            st.dataframe(
                disp.style.format({
                    a_label: fmt_currency,
                    b_label: fmt_currency,
                    "Δ Sales": fmt_currency,
                    "Sales %": lambda v: f"{v*100:.1f}%" if pd.notna(v) else "—",
                    f"Units {a_label}": fmt_int,
                    f"Units {b_label}": fmt_int,
                    "Δ Units": fmt_int,
                    "Units %": lambda v: f"{v*100:.1f}%" if pd.notna(v) else "—",
                    "Retailers": fmt_int,
                    "ActiveWeeks": fmt_int,
                }),
                use_container_width=True,
                hide_index=True,
                height=_table_height(disp, max_px=1200)
            )

    def render_lost_sales():
            st.subheader("Lost Sales Detector")

            if df_all.empty:
                st.info("No sales data yet.")
                return

            d = df_all.copy()
            d["StartDate"] = pd.to_datetime(d["StartDate"], errors="coerce")
            d = d[d["StartDate"].notna()].copy()
            d["Year"] = d["StartDate"].dt.year.astype(int)
            d["Month"] = d["StartDate"].dt.month.astype(int)

            years = sorted(d["Year"].unique().tolist())
            month_name = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}
            month_list = [month_name[i] for i in range(1,13)]

            c1, c2, c3 = st.columns([1, 1, 2])
            with c1:
                base_year = st.selectbox("Base year", options=years, index=max(0, len(years)-2), key="ls_base")
            with c2:
                comp_year = st.selectbox("Compare to", options=years, index=len(years)-1, key="ls_comp")
            with c3:
                basis = st.radio("Basis", options=["Sales", "Units"], index=0, horizontal=True, key="ls_basis")

            pmode = st.selectbox("Period", options=["Full year", "Specific months"], index=0, key="ls_period_mode")
            sel_months = list(range(1,13))
            if pmode == "Specific months":
                sel_names = st.multiselect("Months", options=month_list, default=[month_list[0]], key="ls_months_pick")
                sel_months = [k for k,v in month_name.items() if v in sel_names]

            value_col = "Sales" if basis == "Sales" else "Units"

            a = d[(d["Year"] == int(base_year)) & (d["Month"].isin(sel_months))].copy()
            b = d[(d["Year"] == int(comp_year)) & (d["Month"].isin(sel_months))].copy()

            # Build SKU-level totals for both Units and Sales (for summary + gained/lost)
            ga_all = a.groupby("SKU", as_index=False).agg(Units_A=("Units","sum"), Sales_A=("Sales","sum"))
            gb_all = b.groupby("SKU", as_index=False).agg(Units_B=("Units","sum"), Sales_B=("Sales","sum"))
            sku_all = ga_all.merge(gb_all, on="SKU", how="outer").fillna(0.0)

            lost_mask = ((sku_all["Units_A"] > 0) | (sku_all["Sales_A"] > 0)) & (sku_all["Units_B"] == 0) & (sku_all["Sales_B"] == 0)
            gained_mask = (sku_all["Units_A"] == 0) & (sku_all["Sales_A"] == 0) & ((sku_all["Units_B"] > 0) | (sku_all["Sales_B"] > 0))

            lost_all = sku_all[lost_mask].copy()
            gained_all = sku_all[gained_mask].copy()

            # Summary totals
            lost_units = float(lost_all["Units_A"].sum()) if not lost_all.empty else 0.0
            lost_sales = float(lost_all["Sales_A"].sum()) if not lost_all.empty else 0.0
            gained_units = float(gained_all["Units_B"].sum()) if not gained_all.empty else 0.0
            gained_sales = float(gained_all["Sales_B"].sum()) if not gained_all.empty else 0.0

            net_units = gained_units - lost_units
            net_sales = gained_sales - lost_sales

            # Show net impact (green if net positive, red if net negative)
            net_color = "#2ecc71" if net_sales > 0 else ("#e74c3c" if net_sales < 0 else "#999999")
            st.markdown(
                f"""<div style="padding:10px 12px; border-radius:10px; border:1px solid #2a2a2a;">
                <div style="font-size:0.95rem; font-weight:700;">Net change (Gained − Lost)</div>
                <div style="margin-top:6px; color:{net_color}; font-weight:800; font-size:1.15rem;">
                  {fmt_currency(net_sales)} &nbsp;|&nbsp; {fmt_int(net_units)} units
                </div>
                <div style="margin-top:6px; font-size:0.9rem; color:#aaaaaa;">
                  Lost: {fmt_currency(lost_sales)} / {fmt_int(lost_units)} units &nbsp;&nbsp;•&nbsp;&nbsp;
                  Gained: {fmt_currency(gained_sales)} / {fmt_int(gained_units)} units
                </div>
                </div>""",
                unsafe_allow_html=True
            )

            # Basis-specific (existing behavior)
            ga = a.groupby("SKU", as_index=False).agg(A=(value_col,"sum"))
            gb = b.groupby("SKU", as_index=False).agg(B=(value_col,"sum"))
            sku = ga.merge(gb, on="SKU", how="outer").fillna(0.0)
            sku["Delta"] = sku["A"] - sku["B"]
            sku["Pct"] = sku["Delta"] / sku["A"].replace(0, np.nan)

            lost = sku[(sku["A"] > 0) & (sku["B"] == 0)].copy().sort_values("A", ascending=False).head(200)
            gained = sku[(sku["A"] == 0) & (sku["B"] > 0)].copy().sort_values("B", ascending=False).head(200)
            drops = sku[(sku["A"] > 0) & (sku["B"] > 0) & (sku["Delta"] < 0)].copy().sort_values("Delta").head(200)

            ra = a.groupby(["SKU","Retailer"], as_index=False).agg(A=(value_col,"sum"))
            rb = b.groupby(["SKU","Retailer"], as_index=False).agg(B=(value_col,"sum"))
            rr = ra.merge(rb, on=["SKU","Retailer"], how="outer").fillna(0.0)
            rr["Delta"] = rr["A"] - rr["B"]
            lost_retail = rr[(rr["A"] > 0) & (rr["B"] == 0)].copy().sort_values("A", ascending=False).head(300)

            try:
                if isinstance(vmap, pd.DataFrame) and "SKU" in vmap.columns and "Vendor" in vmap.columns:
                    vend = vmap[["SKU","Vendor"]].drop_duplicates()
                    lost = lost.merge(vend, on="SKU", how="left")
                    drops = drops.merge(vend, on="SKU", how="left")
                    lost_retail = lost_retail.merge(vend, on="SKU", how="left")
            except Exception:
                pass

            def _fmt(v):
                return fmt_currency(v) if value_col == "Sales" else fmt_int(v)

            st.markdown("### Lost SKUs (sold in base period, zero in compare period)")
            lost_disp = lost[["SKU"] + (["Vendor"] if "Vendor" in lost.columns else []) + ["A"]].copy().rename(columns={"A": str(base_year)})
            lost_disp = make_unique_columns(lost_disp)
            st.dataframe(lost_disp.style.format({str(base_year): _fmt}), use_container_width=True, hide_index=True, height=650)

            st.markdown("### Gained SKUs (zero in base period, sold in compare period)")
            gained_disp = gained[["SKU"] + (["Vendor"] if "Vendor" in gained.columns else []) + ["B"]].copy().rename(columns={"B": str(comp_year)})
            gained_disp = make_unique_columns(gained_disp)
            st.dataframe(gained_disp.style.format({str(comp_year): _fmt}), use_container_width=True, hide_index=True, height=650)

            st.markdown("### Biggest declines (still selling, but down)")
            drops_disp = drops[["SKU"] + (["Vendor"] if "Vendor" in drops.columns else []) + ["A","B","Delta","Pct"]].copy()
            drops_disp = drops_disp.rename(columns={"A": str(base_year), "B": str(comp_year)})
            drops_disp = make_unique_columns(drops_disp)
            st.dataframe(
                drops_disp.style.format({
                    str(base_year): _fmt,
                    str(comp_year): _fmt,
                    "Delta": _fmt,
                    "Pct": lambda v: f"{v*100:.1f}%" if pd.notna(v) else "—",
                }),
                use_container_width=True,
                hide_index=True,
                height=650
            )

            st.markdown("### Lost retailers for specific SKUs")
            lost_retail_disp = lost_retail[["SKU","Retailer"] + (["Vendor"] if "Vendor" in lost_retail.columns else []) + ["A"]].copy()
            lost_retail_disp = lost_retail_disp.rename(columns={"A": str(base_year)})
            lost_retail_disp = make_unique_columns(lost_retail_disp)
            st.dataframe(lost_retail_disp.style.format({str(base_year): _fmt}), use_container_width=True, hide_index=True, height=700)

    def render_data_inventory():
            st.subheader("Data Inventory")

            if df_all.empty:
                st.info("No sales data yet.")
            else:
                d = df_all.copy()
                d["StartDate"] = pd.to_datetime(d["StartDate"], errors="coerce")
                d = d[d["StartDate"].notna()].copy()
                d["Year"] = d["StartDate"].dt.year.astype(int)

                st.markdown("### By year")
                by_year = d.groupby("Year", as_index=False).agg(
                    Units=("Units","sum"),
                    Sales=("Sales","sum"),
                    Retailers=("Retailer","nunique"),
                    Vendors=("Vendor","nunique"),
                    SKUs=("SKU","nunique"),
                ).sort_values("Year", ascending=False)
                st.dataframe(by_year.style.format({
                    "Units": fmt_int, "Sales": fmt_currency,
                    "Retailers": fmt_int, "Vendors": fmt_int, "SKUs": fmt_int
                }), use_container_width=True, hide_index=True)

                st.markdown("### By retailer (selected year)")
                years = sorted(d["Year"].unique().tolist())
                sel_y = st.selectbox("Year", options=years, index=len(years)-1, key="inv_year")
                dy = d[d["Year"] == int(sel_y)].copy()
                if "SourceFile" not in dy.columns:
                    dy["SourceFile"] = ""
                by_ret = dy.groupby("Retailer", as_index=False).agg(
                    Units=("Units","sum"),
                    Sales=("Sales","sum"),
                    SKUs=("SKU","nunique"),
                    Sources=("SourceFile","nunique"),
                ).sort_values("Sales", ascending=False)
                st.dataframe(by_ret.style.format({
                    "Units": fmt_int, "Sales": fmt_currency, "SKUs": fmt_int, "Sources": fmt_int
                }), use_container_width=True, height=_table_height(by_ret, max_px=900), hide_index=True)

                st.markdown("### By source file (selected year)")
                by_src = dy.groupby("SourceFile", as_index=False).agg(
                    Units=("Units","sum"),
                    Sales=("Sales","sum"),
                    Retailers=("Retailer","nunique"),
                    SKUs=("SKU","nunique"),
                ).sort_values("Sales", ascending=False)
                st.dataframe(by_src.style.format({
                    "Units": fmt_int, "Sales": fmt_currency, "Retailers": fmt_int, "SKUs": fmt_int
                }), use_container_width=True, height=_table_height(by_src, max_px=900), hide_index=True)




        # -------------------------
        # Insights & Alerts
        # -------------------------


    def render_edit_vendor_map():
            st.subheader("Edit Vendor Map")
            st.caption("Edit Vendor and Price. Click Save to update the default vendor map file used by the app.")
            vmap_disp = vmap[["Retailer","SKU","Vendor","Price","MapOrder"]].copy().sort_values(["Retailer","MapOrder"])
            show = vmap_disp.drop(columns=["MapOrder"]).copy()

            if edit_mode:
                edited = st.data_editor(show, use_container_width=True, hide_index=True, num_rows="dynamic")
                if st.button("Save Vendor Map"):
                    updated = edited.copy()
                    updated["Retailer"] = updated["Retailer"].map(_normalize_retailer)
                    updated["SKU"] = updated["SKU"].map(_normalize_sku)
                    updated["Vendor"] = updated["Vendor"].astype(str).str.strip()
                    updated["Price"] = pd.to_numeric(updated["Price"], errors="coerce")

                    # MapOrder based on current row order per retailer
                    updated["MapOrder"] = 0
                    for r, grp in updated.groupby("Retailer", sort=False):
                        for j, ix in enumerate(grp.index.tolist()):
                            updated.loc[ix, "MapOrder"] = j

                    updated.to_excel(DEFAULT_VENDOR_MAP, index=False)
                    st.success("Saved vendor map. Reloading…")
                    st.rerun()
            else:
                st.info("Turn on Edit Mode in the sidebar to edit.")
                st.dataframe(show, use_container_width=True, height=_table_height(show, max_px=1400), hide_index=True)

        # Backup / Restore


    def render_backup_restore():
            st.subheader("Backup / Restore")

            st.markdown("### Backup files")
            c1, c2, c3 = st.columns(3)

            with c1:
                st.markdown("#### Sales database")
                if DEFAULT_SALES_STORE.exists():
                    st.download_button("Download sales_store.csv", data=DEFAULT_SALES_STORE.read_bytes(), file_name="sales_store.csv", mime="text/csv")
                else:
                    st.info("No sales_store.csv yet.")

                up = st.file_uploader("Restore sales_store.csv", type=["csv"], key="restore_sales_csv")
                if st.button("Restore sales_store.csv", disabled=up is None, key="btn_restore_sales"):
                    DEFAULT_SALES_STORE.write_bytes(up.getbuffer())
                    st.success("Restored sales_store.csv. Reloading…")
                    st.rerun()

            with c2:
                st.markdown("#### Vendor map")
                if DEFAULT_VENDOR_MAP.exists():
                    st.download_button("Download vendor_map.xlsx", data=DEFAULT_VENDOR_MAP.read_bytes(), file_name="vendor_map.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                else:
                    st.info("No vendor_map.xlsx yet.")

                up2 = st.file_uploader("Restore vendor_map.xlsx", type=["xlsx"], key="restore_vm_xlsx")
                if st.button("Restore vendor_map.xlsx", disabled=up2 is None, key="btn_restore_vm"):
                    DEFAULT_VENDOR_MAP.write_bytes(up2.getbuffer())
                    st.success("Restored vendor_map.xlsx. Reloading…")
                    st.rerun()

            with c3:
                st.markdown("#### Price history")
                if DEFAULT_PRICE_HISTORY.exists():
                    st.download_button("Download price_history.csv", data=DEFAULT_PRICE_HISTORY.read_bytes(), file_name="price_history.csv", mime="text/csv")
                else:
                    st.info("No price_history.csv yet.")

                up3 = st.file_uploader("Restore price_history.csv", type=["csv"], key="restore_ph_csv")
                if st.button("Restore price_history.csv", disabled=up3 is None, key="btn_restore_ph"):
                    DEFAULT_PRICE_HISTORY.write_bytes(up3.getbuffer())
                    st.success("Restored price_history.csv. Reloading…")
                    st.rerun()

            st.markdown("#### Year locks")
            if DEFAULT_YEAR_LOCKS.exists():
                st.download_button("Download year_locks.json", data=DEFAULT_YEAR_LOCKS.read_bytes(), file_name="year_locks.json", mime="application/json")
            else:
                st.info("No year locks saved yet.")

            up4 = st.file_uploader("Restore year_locks.json", type=["json"], key="restore_year_locks")
            if st.button("Restore year_locks.json", disabled=up4 is None, key="btn_restore_year_locks"):
                DEFAULT_YEAR_LOCKS.write_bytes(up4.getbuffer())
                st.success("Restored year locks. Reloading…")
                st.rerun()

            st.divider()

            st.markdown("### Price changes (effective date)")
            st.caption("Upload a sheet with SKU + Price + StartDate. Optional Retailer column. Prices apply from StartDate forward and never change earlier weeks.")

            tmpl = pd.DataFrame([
                {"Retailer":"*", "SKU":"ABC123", "Price": 19.99, "StartDate":"2026-02-01"},
                {"Retailer":"home depot", "SKU":"XYZ999", "Price": 24.99, "StartDate":"2026-03-15"},
            ])
            st.download_button("Download template CSV", data=tmpl.to_csv(index=False).encode("utf-8"),
                               file_name="price_history_template.csv", mime="text/csv")

            ph_up = st.file_uploader("Upload price history (CSV or Excel)", type=["csv","xlsx"], key="ph_upload")
            if ph_up is not None:
                try:
                    if ph_up.name.lower().endswith(".csv"):
                        ph_new = pd.read_csv(ph_up)
                    else:
                        ph_new = pd.read_excel(ph_up)

                    st.markdown("#### Preview upload")
                    st.dataframe(ph_new.head(50), use_container_width=True, hide_index=True)

                    # Normalize + ignore blanks safely
                    cur_ph = load_price_history()
                    incoming, ignored = _prepare_price_history_upload(ph_new)
                    diff = _price_history_diff(cur_ph, incoming)

                    st.divider()
                    st.markdown("#### What will change")
                    c1, c2, c3, c4 = st.columns(4)
                    c1.metric("Rows uploaded", int(len(ph_new)))
                    c2.metric("Rows ignored (blank/invalid)", int(len(ignored)))
                    c3.metric("Inserts", int((diff["Action"] == "insert").sum()) if not diff.empty else 0)
                    c4.metric("Updates", int((diff["Action"] == "update").sum()) if not diff.empty else 0)

                    show_diff = diff.copy()
                    if not show_diff.empty:
                        show_diff["StartDate"] = pd.to_datetime(show_diff["StartDate"], errors="coerce").dt.date
                        sty = show_diff.style.format({
                            "OldPrice": lambda v: fmt_currency(v) if pd.notna(v) else "—",
                            "Price": lambda v: fmt_currency(v),
                            "PriceDiff": lambda v: fmt_currency(v) if pd.notna(v) else "—",
                        }).applymap(lambda v: "font-weight:700;" if str(v) in ["insert","update"] else "", subset=["Action"])
                        st.dataframe(sty, use_container_width=True, height=_table_height(show_diff, max_px=900), hide_index=True)
                        st.download_button("Download change preview (CSV)", data=show_diff.to_csv(index=False).encode("utf-8"),
                        file_name="price_history_changes_preview.csv", mime="text/csv")
                    else:
                        st.info("No valid rows found in this upload (all prices were blank/invalid).")

                    if not ignored.empty:
                        st.markdown("#### Ignored rows")
                        ign = ignored.copy()
                        ign["StartDate"] = pd.to_datetime(ign["StartDate"], errors="coerce").dt.date
                        st.dataframe(ign.head(200), use_container_width=True, height=_table_height(ign, max_px=600), hide_index=True)
                    st.download_button("Download ignored rows (CSV)", data=ign.to_csv(index=False).encode("utf-8"),
                        file_name="price_history_ignored_rows.csv", mime="text/csv")

                    if st.button("Apply price changes", key="btn_apply_prices"):
                        ins, upd, noop = upsert_price_history(ph_new)
                        st.success(f"Price history updated. Inserts: {ins}, Updates: {upd}, Unchanged: {noop}. Reloading…")
                        st.rerun()
                except Exception as e:
                    st.error(f"Could not read this file: {e}")

            if DEFAULT_PRICE_HISTORY.exists():
                if st.button("Clear ALL price history", key="btn_clear_ph"):
                    DEFAULT_PRICE_HISTORY.unlink(missing_ok=True)
                    st.success("Cleared. Reloading…")
                    st.rerun()

            st.divider()

            st.markdown("### Export enriched sales")
            if not df.empty:
                ex = df.copy()
                ex["StartDate"] = pd.to_datetime(ex["StartDate"], errors="coerce").dt.strftime("%Y-%m-%d")
                ex["EndDate"] = pd.to_datetime(ex["EndDate"], errors="coerce").dt.strftime("%Y-%m-%d")
                st.download_button("Download enriched_sales.csv", data=ex.to_csv(index=False).encode("utf-8"),
                                   file_name="enriched_sales.csv", mime="text/csv")
            else:
                st.info("No sales yet.")



        # -------------------------
        # Bulk Data Upload
        # -------------------------


    def render_bulk_data_upload():
        st.subheader("Bulk Data Upload (Multi-week / Multi-month)")

        st.markdown(
            """
            Use this when you get a **wide** retailer file (not week-by-week uploads).

            Expected format:
            - One sheet per retailer (or retailer name in cell **A1**)
            - Column **A** = SKU (starting row 2)
            - Row **1** from column **B** onward = week ranges (example: `1-1 / 1-3`)
            - Cells = Units sold for that SKU in that week
            - Sales uses your **current pricing** (Vendor Map / Price History). `UnitPrice` is left blank.
            """
        )

        locked_years = load_year_locks()
        years_opt = list(range(this_year - 6, this_year + 2))

        st.markdown("### Year locks")
        cL1, cL2 = st.columns([2, 1])
        with cL1:
            lock_pick = st.multiselect("Locked years (prevent edits)", options=years_opt, default=sorted(list(locked_years)), key="lock_pick")
        with cL2:
            if st.button("Save locks", key="btn_save_locks"):
                save_year_locks(set(int(y) for y in lock_pick))
                st.success("Saved year locks.")
                st.rerun()

        st.divider()

        bulk_upload = st.file_uploader(
            "Upload bulk data workbook (.xlsx)",
            type=["xlsx"],
            key="bulk_up_tab"
        )

        data_year = st.selectbox(
            "Data Year (for header parsing)",
            options=years_opt,
            index=years_opt.index(int(view_year)) if int(view_year) in years_opt else years_opt.index(this_year),
            key="bulk_data_year"
        )

        mode = st.radio(
            "Ingest mode",
            options=["Append (add rows)", "Overwrite year + retailer(s) (replace)"],
            index=0,
            horizontal=True,
            key="bulk_mode"
        )

        is_locked = int(data_year) in load_year_locks()
        if is_locked:
            st.error(f"Year {int(data_year)} is locked. Unlock it above to ingest data for this year.")

        c1, c2 = st.columns([1, 3])
        with c1:
            if st.button("Ingest Bulk Workbook", disabled=(bulk_upload is None) or is_locked, key="btn_ingest_bulk"):
                new_rows = read_yow_workbook(bulk_upload, year=int(data_year))

                if mode.startswith("Overwrite"):
                    retailers = set(new_rows["Retailer"].dropna().unique().tolist()) if not new_rows.empty else set()
                    overwrite_sales_rows(int(data_year), retailers)

                append_sales_to_store(new_rows)
                st.success("Bulk workbook ingested successfully.")
                st.rerun()

        with c2:
            st.caption("Append = adds rows. Overwrite = deletes existing rows for that year + retailer(s) found in the upload, then re-adds.")


    def render_seasonality():
            st.subheader("Seasonality (Top 20 seasonal SKUs)")

            if df_all.empty:
                st.info("No sales data yet.")
            else:
                d = df_all.copy()
                d["StartDate"] = pd.to_datetime(d["StartDate"], errors="coerce")
                d = d[d["StartDate"].notna()].copy()
                d["Year"] = d["StartDate"].dt.year.astype(int)
                d["Month"] = d["StartDate"].dt.month.astype(int)
                d["MonthP"] = d["StartDate"].dt.to_period("M")

                month_name = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}
                month_list = [month_name[i] for i in range(1,13)]

                years = sorted(d["Year"].unique().tolist())

                c1, c2, c3 = st.columns([1, 2, 2])
                with c1:
                    basis = st.radio("Basis", options=["Units", "Sales"], index=0, horizontal=True, key="sea_basis")
                with c2:
                    mode = st.selectbox("Timeframe", options=["Pick year", "Lookback"], index=0, key="sea_tf_mode")
                with c3:
                    # pick year or lookback window
                    if mode == "Pick year":
                        pick_year = st.selectbox("Year", options=["All years"] + [str(y) for y in years], index=0, key="sea_year")
                        month_mode = st.radio("Months", options=["All months (Jan–Dec)", "Custom months"], index=0, horizontal=True, key="sea_month_mode")
                        if month_mode == "Custom months":
                            sel_month_names = st.multiselect("Select months", options=month_list, default=month_list, key="sea_months_pick")
                            sel_months = [k for k,v in month_name.items() if v in sel_month_names]
                        else:
                            sel_months = list(range(1,13))
                        # Apply filters
                        d2 = d[d["Month"].isin(sel_months)].copy()
                        if pick_year != "All years":
                            d2 = d2[d2["Year"] == int(pick_year)].copy()
                    else:
                        lookback = st.selectbox("Look back", options=["12 months","24 months","36 months","All available"], index=0, key="sea_lookback")
                        if lookback == "All available":
                            d2 = d.copy()
                        else:
                            n = int(lookback.split()[0])
                            months = sorted(d["MonthP"].dropna().unique().tolist())
                            usem = months[-n:] if len(months) >= n else months
                            d2 = d[d["MonthP"].isin(usem)].copy()

                min_units = st.number_input(
                    "Minimum total units (within selected timeframe) to include a SKU",
                    min_value=0, max_value=1_000_000, value=20, step=5, key="sea_min_units"
                )

                value_col = "Units" if basis == "Units" else "Sales"

                # Monthly totals per SKU (within timeframe)
                m = d2.groupby(["SKU","MonthP"], as_index=False).agg(v=(value_col,"sum"))

                # Seasonality score computed on month-of-year buckets (Jan..Dec) from the same timeframe
                m_y = d2.groupby(["SKU","Month"], as_index=False).agg(v=(value_col,"sum"))
                tot = m_y.groupby("SKU", as_index=False).agg(total=("v","sum"))
                mx = m_y.sort_values("v", ascending=False).groupby("SKU", as_index=False).first().rename(columns={"Month":"PeakMonth","v":"PeakVal"})
                s = tot.merge(mx, on="SKU", how="left")
                s["SeasonalityScore"] = s["PeakVal"] / s["total"].replace(0, np.nan)

                # Filter by units sold in the timeframe (always units)
                units_tot = d2.groupby("SKU", as_index=False).agg(TotalUnits=("Units","sum"))
                s = s.merge(units_tot, on="SKU", how="left").fillna({"TotalUnits": 0})
                s = s[s["TotalUnits"] >= float(min_units)].copy()

                # Vendor labels
                try:
                    if isinstance(vmap, pd.DataFrame) and "SKU" in vmap.columns and "Vendor" in vmap.columns:
                        s = s.merge(vmap[["SKU","Vendor"]].drop_duplicates(), on="SKU", how="left")
                except Exception:
                    pass

                s = s.sort_values("SeasonalityScore", ascending=False)
                top = s.head(20).copy()
                top["PeakMonthName"] = top["PeakMonth"].map(month_name)

                st.markdown("### Top seasonal SKUs")
                tbl_cols = ["SKU"]
                if "Vendor" in top.columns:
                    tbl_cols.append("Vendor")
                tbl_cols += ["PeakMonthName","SeasonalityScore","TotalUnits"]

                tbl = top[tbl_cols].copy().rename(columns={
                    "PeakMonthName": "Peak Month",
                    "SeasonalityScore": "Seasonality",
                    "TotalUnits": "Total Units",
                })
                tbl = tbl.loc[:, ~tbl.columns.duplicated()].copy()

                st.dataframe(
                    tbl.style.format({
                        "Seasonality": lambda v: f"{v*100:.1f}%" if pd.notna(v) else "—",
                        "Total Units": fmt_int,
                    }),
                    use_container_width=True,
                    hide_index=True
                )

                st.divider()
                st.markdown("### Seasonal profiles (monthly totals in timeframe)")

                # Create a complete month index for charting, preserving chronological order
                months_all = sorted(d2["MonthP"].dropna().unique().tolist())
                if not months_all:
                    st.info("No months found in the selected timeframe.")
                    return

                dt_index = pd.PeriodIndex(months_all, freq="M").to_timestamp()

                for _, row in top.iterrows():
                    sku0 = row["SKU"]
                    vend0 = row.get("Vendor", "")
                    peak0 = row.get("PeakMonthName", "")
                    score0 = row.get("SeasonalityScore", np.nan)

                    title = f"{sku0}"
                    if pd.notna(vend0) and str(vend0).strip():
                        title += f" — {vend0}"
                    if pd.notna(score0):
                        title += f" | Peak: {peak0} | Seasonality: {score0*100:.1f}%"

                    st.markdown(f"**{title}**")

                    prof = m[m["SKU"] == sku0][["MonthP","v"]].copy()
                    prof["MonthP"] = prof["MonthP"].astype("period[M]")
                    prof = prof.set_index("MonthP").reindex(months_all).fillna(0.0)

                    chart_df = pd.DataFrame({f"{basis}": prof["v"].to_numpy()}, index=dt_index)
                    st.line_chart(chart_df)

    def render_runrate():
            st.subheader("Run-Rate Forecast")

            if df.empty:
                st.info("No sales data yet.")
            else:
                window = st.selectbox("Forecast window (weeks)", options=[4, 8, 12], index=0, key="rr_window")
                lookback = st.selectbox("Lookback for avg", options=[4, 8, 12], index=1, key="rr_lookback")
                level = st.selectbox("Level", options=["SKU", "Vendor", "Retailer"], index=0, key="rr_level")

                d = add_week_col(df)
                weeks = last_n_weeks(d, lookback)
                d = d[d["Week"].isin(weeks)].copy()

                if level == "SKU":
                    grp = ["Retailer","Vendor","SKU"]
                elif level == "Vendor":
                    grp = ["Vendor"]
                else:
                    grp = ["Retailer"]

                base = d.groupby(grp + ["Week"], as_index=False).agg(Units=("Units","sum"), Sales=("Sales","sum"))
                units_piv = base.pivot_table(index=grp, columns="Week", values="Units", aggfunc="sum", fill_value=0.0)
                sales_piv = base.pivot_table(index=grp, columns="Week", values="Sales", aggfunc="sum", fill_value=0.0)

                avg_units = nonzero_mean_rowwise(units_piv).fillna(0.0)
                avg_sales = nonzero_mean_rowwise(sales_piv).fillna(0.0)

                out = avg_units.reset_index().rename(columns={0:"AvgWeeklyUnits"})
                out["AvgWeeklySales"] = avg_sales.values
                out["ProjectedUnits"] = out["AvgWeeklyUnits"] * window
                out["ProjectedSales"] = out["AvgWeeklySales"] * window
                out = out.sort_values("ProjectedSales", ascending=False)

                disp = out.copy()
                disp["AvgWeeklyUnits"] = disp["AvgWeeklyUnits"].round(2)
                disp["ProjectedUnits"] = disp["ProjectedUnits"].round(0).astype(int)

                sty = disp.style.format({
                    "AvgWeeklyUnits": lambda v: fmt_2(v),
                    "AvgWeeklySales": lambda v: fmt_currency(v),
                    "ProjectedUnits": lambda v: fmt_int(v),
                    "ProjectedSales": lambda v: fmt_currency(v),
                })
                st.dataframe(sty, use_container_width=True, height=_table_height(disp, max_px=1200), hide_index=True)

                # SKU lookup (only shows when Level includes SKU)
                if isinstance(disp, pd.DataFrame) and (not disp.empty) and ('SKU' in disp.columns):
                    st.markdown('---')
                    st.markdown('### SKU lookup')
                    _sku_list = sorted(disp['SKU'].astype(str).dropna().unique().tolist())
                    sel_sku = st.selectbox('Select SKU', options=_sku_list, index=0, key='rr_sku_lookup') if _sku_list else None
                    if sel_sku:
                        row_df = disp[disp['SKU'].astype(str) == str(sel_sku)].copy()
                        # keep same formatting
                        row_sty = row_df.style.format({
                            'AvgWeeklyUnits': lambda v: fmt_2(v),
                            'AvgWeeklySales': lambda v: fmt_currency(v),
                            'ProjectedUnits': lambda v: fmt_int(v),
                            'ProjectedSales': lambda v: fmt_currency(v),
                        })
                        st.dataframe(row_sty, use_container_width=True, hide_index=True)


        # -------------------------
        # Seasonality Heatmap
        # -------------------------



    def render_alerts():
            st.subheader("Insights & Alerts")

            if df_all.empty:
                st.info("No sales data yet.")
                return

            d = df_all.copy()
            d["StartDate"] = pd.to_datetime(d["StartDate"], errors="coerce")
            d = d[d["StartDate"].notna()].copy()
            d["Year"] = d["StartDate"].dt.year.astype(int)
            d["MonthP"] = d["StartDate"].dt.to_period("M")

            years = sorted(d["Year"].unique().tolist())
            months = sorted(d["MonthP"].unique().tolist())
            month_labels = [m.to_timestamp().strftime("%B %Y") for m in months]
            label_to_period = dict(zip(month_labels, months))

            # --- Period selection ---
            period_mode = st.radio(
                "Period selection",
                options=["Full year (Year vs Year)", "Specific months (A vs B)"],
                index=0,
                horizontal=True,
                key="al_period_mode",
            )

            def _summarize_months(pers: list[pd.Period]) -> str:
                if not pers:
                    return "—"
                pers_sorted = sorted(pers)
                labels = [p.to_timestamp().strftime("%b %Y") for p in pers_sorted]
                if len(labels) == 1:
                    return labels[0]
                # If they look contiguous, show range; otherwise show count
                try:
                    diffs = [(pers_sorted[i+1] - pers_sorted[i]).n for i in range(len(pers_sorted)-1)]
                    if diffs and all(int(x) == 1 for x in diffs):
                        return f"{labels[0]}–{labels[-1]}"
                except Exception:
                    pass
                return f"{len(labels)} months"

            if period_mode.startswith("Full year"):
                c1, c2 = st.columns(2)
                with c1:
                    base_year = st.selectbox("Base Year", options=years, index=0, key="al_base")
                with c2:
                    comp_opts = [y for y in years if y != int(base_year)]
                    if not comp_opts:
                        st.warning("Only one year of data available. Add another year to compare, or use Specific months.")
                        comp_year = int(base_year)
                    else:
                        comp_year = st.selectbox("Comparison Year", options=comp_opts, index=0, key="al_comp")

                a = d[d["Year"] == int(base_year)].copy()
                b = d[d["Year"] == int(comp_year)].copy()
                label_a = str(base_year)
                label_b = str(comp_year)

            else:
                c1, c2 = st.columns(2)
                with c1:
                    a_pick = st.multiselect(
                        "Selection A months",
                        options=month_labels,
                        default=month_labels[-1:] if month_labels else [],
                        key="al_a_months",
                    )
                with c2:
                    b_pick = st.multiselect(
                        "Selection B months",
                        options=month_labels,
                        default=month_labels[-2:-1] if len(month_labels) >= 2 else [],
                        key="al_b_months",
                    )

                a_periods = [label_to_period[x] for x in a_pick if x in label_to_period]
                b_periods = [label_to_period[x] for x in b_pick if x in label_to_period]

                # Store for Top SKU movers table
                st.session_state["movers_a_periods"] = [str(p) for p in a_periods]
                st.session_state["movers_b_periods"] = [str(p) for p in b_periods]

                if not a_periods or not b_periods:
                    st.info("Pick at least one month in Selection A and Selection B to generate alerts.")
                    return

                a = d[d["MonthP"].isin(a_periods)].copy()
                b = d[d["MonthP"].isin(b_periods)].copy()
                label_a = _summarize_months(a_periods)
                label_b = _summarize_months(b_periods)

            basis = st.radio("Basis", options=["Sales", "Units"], index=0, horizontal=True, key="al_basis")
            value_col = "Sales" if basis == "Sales" else "Units"

            insights = []

            # Vendor deltas (worst 5)
            va = a.groupby("Vendor", as_index=False).agg(A=(value_col, "sum"))
            vb = b.groupby("Vendor", as_index=False).agg(B=(value_col, "sum"))
            v = va.merge(vb, on="Vendor", how="outer").fillna(0.0)
            v["Delta"] = v["A"] - v["B"]
            v = v.sort_values("Delta")

            def _fmt(vv):
                return fmt_currency(vv) if value_col == "Sales" else fmt_int(vv)

            for _, row in v.head(5).iterrows():
                if row["Delta"] < 0:
                    insights.append(f"🔻 Vendor **{row['Vendor']}** down {_fmt(row['Delta'])} ({label_a} → {label_b}).")

            # Retailer concentration warning (top 1 >= 40%)
            g = b.groupby("Retailer", as_index=False).agg(val=(value_col, "sum")).sort_values("val", ascending=False)
            total = float(g["val"].sum())
            if total > 0 and not g.empty:
                top1_share = float(g.iloc[0]["val"]) / total
                if top1_share >= 0.40:
                    insights.append(f"⚠️ Concentration risk: **{g.iloc[0]['Retailer']}** is {top1_share*100:.1f}% of {label_b} ({value_col}).")

            # Growth driven by few SKUs (top10 >= 60% of positive delta)
            sa = a.groupby("SKU", as_index=False).agg(A=(value_col, "sum"))
            sb = b.groupby("SKU", as_index=False).agg(B=(value_col, "sum"))
            sku = sa.merge(sb, on="SKU", how="outer").fillna(0.0)
            sku["Delta"] = sku["A"] - sku["B"]

            pos = sku[sku["Delta"] > 0].sort_values("Delta", ascending=False)
            if not pos.empty:
                top10 = float(pos.head(10)["Delta"].sum())
                total_pos = float(pos["Delta"].sum())
                share = (top10 / total_pos) if total_pos else 0.0
                if share >= 0.60:
                    insights.append(f"📈 Growth concentration: top 10 SKUs drive {share*100:.1f}% of positive change ({value_col}) ({label_a} → {label_b}).")

            # Lost SKUs count (had A but not B)
            lost = int(((sku["A"] > 0) & (sku["B"] == 0)).sum())
            if lost:
                insights.append(f"🧯 Lost SKUs: **{lost}** SKUs sold in {label_a} but not in {label_b}.")

            # Year locks notice
            locked = sorted(list(load_year_locks()))
            if locked:
                insights.append(f"🔒 Locked years: {', '.join(str(y) for y in locked)} (bulk ingest blocked).")

            if not insights:
                st.success("No major alerts detected with the current settings.")
            else:
                st.markdown("### Highlights")
                for s in insights:
                    st.markdown(f"- {s}")

            with st.expander("Details (tables)", expanded=False):
                st.markdown("**Worst vendors**")
                st.dataframe(
                    v.head(15).style.format({"A": _fmt, "B": _fmt, "Delta": _fmt})
                    .applymap(lambda v: f"color: {_color(v)};", subset=["Delta"]),
                    use_container_width=True,
                    hide_index=True
                )

                st.markdown("**Top SKU movers**")
                movers = sku.sort_values("Delta", ascending=False).head(15).copy()
                st.dataframe(
                    movers.style.format({"A": _fmt, "B": _fmt, "Delta": _fmt})
                    .applymap(lambda v: f"color: {_color(v)};", subset=["Delta"]),
                    use_container_width=True,
                    hide_index=True
                )


        # Run-Rate Forecast
        # -------------------------


    def render_no_sales():
            st.subheader("No Sales SKUs")
            weeks = st.selectbox("Timeframe (weeks)", options=[3,6,8,12], index=0, key="ns_weeks")
            retailers = sorted(vmap["Retailer"].dropna().unique().tolist())
            sel_r = st.selectbox("Retailer", options=["All"] + retailers, index=0, key="ns_retailer")

            if df.empty:
                st.info("No sales data yet.")
            else:
                d2 = df.copy()
                d2["StartDate"] = pd.to_datetime(d2["StartDate"], errors="coerce")
                periods = sorted(d2["StartDate"].dropna().dt.date.unique().tolist())
                use = periods[-weeks:] if len(periods) >= weeks else periods

                if not use:
                    st.info("No periods found yet.")
                else:
                    sold = d2[d2["StartDate"].dt.date.isin(use)].groupby(["Retailer","SKU"], as_index=False).agg(Units=("Units","sum"), Sales=("Sales","sum"))
                    ref = vmap[["Retailer","SKU","Vendor","MapOrder"]].copy()
                    if sel_r != "All":
                        ref = ref[ref["Retailer"] == sel_r].copy()

                    merged = ref.merge(sold, on=["Retailer","SKU"], how="left")
                    merged["Units"] = merged["Units"].fillna(0.0)
                    merged["Sales"] = merged["Sales"].fillna(0.0)

                    nos = merged[(merged["Units"] <= 0) & (merged["Sales"] <= 0)].copy()
                    nos["Status"] = f"No sales in last {weeks} weeks"
                    nos = nos.sort_values(["Retailer","MapOrder","SKU"], ascending=[True, True, True])

                    out = nos[["Retailer","Vendor","SKU","Status"]].copy()
                    st.dataframe(out, use_container_width=True, height=_table_height(out, max_px=1400), hide_index=True)


        # -------------------------
        # WoW Exceptions
        # -------------------------


    def keep_total_last(df_in: pd.DataFrame, label_col: str) -> pd.DataFrame:
        """After sorting, keep the TOTAL row (if present) pinned to the bottom."""
        try:
            if df_in is None or df_in.empty or label_col not in df_in.columns:
                return df_in
            m = df_in[label_col].astype(str).str.upper().eq("TOTAL")
            if not m.any():
                return df_in
            total = df_in.loc[m].copy()
            rest = df_in.loc[~m].copy()
            return pd.concat([rest, total], ignore_index=True)
        except Exception:
            return df_in


    def resolve_week_dates(periods: list, window):
        """
        periods: sorted list of datetime.date representing week start dates.
        window: int weeks or string like "6 months".
        Returns list of week dates to include, ordered ascending.
        """
        if not periods:
            return []
        if isinstance(window, int):
            return periods[-window:] if len(periods) >= window else periods
        if isinstance(window, str) and "month" in window:
            try:
                n = int(window.split()[0])
            except Exception:
                n = 6
            # get last n unique months present in periods
            months = [pd.Timestamp(d).to_period("M") for d in periods]
            uniq = []
            for p in months:
                if p not in uniq:
                    uniq.append(p)
            usem = uniq[-n:] if len(uniq) >= n else uniq
            use = [d for d in periods if pd.Timestamp(d).to_period("M") in usem]
            return use
        return periods


    def make_totals_tables(base: pd.DataFrame, group_col: str, tf_weeks, avg_weeks):
        if base.empty:
            return pd.DataFrame(), pd.DataFrame()
        base = base.copy()
        base["StartDate"] = pd.to_datetime(base["StartDate"], errors="coerce")
        periods = sorted(base["StartDate"].dropna().dt.date.unique().tolist())
        first_week = periods[0] if periods else None
        if not periods:
            return pd.DataFrame(), pd.DataFrame()

        use = resolve_week_dates(periods, tf_weeks)
        d = base[base["StartDate"].dt.date.isin(use)].copy()
        d["Week"] = d["StartDate"].dt.date

        sales_p = d.pivot_table(index=group_col, columns="Week", values="Sales", aggfunc="sum", fill_value=0.0).reindex(columns=use, fill_value=0.0)
        units_p = d.pivot_table(index=group_col, columns="Week", values="Units", aggfunc="sum", fill_value=0.0).reindex(columns=use, fill_value=0.0)

        if len(use) >= 2:
            sales_p["Diff"] = sales_p[use[-1]] - sales_p[use[-2]]
            units_p["Diff"] = units_p[use[-1]] - units_p[use[-2]]
        else:
            sales_p["Diff"] = 0.0
            units_p["Diff"] = 0.0

        # Determine which weeks to average based on selected average window
        current_year = int(pd.to_datetime(base["StartDate"], errors="coerce").dt.year.max() or date.today().year)

        # Choose avg weeks from ALL available periods in this filtered dataset (not just the displayed window)
        avg_use = resolve_avg_use(avg_weeks, periods, current_year)

        # Ignore the very first week present (often a partial week)
        if first_week is not None and avg_use:
            avg_use = [w for w in avg_use if pd.to_datetime(w, errors="coerce").date() != first_week]

        # Compute Avg from underlying data so month-year windows can work even if not currently displayed
        if avg_use:
            tmp = base[base["StartDate"].dt.date.isin(avg_use)].copy()
            tmp["Week"] = tmp["StartDate"].dt.date
            s_week = tmp.pivot_table(index=group_col, columns="Week", values="Sales", aggfunc="sum", fill_value=0.0)
            u_week = tmp.pivot_table(index=group_col, columns="Week", values="Units", aggfunc="sum", fill_value=0.0)
            sales_avg = s_week.replace(0, np.nan).mean(axis=1)
            units_avg = u_week.replace(0, np.nan).mean(axis=1)
            sales_p["Avg"] = sales_p.index.to_series().map(sales_avg).fillna(0.0)
            units_p["Avg"] = units_p.index.to_series().map(units_avg).fillna(0.0)
        else:
            sales_p["Avg"] = 0.0
            units_p["Avg"] = 0.0

        # Diff vs Avg uses the last week displayed minus Avg
        if use:
            sales_p["Diff vs Avg"] = sales_p[use[-1]] - sales_p["Avg"]
            units_p["Diff vs Avg"] = units_p[use[-1]] - units_p["Avg"]
        else:
            sales_p["Diff vs Avg"] = 0.0
            units_p["Diff vs Avg"] = 0.0

        sales_p = sales_p.sort_index()
        units_p = units_p.sort_index()

        sales_p.loc["TOTAL"] = sales_p.sum(axis=0)
        units_p.loc["TOTAL"] = units_p.sum(axis=0)

        # Recompute TOTAL Avg and Diff vs Avg from totals row values
        if "Avg" in sales_p.columns and use:
            # Avg already computed; just ensure TOTAL row is numeric
            try:
                sales_p.loc["TOTAL","Avg"] = float(sales_p.loc["TOTAL","Avg"])
                units_p.loc["TOTAL","Avg"] = float(units_p.loc["TOTAL","Avg"])
            except Exception:
                pass
            sales_p.loc["TOTAL","Diff vs Avg"] = sales_p.loc["TOTAL", use[-1]] - sales_p.loc["TOTAL","Avg"]
            units_p.loc["TOTAL","Diff vs Avg"] = units_p.loc["TOTAL", use[-1]] - units_p.loc["TOTAL","Avg"]

        def wlab(c):
            try:
                return pd.Timestamp(c).strftime("%m-%d")
            except Exception:
                return c

        sales_p = sales_p.rename(columns={c: wlab(c) for c in sales_p.columns})
        units_p = units_p.rename(columns={c: wlab(c) for c in units_p.columns})

        return sales_p.reset_index(), units_p.reset_index()

    # Retailer Totals

    # -------------------------
    # Tabs (top navigation)
    # -------------------------
    (tab_overview,
     tab_totals_dash,
     tab_momentum,
     tab_action_center,
     tab_top_skus,
     tab_exec,
     tab_comparisons,
     tab_wow_exc,
     tab_sku_intel,
     tab_forecasting,
     tab_year_summary,
     tab_alerts,
     tab_data_mgmt) = st.tabs([
        "Overview",
        "Totals Dashboards",
        "Momentum",
        "Action Center",
        "Top SKUs",
        "Executive Summary",
        "Comparisons",
        "WoW Exceptions",
        "SKU Intelligence",
        "Forecasting",
        "Year Summary",
        "Alerts",
        "Data Management",
    ])




    # -------------------------
    # Momentum + Action Center helpers
    # -------------------------

    def compute_momentum_scores(df_all: pd.DataFrame, window: int = 8) -> pd.DataFrame:
        """Return SKU momentum scores (0-100) using the last `window` weeks.

        Score blends:
        - Recent growth (last vs first in window)
        - Trend slope (linear fit)
        - Consistency (how often WoW is positive)

        Works off Units and Sales; final score is an average of the two.
        """
        if df_all_raw is None or df_all_raw.empty:
            return pd.DataFrame(columns=["SKU","Momentum","Momentum_Sales","Momentum_Units","Weeks","Lookback Weeks","Up Weeks","Down Weeks","Units_Last","Sales_Last"])

        d = df_all.copy()
        d["EndDate"] = pd.to_datetime(d["EndDate"], errors="coerce")
        d = d[d["EndDate"].notna()].copy()
        if d.empty:
            return pd.DataFrame(columns=["SKU","Momentum","Momentum_Sales","Momentum_Units","Weeks","Lookback Weeks","Up Weeks","Down Weeks","Units_Last","Sales_Last"])

        # Use last N distinct weeks
        weeks = sorted(d["EndDate"].unique())
        weeks = weeks[-window:]
        d = d[d["EndDate"].isin(weeks)].copy()

        sku_week = d.groupby(["SKU","EndDate"], as_index=False)[["Units","Sales"]].sum()

        def _score_series(vals: np.ndarray):
            vals = np.asarray(vals, dtype=float)
            vals = np.nan_to_num(vals, nan=0.0)
            if len(vals) < 2:
                return (0.0, 0.0, 0.0)
            x = np.arange(len(vals), dtype=float)
            # slope
            try:
                slope = np.polyfit(x, vals, 1)[0]
            except Exception:
                slope = 0.0
            # growth
            first = float(vals[0])
            last = float(vals[-1])
            growth = (last - first)
            # consistency: % of positive WoW changes
            diffs = np.diff(vals)
            if len(diffs) == 0:
                pos_rate = 0.0
            else:
                pos_rate = float((diffs > 0).mean())

            # Normalize components within plausible bounds
            # Use robust scaling by median absolute values across series later; here return raw tuple.
            return slope, growth, pos_rate

        rows = []
        for sku, g in sku_week.groupby("SKU"):
            g = g.sort_values("EndDate")
            # Series for momentum window
            s = pd.to_numeric(g['Sales'], errors='coerce').fillna(0.0).to_numpy(dtype=float)
            u = pd.to_numeric(g['Units'], errors='coerce').fillna(0.0).to_numpy(dtype=float)
            up_weeks = int((np.diff(s) > 0).sum()) if len(s) >= 2 else 0
            down_weeks = int((np.diff(s) < 0).sum()) if len(s) >= 2 else 0
            lookback_weeks = int(len(s))

            u = g["Units"].to_numpy(dtype=float)
            s = g["Sales"].to_numpy(dtype=float)
            su = _score_series(u)
            ss = _score_series(s)
            rows.append({
                "SKU": str(sku),
                "Weeks": len(g),
                "Lookback Weeks": int(len(s)),
                "Up Weeks": int((np.diff(s) > 0).sum()) if len(s) >= 2 else 0,
                "Down Weeks": int((np.diff(s) < 0).sum()) if len(s) >= 2 else 0,
                "Lookback Weeks": len(g),
                "Up Weeks": int((np.diff(s) > 0).sum()) if len(s) >= 2 else 0,
                "Down Weeks": int((np.diff(s) < 0).sum()) if len(s) >= 2 else 0,
                "_u_slope": su[0], "_u_growth": su[1], "_u_pos": su[2],
                "_s_slope": ss[0], "_s_growth": ss[1], "_s_pos": ss[2],
                "Units_Last": float(u[-1]) if len(u) else 0.0,
                "Sales_Last": float(s[-1]) if len(s) else 0.0,
            })

        out = pd.DataFrame(rows)
        if out.empty:
            return pd.DataFrame(columns=["SKU","Momentum","Momentum_Sales","Momentum_Units","Weeks","Lookback Weeks","Up Weeks","Down Weeks","Units_Last","Sales_Last"])

        def _robust_norm(col: str) -> pd.Series:
            v = out[col].astype(float)
            med = float(v.median())
            mad = float((v - med).abs().median())
            if mad == 0:
                return pd.Series(np.zeros(len(v)))
            z = (v - med) / mad
            # squash to 0..1 via logistic
            return 1.0 / (1.0 + np.exp(-z))

        u_slope = _robust_norm("_u_slope")
        u_growth = _robust_norm("_u_growth")
        s_slope = _robust_norm("_s_slope")
        s_growth = _robust_norm("_s_growth")

        # pos_rate already 0..1
        u_pos = out["_u_pos"].clip(0,1)
        s_pos = out["_s_pos"].clip(0,1)

        out["Momentum_Units"] = (0.40*u_growth + 0.35*u_slope + 0.25*u_pos) * 100
        out["Momentum_Sales"] = (0.40*s_growth + 0.35*s_slope + 0.25*s_pos) * 100
        out["Momentum"] = (out["Momentum_Units"] + out["Momentum_Sales"]) / 2.0

        return out[["SKU","Momentum","Momentum_Sales","Momentum_Units","Weeks","Units_Last","Sales_Last"]].sort_values("Momentum", ascending=False)


    def forecast_next_weeks(series: pd.Series, periods: int = 4) -> pd.DataFrame:
        """Simple, fast forecast: linear trend + moving average baseline.
        Returns df with columns: t, yhat_trend, yhat_ma, yhat (blend).
        """
        y = pd.to_numeric(series, errors="coerce").fillna(0.0).to_numpy(dtype=float)
        n = len(y)
        if n == 0:
            return pd.DataFrame({"t": list(range(1, periods+1)), "yhat": [0.0]*periods})

        x = np.arange(n, dtype=float)
        try:
            slope, intercept = np.polyfit(x, y, 1)
        except Exception:
            slope, intercept = 0.0, float(y.mean()) if n else 0.0

        x_future = np.arange(n, n+periods, dtype=float)
        yhat_trend = intercept + slope * x_future

        # Moving average baseline (last min(4,n) points)
        k = min(4, n)
        ma = float(np.mean(y[-k:]))
        yhat_ma = np.array([ma]*periods, dtype=float)

        # Blend 60% MA + 40% trend (stable)
        yhat = 0.6*yhat_ma + 0.4*yhat_trend

        yhat = np.clip(yhat, 0, None)
        return pd.DataFrame({"t": range(1, periods+1), "yhat_trend": yhat_trend, "yhat_ma": yhat_ma, "yhat": yhat})

    # BULLETPROOF_TABS



    def compute_momentum_table(df_all: pd.DataFrame, window: int = 12) -> pd.DataFrame:
        """
        Rebuilt momentum table:
        - Uses the last `window` weeks (based on EndDate global weeks) to compute:
          SKU | Momentum score (0-100) | Up Weeks | Down Weeks | Units Last | Sales Last
        - Momentum score is a rank-based composite of slope, growth, positive-week count, and current sales.
        """
        if df_all_raw is None or df_all_raw.empty:
            return pd.DataFrame(columns=["SKU","Momentum","Up Weeks","Down Weeks","Units_Last","Sales_Last"])

        df = df_all.copy()
        if "EndDate" not in df.columns or "SKU" not in df.columns:
            return pd.DataFrame(columns=["SKU","Momentum","Up Weeks","Down Weeks","Units_Last","Sales_Last"])

        df["EndDate"] = pd.to_datetime(df["EndDate"], errors="coerce")
        df = df[df["EndDate"].notna()].copy()
        if df.empty:
            return pd.DataFrame(columns=["SKU","Momentum","Up Weeks","Down Weeks","Units_Last","Sales_Last"])

        # Global last N weeks
        weeks = sorted(df["EndDate"].dropna().unique())
        if not weeks:
            return pd.DataFrame(columns=["SKU","Momentum","Up Weeks","Down Weeks","Units_Last","Sales_Last"])
        window = int(max(2, min(int(window), len(weeks))))
        use_weeks = weeks[-window:]
        sub = df[df["EndDate"].isin(use_weeks)].copy()

        # Aggregate weekly by SKU
        wk = sub.groupby(["SKU","EndDate"], as_index=False)[["Units","Sales"]].sum()
        wk = wk.sort_values(["SKU","EndDate"])

        rows = []
        for sku, g in wk.groupby("SKU", sort=False):
            g = g.sort_values("EndDate")
            s = pd.to_numeric(g["Sales"], errors="coerce").fillna(0.0).to_numpy(dtype=float)
            u = pd.to_numeric(g["Units"], errors="coerce").fillna(0.0).to_numpy(dtype=float)
            if len(s) == 0:
                continue

            diffs = (s[1:] - s[:-1]) if len(s) >= 2 else np.array([])
            up_weeks = int((diffs > 0).sum()) if diffs.size else 0
            down_weeks = int((diffs < 0).sum()) if diffs.size else 0

            # Trend features
            if len(s) >= 2:
                x = np.arange(len(s), dtype=float)
                # slope in $ per week
                try:
                    slope = float(np.polyfit(x, s, 1)[0])
                except Exception:
                    slope = float(s[-1] - s[0])
                growth = float((s[-1] - s[0]) / (abs(s[0]) + 1e-9))
            else:
                slope = 0.0
                growth = 0.0

            units_last = float(u[-1]) if len(u) else 0.0
            sales_last = float(s[-1]) if len(s) else 0.0

            rows.append({
                "SKU": str(sku),
                "_slope": slope,
                "_growth": growth,
                "_up": up_weeks,
                "_sales_last": sales_last,
                "Up Weeks": up_weeks,
                "Down Weeks": down_weeks,
                "Units_Last": units_last,
                "Sales_Last": sales_last,
            })

        out = pd.DataFrame(rows)
        if out.empty:
            return pd.DataFrame(columns=["SKU","Momentum","Up Weeks","Down Weeks","Units_Last","Sales_Last"])

        # Rank-based momentum score (0-100) so it is stable and comparable across SKUs each run
        def _pct_rank(series: pd.Series) -> pd.Series:
            try:
                return series.rank(pct=True, method="average").fillna(0.0)
            except Exception:
                return pd.Series([0.0]*len(series), index=series.index)

        slope_p = _pct_rank(out["_slope"])
        growth_p = _pct_rank(out["_growth"])
        up_p = _pct_rank(out["_up"])
        sales_p = _pct_rank(out["_sales_last"])

        score = 40.0 * (0.40*slope_p + 0.30*growth_p + 0.20*up_p + 0.10*sales_p)
        out["Momentum"] = score.round(0).astype(int)

        # Final columns
        final = out[["SKU","Momentum","Up Weeks","Down Weeks","Units_Last","Sales_Last"]].copy()
        final = final.sort_values(["Momentum","Sales_Last"], ascending=[False, False]).reset_index(drop=True)
        return final



    def _get_current_and_prev_week(df: pd.DataFrame):
        if df is None or df.empty:
            return None, None, None, None
        d = df.copy()
        d["StartDate"] = pd.to_datetime(d["StartDate"], errors="coerce")
        d["EndDate"] = pd.to_datetime(d["EndDate"], errors="coerce")
        d = d[d["StartDate"].notna() & d["EndDate"].notna()].copy()
        if d.empty:
            return None, None, None, None
        ends = sorted(d["EndDate"].dropna().unique())
        cur_end = ends[-1]
        prev_end = ends[-2] if len(ends) >= 2 else None
        # Use the most common start date for that end date (handles any stray rows)
        cur_start = d.loc[d["EndDate"] == cur_end, "StartDate"].mode()
        cur_start = cur_start.iloc[0] if not cur_start.empty else d["StartDate"].min()
        prev_start = None
        if prev_end is not None:
            prev_start_s = d.loc[d["EndDate"] == prev_end, "StartDate"].mode()
            prev_start = prev_start_s.iloc[0] if not prev_start_s.empty else None
        cur = d[d["EndDate"] == cur_end].copy()
        prev = d[d["EndDate"] == prev_end].copy() if prev_end is not None else d.iloc[0:0].copy()
        return cur, prev, cur_start, cur_end


    def _fmt_week_range(start_dt, end_dt) -> str:
        try:
            s = pd.to_datetime(start_dt).date()
            e = pd.to_datetime(end_dt).date()
            return f"{s.strftime('%b %d, %Y')} – {e.strftime('%b %d, %Y')}"
        except Exception:
            return "Current week"


    def render_tab_overview():
        with tab_overview:
            st.subheader("Overview")

            if df_all_raw is None or df_all_raw.empty:
                st.info("No sales data yet. Upload your sales_store.csv in Data Management.")
                return

            cur, prev, cur_start, cur_end = _get_current_and_prev_week(df_all_raw)
            if cur is None or cur.empty:
                st.info("No valid weekly rows found (missing StartDate/EndDate).")
                return

            st.caption(f"Week: {_fmt_week_range(cur_start, cur_end)}")

            # What changed? (auto-summary)
            total_sales_cur = float(cur["Sales"].sum())
            total_sales_prev = float(prev["Sales"].sum()) if (prev is not None and not prev.empty) else 0.0
            wow_sales_total = total_sales_cur - total_sales_prev

            # Drivers by Retailer & Vendor (for narrative)
            def _wow_by_key(cur_df, prev_df, key):
                c = cur_df.groupby(key, as_index=False)[["Sales","Units"]].sum()
                p = prev_df.groupby(key, as_index=False)[["Sales","Units"]].sum() if (prev_df is not None and not prev_df.empty) else pd.DataFrame(columns=[key,"Sales","Units"])
                m2 = c.merge(p[[key,"Sales"]].rename(columns={"Sales":"Sales_prev"}), on=key, how="left")
                m2["Sales_prev"] = m2["Sales_prev"].fillna(0.0)
                m2["WoW Sales"] = m2["Sales"] - m2["Sales_prev"]
                # hide entities that have no sales in the selected week (helps prevent 0/0 rows)
                m2 = m2[(pd.to_numeric(m2["Sales"], errors="coerce").fillna(0) > 0) | (pd.to_numeric(m2["Units"], errors="coerce").fillna(0) > 0)].copy()
                return m2.sort_values("WoW Sales")

            wow_r = _wow_by_key(cur, prev, "Retailer") if "Retailer" in cur.columns else pd.DataFrame()
            wow_v = _wow_by_key(cur, prev, "Vendor") if "Vendor" in cur.columns else pd.DataFrame()

            def _money2(v):
                try:
                    v = float(v)
                except Exception:
                    return "—"
                # Always show 2 decimals and standard negative currency style
                return f"-${abs(v):,.2f}" if v < 0 else f"${v:,.2f}"

            direction = "up" if wow_sales_total > 0 else ("down" if wow_sales_total < 0 else "flat")
            top_r_down = wow_r.head(1) if (wow_r is not None and not wow_r.empty) else pd.DataFrame()
            top_r_up   = wow_r.tail(1) if (wow_r is not None and not wow_r.empty) else pd.DataFrame()
            top_v_down = wow_v.head(1) if (wow_v is not None and not wow_v.empty) else pd.DataFrame()
            top_v_up   = wow_v.tail(1) if (wow_v is not None and not wow_v.empty) else pd.DataFrame()

            with st.expander("🧠 What changed? (auto-summary)", expanded=True):
                st.write(f"WoW sales were **{direction}** {_money2(wow_sales_total)}.")
                if not top_r_down.empty and float(top_r_down["WoW Sales"].iloc[0]) < 0:
                    st.write(f"• Biggest retailer headwind: {str(top_r_down['Retailer'].iloc[0])} ({_money2(top_r_down['WoW Sales'].iloc[0])})")
                if not top_v_down.empty and float(top_v_down["WoW Sales"].iloc[0]) < 0:
                    st.write(f"• Biggest vendor headwind: {str(top_v_down['Vendor'].iloc[0])} ({_money2(top_v_down['WoW Sales'].iloc[0])})")
                if not top_r_up.empty and float(top_r_up["WoW Sales"].iloc[0]) > 0:
                    st.write(f"• Top retailer offset: {str(top_r_up['Retailer'].iloc[0])} ({_money2(top_r_up['WoW Sales'].iloc[0])})")
                if not top_v_up.empty and float(top_v_up["WoW Sales"].iloc[0]) > 0:
                    st.write(f"• Top vendor offset: {str(top_v_up['Vendor'].iloc[0])} ({_money2(top_v_up['WoW Sales'].iloc[0])})")


            # Aggregate current + previous
            cur_r = cur.groupby("Retailer", as_index=False)[["Units","Sales"]].sum()
            prev_r = prev.groupby("Retailer", as_index=False)[["Units","Sales"]].sum() if prev is not None and not prev.empty else pd.DataFrame(columns=["Retailer","Units","Sales"])

            cur_v = cur.groupby("Vendor", as_index=False)[["Units","Sales"]].sum()
            prev_v = prev.groupby("Vendor", as_index=False)[["Units","Sales"]].sum() if prev is not None and not prev.empty else pd.DataFrame(columns=["Vendor","Units","Sales"])

            cur_s = cur.groupby("SKU", as_index=False)[["Units","Sales"]].sum()
            prev_s = prev.groupby("SKU", as_index=False)[["Units","Sales"]].sum() if prev is not None and not prev.empty else pd.DataFrame(columns=["SKU","Units","Sales"])

            # Helpers
            def _delta(cur_df, prev_df, key, val_col, key_val):
                try:
                    c = float(cur_df.loc[cur_df[key]==key_val, val_col].sum())
                except Exception:
                    c = 0.0
                try:
                    p = float(prev_df.loc[prev_df[key]==key_val, val_col].sum()) if prev_df is not None and not prev_df.empty else 0.0
                except Exception:
                    p = 0.0
                return c, (c - p)

            def _fmt_currency(x):
                """Currency formatter where negatives start with "-" so st.metric colors correctly."""
                try:
                    v = float(x)
                except Exception:
                    return ""
                return f"-${abs(v):,.2f}" if v < 0 else f"${v:,.2f}"

            def _fmt_int(x):
                try:
                    return f"{int(round(float(x))):,}"
                except Exception:
                    return ""

            def _posneg_color(v):
                try:
                    v = float(v)
                except Exception:
                    return ""
                if v > 0:
                    return "color: #0f7b0f; font-weight: 600;"
                if v < 0:
                    return "color: #b00020; font-weight: 600;"
                return "color: #666;"

            def _top2_cards(label, metric_col):
                is_currency = (metric_col == "Sales")

                top_retailers = cur_r.sort_values([metric_col, ("Units" if metric_col=="Sales" else "Sales")], ascending=False).head(2)
                top_vendors   = cur_v.sort_values([metric_col, ("Units" if metric_col=="Sales" else "Sales")], ascending=False).head(2)
                top_skus      = cur_s.sort_values([metric_col, ("Units" if metric_col=="Sales" else "Sales")], ascending=False).head(2)

                c1, c2, c3 = st.columns(3)

                def _render_two(col, title, key_name, df_top, prev_df):
                    col.markdown(f"**{title} ({label})**")
                    if df_top.empty:
                        col.caption("—")
                        return
                    for i, row in enumerate(df_top.itertuples(index=False), start=1):
                        name = getattr(row, key_name)
                        val, dlt = _delta(df_top, prev_df, key_name, metric_col, name)  # delta helper expects same schema, ok for val
                        # IMPORTANT: delta should be computed from full aggregates, not top slices
                        val, dlt = _delta({"Retailer":cur_r,"Vendor":cur_v,"SKU":cur_s}[key_name],
                                          {"Retailer":prev_r,"Vendor":prev_v,"SKU":prev_s}[key_name],
                                          key_name, metric_col, name)

                        value_str = _fmt_currency(val) if is_currency else _fmt_int(val)
                        delta_str = _fmt_currency(dlt) if is_currency else _fmt_int(dlt)
                        col.metric(f"#{i}: {name}", value_str, delta_str)

                # Render each category with two metrics
                _render_two(c1, "Top Retailers", "Retailer", top_retailers, prev_r)
                _render_two(c2, "Top Vendors", "Vendor", top_vendors, prev_v)
                # Ensure SKU shown as string
                top_skus = top_skus.copy()
                top_skus["SKU"] = top_skus["SKU"].astype(str)
                _render_two(c3, "Top SKUs", "SKU", top_skus, prev_s)

            # Top 2 (Sales) + Top 2 (Units)
            _top2_cards("Sales", "Sales")
            _top2_cards("Units", "Units")

            st.divider()

            # Biggest movers (Top 10 by absolute WoW change)
            movers = cur_s.merge(prev_s, on="SKU", how="left", suffixes=("_cur","_prev"))
            movers["Sales_prev"] = movers["Sales_prev"].fillna(0.0)
            movers["Units_prev"] = movers["Units_prev"].fillna(0.0)
            movers["WoW Sales"] = movers["Sales_cur"] - movers["Sales_prev"]
            movers["WoW Units"] = movers["Units_cur"] - movers["Units_prev"]

            m_sales = movers.assign(_abs=movers["WoW Sales"].abs()).sort_values("_abs", ascending=False).head(10).drop(columns=["_abs"])
            m_units = movers.assign(_abs=movers["WoW Units"].abs()).sort_values("_abs", ascending=False).head(10).drop(columns=["_abs"])

            st.markdown("### Biggest movers this week")
            t1, t2 = st.tabs(["By Sales ($)", "By Units"])

            def _render_movers(df_m):
                if df_m.empty:
                    st.info("No movers found (need at least 2 weeks of data).")
                    return
                show = df_m[[
                    "SKU",
                    "Sales_cur","Sales_prev","WoW Sales",
                    "Units_cur","Units_prev","WoW Units"
                ]].rename(columns={
                    "Sales_cur":"Sales (This Week)",
                    "Sales_prev":"Sales (Prev Week)",
                    "Units_cur":"Units (This Week)",
                    "Units_prev":"Units (Prev Week)",
                }).copy()

                sty = show.style.format({
                    "Sales (This Week)": _fmt_currency,
                    "Sales (Prev Week)": _fmt_currency,
                    "WoW Sales": _fmt_currency,
                    "Units (This Week)": _fmt_int,
                    "Units (Prev Week)": _fmt_int,
                    "WoW Units": _fmt_int,
                }).applymap(_posneg_color, subset=["WoW Sales","WoW Units"])

                st.dataframe(sty, use_container_width=True, hide_index=True)

            with t1:
                _render_movers(m_sales)
            with t2:
                _render_movers(m_units)

            st.divider()

            c4, c5 = st.columns(2)

            # New SKU detected (only show when the SKU has a NEW SALE this week)
            prior = df_all.copy()
            prior["EndDate"] = pd.to_datetime(prior["EndDate"], errors="coerce")
            prior = prior[prior["EndDate"].notna()].copy()
            prior = prior[prior["EndDate"] < cur_end]

            # Treat "sale" as Units > 0 OR Sales > 0. This avoids flagging SKUs that merely appear on a list/map with zero sales.
            cur_sold = cur.copy()
            cur_sold["Units"] = pd.to_numeric(cur_sold.get("Units"), errors="coerce").fillna(0.0)
            cur_sold["Sales"] = pd.to_numeric(cur_sold.get("Sales"), errors="coerce").fillna(0.0)
            cur_sold = cur_sold[(cur_sold["Units"] > 0) | (cur_sold["Sales"] > 0)].copy()

            prior_sold = prior.copy()
            prior_sold["Units"] = pd.to_numeric(prior_sold.get("Units"), errors="coerce").fillna(0.0)
            prior_sold["Sales"] = pd.to_numeric(prior_sold.get("Sales"), errors="coerce").fillna(0.0)
            prior_sold = prior_sold[(prior_sold["Units"] > 0) | (prior_sold["Sales"] > 0)].copy()

            sold_skus_cur = set(cur_sold["SKU"].astype(str))
            sold_skus_prior = set(prior_sold["SKU"].astype(str))
            new_sale_skus = sorted(sold_skus_cur - sold_skus_prior)

            c4.markdown("### New SKU sales detected")
            if not new_sale_skus:
                c4.success("No new-SKU sales this week (nothing sold for the first time).")
            else:
                # Show Units + Sales for the current week for each new-sale SKU
                show = (cur_sold[cur_sold["SKU"].astype(str).isin(new_sale_skus)]
                        .groupby("SKU", as_index=False)[["Units", "Sales"]].sum())

                # Friendly formatting
                show["Units"] = show["Units"].round(0).astype(int)
                show["Sales"] = show["Sales"].apply(_fmt_currency)

                c4.warning(f"{len(new_sale_skus)} SKU(s) recorded their first sale this week.")
                c4.dataframe(show.rename(columns={"SKU": "New Sale SKU"}), hide_index=True, use_container_width=True)

            # Declining alerts
            c5.markdown("### Declining alerts")
            if prev is None or prev.empty:
                c5.info("Need at least 2 weeks of data to calculate declines.")
            else:
                # --- Vendor declines ---
                c5.markdown("**Declining vendor alert**")
                vcmp = cur_v.merge(prev_v, on="Vendor", how="left", suffixes=("_cur","_prev"))
                vcmp["Sales_prev"] = vcmp["Sales_prev"].fillna(0.0)
                vcmp["Units_prev"] = vcmp["Units_prev"].fillna(0.0)
                vcmp["ΔSales"] = vcmp["Sales_cur"] - vcmp["Sales_prev"]
                vcmp["ΔUnits"] = vcmp["Units_cur"] - vcmp["Units_prev"]
                vcmp["%ΔSales"] = np.where(vcmp["Sales_prev"] > 0, vcmp["ΔSales"] / vcmp["Sales_prev"], np.nan)
                vcmp["%ΔUnits"] = np.where(vcmp["Units_prev"] > 0, vcmp["ΔUnits"] / vcmp["Units_prev"], np.nan)

                min_prev_sales = 1000.0
                alerts_v = vcmp[(vcmp["Sales_prev"] >= min_prev_sales) & (vcmp["%ΔSales"] <= -0.25)].copy().sort_values("%ΔSales").head(8)
                if alerts_v.empty:
                    c5.success("No major vendor sales declines this week.")
                else:
                    c5.warning(f"{len(alerts_v)} vendor(s) down 25%+ WoW (Sales).")
                    show_v = alerts_v[["Vendor","Sales_cur","Sales_prev","ΔSales","Units_cur","Units_prev","ΔUnits"]].rename(columns={
                        "Sales_cur":"Sales (This Week)",
                        "Sales_prev":"Sales (Prev Week)",
                        "ΔSales":"WoW Sales",
                        "Units_cur":"Units (This Week)",
                        "Units_prev":"Units (Prev Week)",
                        "ΔUnits":"WoW Units",
                    }).copy()
                    sty_v = show_v.style.format({
                        "Sales (This Week)": _fmt_currency,
                        "Sales (Prev Week)": _fmt_currency,
                        "WoW Sales": _fmt_currency,
                        "Units (This Week)": _fmt_int,
                        "Units (Prev Week)": _fmt_int,
                        "WoW Units": _fmt_int,
                    }).applymap(_posneg_color, subset=["WoW Sales","WoW Units"])
                    c5.dataframe(sty_v, hide_index=True, use_container_width=True)

                c5.divider()

                # --- Retailer declines ---
                c5.markdown("**Declining retailer alert**")
                rcmp = cur_r.merge(prev_r, on="Retailer", how="left", suffixes=("_cur","_prev"))
                rcmp["Sales_prev"] = rcmp["Sales_prev"].fillna(0.0)
                rcmp["Units_prev"] = rcmp["Units_prev"].fillna(0.0)
                rcmp["ΔSales"] = rcmp["Sales_cur"] - rcmp["Sales_prev"]
                rcmp["ΔUnits"] = rcmp["Units_cur"] - rcmp["Units_prev"]
                rcmp["%ΔSales"] = np.where(rcmp["Sales_prev"] > 0, rcmp["ΔSales"] / rcmp["Sales_prev"], np.nan)

                min_prev_sales_r = 5000.0  # retailers are usually larger; avoids noise
                alerts_r = rcmp[(rcmp["Sales_prev"] >= min_prev_sales_r) & (rcmp["%ΔSales"] <= -0.25)].copy().sort_values("%ΔSales").head(8)

                if alerts_r.empty:
                    c5.success("No major retailer sales declines this week.")
                else:
                    c5.warning(f"{len(alerts_r)} retailer(s) down 25%+ WoW (Sales).")
                    show_r = alerts_r[["Retailer","Sales_cur","Sales_prev","ΔSales","Units_cur","Units_prev","ΔUnits"]].rename(columns={
                        "Sales_cur":"Sales (This Week)",
                        "Sales_prev":"Sales (Prev Week)",
                        "ΔSales":"WoW Sales",
                        "Units_cur":"Units (This Week)",
                        "Units_prev":"Units (Prev Week)",
                        "ΔUnits":"WoW Units",
                    }).copy()
                    sty_r = show_r.style.format({
                        "Sales (This Week)": _fmt_currency,
                        "Sales (Prev Week)": _fmt_currency,
                        "WoW Sales": _fmt_currency,
                        "Units (This Week)": _fmt_int,
                        "Units (Prev Week)": _fmt_int,
                        "WoW Units": _fmt_int,
                    }).applymap(_posneg_color, subset=["WoW Sales","WoW Units"])
                    c5.dataframe(sty_r, hide_index=True, use_container_width=True)




    def render_tab_action_center():
        with tab_action_center:
            st.subheader("Action Center (Rebuilt)")
            st.caption("This page is designed to answer: What do we need to DO this week? (Rule-based thresholds)")

            if df_all_raw is None or df_all_raw.empty:
                st.info("No sales data yet.")
                return

            cur, prev, cur_start, cur_end = _get_current_and_prev_week(df_all_raw)
            if cur is None or cur.empty:
                st.info("No valid weekly rows found (missing StartDate/EndDate).")
                return

            # -------- Thresholds (Option 1) --------
            retailer_min_prev = 5000.0
            vendor_min_prev   = 2000.0
            sku_min_prev      = 500.0
            decline_pct       = -0.25   # -25% or worse
            growth_pct        = 0.30    # +30% or better

            new_window_weeks  = 8
            test_max_weeks    = 10
            test_min_pos_wow  = 3

            st.caption(f"Week: {_fmt_week_range(cur_start, cur_end)}")

            # Weekly aggregates
            cur_r = cur.groupby("Retailer", as_index=False)[["Units","Sales"]].sum()
            cur_v = cur.groupby("Vendor",   as_index=False)[["Units","Sales"]].sum()
            cur_s = cur.groupby("SKU",      as_index=False)[["Units","Sales"]].sum()

            prev_r = prev.groupby("Retailer", as_index=False)[["Units","Sales"]].sum() if prev is not None and not prev.empty else pd.DataFrame(columns=["Retailer","Units","Sales"])
            prev_v = prev.groupby("Vendor",   as_index=False)[["Units","Sales"]].sum() if prev is not None and not prev.empty else pd.DataFrame(columns=["Vendor","Units","Sales"])
            prev_s = prev.groupby("SKU",      as_index=False)[["Units","Sales"]].sum() if prev is not None and not prev.empty else pd.DataFrame(columns=["SKU","Units","Sales"])

            def money(x):
                try: return f"${float(x):,.2f}"
                except: return ""
            def pct(x):
                try: return f"{float(x)*100:,.1f}%"
                except: return ""
            def intfmt(x):
                try: return f"{int(round(float(x))):,}"
                except: return ""

            # =======================
            # 1) CRITICAL DECLINES
            # =======================
            st.markdown("## 🔴 Critical Declines")

            colA, colB = st.columns(2)

            # Retailer declines
            with colA:
                st.markdown("### Retailer declines (Sales)")
                if prev_r.empty:
                    st.info("Need at least 2 weeks of data to compute WoW declines.")
                else:
                    r = cur_r.merge(prev_r, on="Retailer", how="left", suffixes=("_cur","_prev"))
                    r["Sales_prev"] = r["Sales_prev"].fillna(0.0)
                    r["WoW Sales"]  = r["Sales_cur"] - r["Sales_prev"]
                    r["%ΔSales"]    = np.where(r["Sales_prev"] > 0, r["WoW Sales"] / r["Sales_prev"], np.nan)

                    r_alert = r[(r["Sales_prev"] >= retailer_min_prev) & (r["%ΔSales"] <= decline_pct)].copy().sort_values("%ΔSales").head(10)

                    if r_alert.empty:
                        st.success("No retailer met the decline thresholds this week.")
                    else:
                        # SKU drivers per retailer (top 3 ΔSales)
                        cur_rs = cur.groupby(["Retailer","SKU"], as_index=False)[["Sales"]].sum()
                        prev_rs = prev.groupby(["Retailer","SKU"], as_index=False)[["Sales"]].sum() if prev is not None and not prev.empty else pd.DataFrame(columns=["Retailer","SKU","Sales"])
                        rs = cur_rs.merge(prev_rs, on=["Retailer","SKU"], how="left", suffixes=("_cur","_prev"))
                        rs["Sales_prev"] = rs["Sales_prev"].fillna(0.0)
                        rs["ΔSales"] = rs["Sales_cur"] - rs["Sales_prev"]

                        drivers=[]
                        for rr in r_alert["Retailer"].astype(str).tolist():
                            sub = rs[rs["Retailer"].astype(str)==rr].copy().sort_values("ΔSales").head(3)
                            drivers.append(", ".join([f"{str(x.SKU)} ({money(x.ΔSales)})" for x in sub.itertuples(index=False)]))
                        r_alert["Top SKU drivers (ΔSales)"] = drivers

                        show = r_alert[["Retailer","Sales_cur","Sales_prev","WoW Sales","%ΔSales","Top SKU drivers (ΔSales)"]].rename(columns={
                            "Sales_cur":"Sales (This Week)",
                            "Sales_prev":"Sales (Prev Week)",
                        }).copy()
                        show["Sales (This Week)"] = show["Sales (This Week)"].map(money)
                        show["Sales (Prev Week)"] = show["Sales (Prev Week)"].map(money)
                        show["WoW Sales"] = show["WoW Sales"].map(money)
                        show["%ΔSales"] = show["%ΔSales"].map(pct)
                        st.dataframe(show, use_container_width=True, hide_index=True)

            # Vendor declines
            with colB:
                st.markdown("### Vendor declines (Sales)")
                if prev_v.empty:
                    st.info("Need at least 2 weeks of data to compute WoW declines.")
                else:
                    v = cur_v.merge(prev_v, on="Vendor", how="left", suffixes=("_cur","_prev"))
                    v["Sales_prev"] = v["Sales_prev"].fillna(0.0)
                    v["WoW Sales"]  = v["Sales_cur"] - v["Sales_prev"]
                    v["%ΔSales"]    = np.where(v["Sales_prev"] > 0, v["WoW Sales"] / v["Sales_prev"], np.nan)

                    v_alert = v[(v["Sales_prev"] >= vendor_min_prev) & (v["%ΔSales"] <= decline_pct)].copy().sort_values("%ΔSales").head(10)

                    if v_alert.empty:
                        st.success("No vendor met the decline thresholds this week.")
                    else:
                        # Retailer impacts per vendor (top 3 ΔSales)
                        cur_vr = cur.groupby(["Vendor","Retailer"], as_index=False)[["Sales"]].sum()
                        prev_vr = prev.groupby(["Vendor","Retailer"], as_index=False)[["Sales"]].sum() if prev is not None and not prev.empty else pd.DataFrame(columns=["Vendor","Retailer","Sales"])
                        vr = cur_vr.merge(prev_vr, on=["Vendor","Retailer"], how="left", suffixes=("_cur","_prev"))
                        vr["Sales_prev"] = vr["Sales_prev"].fillna(0.0)
                        vr["ΔSales"] = vr["Sales_cur"] - vr["Sales_prev"]

                        impacts=[]
                        for vv in v_alert["Vendor"].astype(str).tolist():
                            sub = vr[vr["Vendor"].astype(str)==vv].copy().sort_values("ΔSales").head(3)
                            impacts.append(", ".join([f"{str(x.Retailer)} ({money(x.ΔSales)})" for x in sub.itertuples(index=False)]))
                        v_alert["Retailer impacts (ΔSales)"] = impacts

                        show = v_alert[["Vendor","Sales_cur","Sales_prev","WoW Sales","%ΔSales","Retailer impacts (ΔSales)"]].rename(columns={
                            "Sales_cur":"Sales (This Week)",
                            "Sales_prev":"Sales (Prev Week)",
                        }).copy()
                        show["Sales (This Week)"] = show["Sales (This Week)"].map(money)
                        show["Sales (Prev Week)"] = show["Sales (Prev Week)"].map(money)
                        show["WoW Sales"] = show["WoW Sales"].map(money)
                        show["%ΔSales"] = show["%ΔSales"].map(pct)
                        st.dataframe(show, use_container_width=True, hide_index=True)

            # =======================
            # 2) OPPORTUNITIES (FAST GROWERS)
            # =======================
            st.markdown("## 🟢 Opportunities")
            st.markdown("### Fast-growing SKUs (Sales)")

            if prev_s.empty:
                st.info("Need at least 2 weeks of data to compute WoW growth.")
                growers = pd.DataFrame()
            else:
                s = cur_s.merge(prev_s, on="SKU", how="left", suffixes=("_cur","_prev"))
                s["Sales_prev"] = s["Sales_prev"].fillna(0.0)
                s["WoW Sales"]  = s["Sales_cur"] - s["Sales_prev"]
                s["%ΔSales"]    = np.where(s["Sales_prev"] > 0, s["WoW Sales"] / s["Sales_prev"], np.nan)

                growers = s[(s["Sales_prev"] >= sku_min_prev) & (s["%ΔSales"] >= growth_pct)].copy().sort_values("%ΔSales", ascending=False).head(15)

                if growers.empty:
                    st.info("No SKUs met the growth thresholds this week.")
                else:
                    cur_drv = cur.groupby(["SKU","Retailer","Vendor"], as_index=False)[["Sales"]].sum()
                    driver=[]
                    for sku in growers["SKU"].astype(str).tolist():
                        sub = cur_drv[cur_drv["SKU"].astype(str)==sku].sort_values("Sales", ascending=False).head(1)
                        driver.append(f"{sub.iloc[0]['Retailer']} / {sub.iloc[0]['Vendor']}" if not sub.empty else "")
                    growers["Primary driver"] = driver

                    show = growers[["SKU","Primary driver","Sales_cur","Sales_prev","WoW Sales","%ΔSales","Units_cur","Units_prev"]].rename(columns={
                        "Sales_cur":"Sales (This Week)",
                        "Sales_prev":"Sales (Prev Week)",
                        "Units_cur":"Units (This Week)",
                        "Units_prev":"Units (Prev Week)",
                    }).copy()
                    show["Sales (This Week)"] = show["Sales (This Week)"].map(money)
                    show["Sales (Prev Week)"] = show["Sales (Prev Week)"].map(money)
                    show["WoW Sales"] = show["WoW Sales"].map(money)
                    show["%ΔSales"] = show["%ΔSales"].map(pct)
                    show["Units (This Week)"] = show["Units (This Week)"].map(intfmt)
                    show["Units (Prev Week)"] = show["Units (Prev Week)"].map(intfmt)
                    st.dataframe(show, use_container_width=True, hide_index=True)

            # =======================
            # 3) NEW SKU PERFORMANCE + TEST SIGNALS
            # =======================
            st.markdown("## 🟡 New SKU performance & test signals")

            hist = df_all.copy()
            hist["EndDate"] = pd.to_datetime(hist["EndDate"], errors="coerce")
            hist = hist[hist["EndDate"].notna()].copy()

            cur_end_ts = pd.to_datetime(cur_end)

            sku_first_seen = hist.groupby(hist["SKU"].astype(str))["EndDate"].min()

            # positive-sales rows (used to detect "first sale ever")
            hist_pos = hist[(pd.to_numeric(hist.get("Sales"), errors="coerce").fillna(0) > 0) | (pd.to_numeric(hist.get("Units"), errors="coerce").fillna(0) > 0)].copy()
            sku_first_sales = hist_pos.groupby(hist_pos["SKU"].astype(str))["EndDate"].min() if not hist_pos.empty else pd.Series(dtype="datetime64[ns]")

            # NEW: first-sale by placement (SKU + Retailer)
            if not hist_pos.empty and "Retailer" in hist_pos.columns:
                hist_pos["SKU"] = hist_pos["SKU"].astype(str)
                hist_pos["Retailer"] = hist_pos["Retailer"].astype(str)
                pair_first_sales = hist_pos.groupby(["SKU","Retailer"])["EndDate"].min()
            else:
                pair_first_sales = pd.Series(dtype="datetime64[ns]")

            cur_skus = set(cur["SKU"].astype(str).unique().tolist())

            true_new = []
            activated = []
            for sku in sorted(cur_skus):
                fs = sku_first_seen.get(sku, pd.NaT)
                fz = sku_first_sales.get(sku, pd.NaT)
                if pd.isna(fs):
                    continue
                if fs == cur_end_ts:
                    true_new.append(sku)
                else:
                    # Activated = existed previously, but first positive sale happens this week
                    if (not pd.isna(fz)) and (fz == cur_end_ts) and (fs < cur_end_ts):
                        activated.append(sku)

            # NEW: "new somewhere" placements — SKU sold at a Retailer for the first time this week
            new_place_rows = []
            if "Retailer" in cur.columns:
                cur_pos = cur[(pd.to_numeric(cur.get("Sales"), errors="coerce").fillna(0) > 0) | (pd.to_numeric(cur.get("Units"), errors="coerce").fillna(0) > 0)].copy()
                cur_pos["SKU"] = cur_pos["SKU"].astype(str)
                cur_pos["Retailer"] = cur_pos["Retailer"].astype(str)

                # aggregate current week by SKU + Retailer (+ Vendor if present)
                gb_cols = ["SKU","Retailer"] + (["Vendor"] if "Vendor" in cur_pos.columns else [])
                cur_pair = cur_pos.groupby(gb_cols, as_index=False)[["Units","Sales"]].sum()

                for _, r in cur_pair.iterrows():
                    sku = str(r["SKU"])
                    ret = str(r["Retailer"])
                    fpair = pair_first_sales.get((sku, ret), pd.NaT)
                    if (not pd.isna(fpair)) and (fpair == cur_end_ts):
                        new_place_rows.append(r.to_dict())


            left, right = st.columns(2)
            with left:
                st.markdown("### New SKUs / New Placements this week")

                if not true_new and not activated and not new_place_rows:
                    st.info("No true-new SKUs, activated SKUs, or new placements this week.")
                else:
                    if true_new:
                        st.success(f"True new (first time ever in data): {len(true_new)}")
                        st.dataframe(pd.DataFrame({"True New SKUs": true_new}), hide_index=True, use_container_width=True)
                    if activated:
                        st.warning(f"Activated (first positive sale this week): {len(activated)}")
                        st.dataframe(pd.DataFrame({"Activated SKUs": activated}), hide_index=True, use_container_width=True)

                    if new_place_rows:
                        df_np = pd.DataFrame(new_place_rows)
                        # Normalize columns and show where it sold
                        cols = []
                        for c in ["SKU","Vendor","Retailer","Units","Sales"]:
                            if c in df_np.columns:
                                cols.append(c)
                        df_np = df_np[cols].copy()
                        if "Units" in df_np.columns:
                            df_np["Units"] = pd.to_numeric(df_np["Units"], errors="coerce").fillna(0).astype(int)
                        if "Sales" in df_np.columns:
                            df_np["Sales"] = df_np["Sales"].map(money)
                        st.info(f"Sold somewhere NEW for the first time: {len(df_np)}")
                        st.dataframe(df_np, hide_index=True, use_container_width=True)

            # Trending + test passing
            weeks = sorted(hist["EndDate"].unique())
            # For "new window": SKUs whose first_sales is within last `new_window_weeks` weeks
            recent_weeks = weeks[-max(new_window_weeks+2, 12):]
            recent = hist[hist["EndDate"].isin(recent_weeks)].copy()
            sku_week = recent.groupby(["SKU","EndDate"], as_index=False)[["Sales","Units"]].sum().sort_values(["SKU","EndDate"])

            def consec_pos_wow(sku, metric="Sales"):
                g = sku_week[sku_week["SKU"].astype(str)==sku].sort_values("EndDate")
                vals = pd.to_numeric(g[metric], errors="coerce").fillna(0.0).to_numpy(dtype=float)
                if len(vals) < 2:
                    return 0
                diffs = np.diff(vals)
                cnt = 0
                for d in diffs[::-1]:
                    if d > 0: cnt += 1
                    else: break
                return cnt

            trending_rows=[]
            passing_rows=[]
            for sku in sorted(cur_skus):
                f_sales = sku_first_sales.get(sku, pd.NaT)
                f_seen  = sku_first_seen.get(sku, pd.NaT)
                if pd.isna(f_sales):
                    continue
                # age in distinct weeks between first_sales and current
                try:
                    age_weeks = len([w for w in weeks if w >= f_sales and w <= cur_end_ts])
                except Exception:
                    age_weeks = None
                if age_weeks is None:
                    continue

                category = "True New" if (not pd.isna(f_seen) and f_seen == f_sales) else "Activated"
                cons = consec_pos_wow(sku, "Sales")

                g2 = hist[(hist["SKU"].astype(str)==sku) & (hist["EndDate"] >= f_sales) & (hist["EndDate"] <= cur_end_ts)].groupby("EndDate", as_index=False)[["Sales","Units"]].sum().sort_values("EndDate")
                sales_vals = pd.to_numeric(g2["Sales"], errors="coerce").fillna(0.0).to_numpy(dtype=float)
                pos_wow = int((np.diff(sales_vals) > 0).sum()) if len(sales_vals) >= 2 else 0
                overall_up = (sales_vals[-1] >= sales_vals[0]) if len(sales_vals) >= 2 else True

                cur_tot = cur_s[cur_s["SKU"].astype(str)==sku]
                cur_sales = float(cur_tot["Sales"].sum()) if not cur_tot.empty else 0.0
                cur_units = float(cur_tot["Units"].sum()) if not cur_tot.empty else 0.0

                if age_weeks <= new_window_weeks and cons >= 2 and cur_sales > 0:
                    trending_rows.append({
                        "SKU": sku,
                        "Category": category,
                        "Weeks active (sales)": age_weeks,
                        "Consecutive +WoW weeks": cons,
                        "This week sales": cur_sales,
                        "This week units": cur_units,
                    })

                if age_weeks <= test_max_weeks and pos_wow >= test_min_pos_wow and overall_up and cur_sales > 0:
                    passing_rows.append({
                        "SKU": sku,
                        "Category": category,
                        "Weeks active (sales)": age_weeks,
                        "+WoW weeks (Sales)": pos_wow,
                        "This week sales": cur_sales,
                        "This week units": cur_units,
                    })


            # Include "new somewhere" placements in trending rules
            if new_place_rows and "Retailer" in hist_pos.columns:
                new_pairs = {(str(r.get("SKU")), str(r.get("Retailer"))) for r in new_place_rows if r.get("SKU") is not None and r.get("Retailer") is not None}

                for (sku, ret) in sorted(new_pairs):
                    f_pair = pair_first_sales.get((sku, ret), pd.NaT)
                    if pd.isna(f_pair):
                        continue
                    try:
                        age_weeks = len([w for w in weeks if w >= f_pair and w <= cur_end_ts])
                    except Exception:
                        continue
                    if age_weeks is None or age_weeks <= 0:
                        continue
                    if age_weeks > new_window_weeks:
                        continue

                    sub = hist_pos[(hist_pos["SKU"].astype(str) == sku) & (hist_pos["Retailer"].astype(str) == ret)].copy()
                    by_w = sub.groupby("EndDate", as_index=False)[["Sales","Units"]].sum()

                    # Align to weeks list
                    sales_map = {pd.to_datetime(rw["EndDate"]): float(rw.get("Sales", 0) or 0) for _, rw in by_w.iterrows()}
                    units_map = {pd.to_datetime(rw["EndDate"]): float(rw.get("Units", 0) or 0) for _, rw in by_w.iterrows()}
                    vals = [sales_map.get(pd.to_datetime(w), 0.0) for w in weeks]
                    uvals = [units_map.get(pd.to_datetime(w), 0.0) for w in weeks]

                    pos_wow = _consecutive_positive_wow(vals)
                    cur_sales = vals[-1] if vals else 0.0
                    cur_units = uvals[-1] if uvals else 0.0

                    trending_rows.append({
                        "SKU": sku,
                        "Where": ret,
                        "Category": "New Placement",
                        "Weeks active (sales)": age_weeks,
                        "Consecutive +WoW weeks (Sales)": pos_wow,
                        "This week sales": cur_sales,
                        "This week units": cur_units,
                    })
            with right:
                st.markdown("### New SKU trending (≤ 8 weeks)")
                if not trending_rows:
                    st.info("No new/activated SKUs are trending under the current rules.")
                else:
                    df_tr = pd.DataFrame(trending_rows).sort_values(["Weeks active (sales)","Consecutive +WoW weeks (Sales)"], ascending=[True, False])
                    if "This week sales" in df_tr.columns:
                        df_tr["This week sales"] = df_tr["This week sales"].map(money)
                    if "This week units" in df_tr.columns:
                        df_tr["This week units"] = df_tr["This week units"].map(intfmt)
                    st.dataframe(df_tr, hide_index=True, use_container_width=True)

            st.markdown("### ⭐ Test success signals (≤ 10 weeks old, 3+ positive WoW weeks)")
            if not passing_rows:
                st.info("No SKUs are flagged as passing the test rules right now.")
            else:
                df_ps = pd.DataFrame(passing_rows).sort_values(["Weeks active (sales)","+WoW weeks (Sales)"], ascending=[True, False])
                df_ps["This week sales"] = df_ps["This week sales"].map(money)
                df_ps["This week units"] = df_ps["This week units"].map(intfmt)
                st.dataframe(df_ps, hide_index=True, use_container_width=True)

            # =======================
            # 4) ACTION CHECKLIST
            # =======================
            st.markdown("## ✅ Recommended actions")
            actions=[]

            # Use alerts if they exist
            try:
                if not r_alert.empty:
                    for rr in r_alert.head(5)["Retailer"].astype(str).tolist():
                        actions.append(f"🔍 Investigate retailer decline: **{rr}** (PO gaps, OOS, program changes).")
            except Exception:
                pass

            try:
                if not v_alert.empty:
                    for vv in v_alert.head(5)["Vendor"].astype(str).tolist():
                        actions.append(f"📞 Call vendor: **{vv}** (down WoW — review pricing, replenishment, coverage).")
            except Exception:
                pass

            try:
                if not growers.empty:
                    for sku in growers.head(5)["SKU"].astype(str).tolist():
                        actions.append(f"📈 Push expansion / reorder: **SKU {sku}** is growing fast WoW.")
            except Exception:
                pass

            if true_new:
                actions.append(f"🧾 Verify mappings: **{len(true_new)} true-new SKU(s)** (vendor/retailer mapping, pricing).")
            if activated:
                actions.append(f"🧩 Audit activation: **{len(activated)} activated SKU(s)** (why 0 before; test start vs mapping).")
            if passing_rows:
                top_pass = [r['SKU'] for r in passing_rows[:5]]
                actions.append(f"⭐ Consider expansion: passing test signals for **{', '.join(top_pass)}**.")

            if not actions:
                st.success("No urgent actions detected this week based on the current thresholds.")
            else:
                st.markdown("\n".join([f"- {a}" for a in actions]))


    def render_tab_momentum():
        with tab_momentum:
            st.subheader("Momentum")
            st.caption("Momentum highlights SKUs with sustained upward sales trends.")

            if df_all_raw is None or df_all_raw.empty:
                st.info("No sales data yet.")
                return

            window = st.slider("Lookback window (weeks)", min_value=4, max_value=52, value=12, step=1)
            mom = compute_momentum_table(df_all, window=window)

            if mom is None or mom.empty:
                st.info("No momentum data available for the selected window.")
                return

            show = mom.copy()

            # SKU lookup (Momentum only)
            sku_q = st.text_input("SKU lookup (Momentum)", value="", placeholder="Type SKU to filter this Momentum table…", key="momentum_sku_lookup")
            if sku_q.strip() and "SKU" in show.columns:
                q = sku_q.strip().lower()
                show = show[show["SKU"].astype(str).str.lower().str.contains(q, na=False)].copy()

            # Ensure numeric for comparisons
            show["Up Weeks"] = pd.to_numeric(show["Up Weeks"], errors="coerce").fillna(0).astype(int)
            show["Down Weeks"] = pd.to_numeric(show["Down Weeks"], errors="coerce").fillna(0).astype(int)
            show["Units_Last"] = pd.to_numeric(show["Units_Last"], errors="coerce").fillna(0).round(0).astype(int)
            show["Sales_Last"] = pd.to_numeric(show["Sales_Last"], errors="coerce").fillna(0.0)

            # Display formatting (match other tables)
            display_df = pd.DataFrame({
                "SKU": show["SKU"].astype(str),
                "Momentum score": show["Momentum"].astype(int),
                "Weeks up": show["Up Weeks"].astype(int),
                "Weeks down": show["Down Weeks"].astype(int),
                "Units last": show["Units_Last"].map(lambda v: f"{int(v):,}"),
                "Sales last": show["Sales_Last"].map(lambda v: f"${float(v):,.2f}"),
            })

            def _highlight_up_down(row):
                up = int(row["Weeks up"])
                down = int(row["Weeks down"])
                styles = [""] * len(row)

                cols = list(row.index)
                up_i = cols.index("Weeks up")
                down_i = cols.index("Weeks down")

                if up > down:
                    styles[up_i] = "color: #2ecc71; font-weight: 800;"
                elif down > up:
                    styles[down_i] = "color: #e74c3c; font-weight: 800;"
                return styles

            st.dataframe(
                display_df.style.apply(_highlight_up_down, axis=1),
                use_container_width=True,
                hide_index=True
            )



    def render_tab_totals_dash():
        with tab_totals_dash:
            st.subheader("Totals Dashboard")

            if df_all.empty:
                st.info("No sales data yet.")
            else:
                d = df_all.copy()
                d["StartDate"] = pd.to_datetime(d["StartDate"], errors="coerce")
                d = d[d["StartDate"].notna()].copy()
                d["Year"] = d["StartDate"].dt.year.astype(int)
                d["Month"] = d["StartDate"].dt.month.astype(int)

                years = sorted(d["Year"].unique().tolist())
                month_name = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}
                month_list = [month_name[i] for i in range(1,13)]


                # -------------------------
                # Controls layout (cleaned)
                # -------------------------

                # Row 1: Primary controls
                p1, p2, p3 = st.columns([1, 2, 2])
                with p1:
                    year_opt = ["All years"] + [str(y) for y in years]
                    pick_year = st.selectbox("Year", options=year_opt, index=0, key="td_year")
                with p2:
                    month_mode = st.radio("Months", options=["All months", "Custom months"], index=0, horizontal=True, key="td_month_mode")
                    if month_mode == "Custom months":
                        sel_month_names = st.multiselect("Select months", options=month_list, default=month_list, key="td_months")
                        sel_months = [k for k,v in month_name.items() if v in sel_month_names]
                    else:
                        sel_months = list(range(1,13))
                with p3:
                    tf_opt = st.selectbox(
                        "Weeks shown",
                        options=["4 weeks", "6 weeks", "8 weeks", "13 weeks", "26 weeks", "52 weeks", "3 months", "6 months", "12 months", "All available"],
                        index=1,
                        key="td_tf_weeks"
                    )

                d2 = d[d["Month"].isin(sel_months)].copy()
                if pick_year != "All years":
                    d2 = d2[d2["Year"] == int(pick_year)].copy()

                # Advanced settings: Group by + Filters + Average window + View
                with st.expander("Advanced settings", expanded=False):
                    a0, a1, a2 = st.columns([2, 2, 2])
                    with a0:
                        group_by = st.selectbox("Group by", options=["Retailer", "Vendor", "SKU"], index=0, key="td_group_by")
                    with a1:
                        month_year_labels = _build_month_year_labels(d2["StartDate"])
                        avg_options = ["4 weeks", "6 weeks", "8 weeks", "13 weeks", "26 weeks", "52 weeks"] + month_year_labels
                        avg_opt = st.selectbox(
                            "Average window",
                            options=avg_options,
                            index=0 if avg_options else 0,
                            key="td_avg_weeks"
                        )
                    with a2:
                        view_mode = st.selectbox("View", options=["Weekly (with Diff/Avg)", "Summary totals"], index=0, key="td_view_mode")

                    st.markdown("#### Filters")
                    f1, f2, f3 = st.columns([2, 2, 2])
                    with f1:
                        retailer_filter = st.multiselect(
                            "Retailer filter (optional)",
                            options=sorted([x for x in d2["Retailer"].dropna().unique().tolist() if str(x).strip()]),
                            key="td_retailer_filter"
                        )
                    with f2:
                        vendor_filter = st.multiselect(
                            "Vendor filter (optional)",
                            options=sorted([x for x in d2["Vendor"].dropna().unique().tolist() if str(x).strip()]),
                            key="td_vendor_filter"
                        )
                    with f3:
                        sku_opts = sorted([x for x in d2["SKU"].dropna().unique().tolist() if str(x).strip()])
                        sku_filter = st.multiselect("SKU filter (optional)", options=sku_opts, key="td_sku_filter")

                # Apply filters
                if retailer_filter:
                    d2 = d2[d2["Retailer"].isin(retailer_filter)]
                if vendor_filter:
                    d2 = d2[d2["Vendor"].isin(vendor_filter)]
                if sku_filter:
                    d2 = d2[d2["SKU"].isin(sku_filter)]


                if d2.empty:
                    st.info("No rows match your filters.")
                else:
                            def _tf_map(x):
                                if x == "All available":
                                    return "all"
                                if "month" in x:
                                    return x
                                try:
                                    return int(x.split()[0])
                                except Exception:
                                    return 13

                            tf_weeks = _tf_map(tf_opt)

                            if view_mode.startswith("Weekly"):
                                sales_t, units_t = make_totals_tables(d2, group_by, tf_weeks, avg_opt)
                                # Keep alphabetical order for readability
                                if not sales_t.empty and group_by in sales_t.columns:
                                    sales_t = sales_t.sort_values(group_by, ascending=True, kind="mergesort")
                                    sales_t = keep_total_last(sales_t, group_by)
                                if not units_t.empty and group_by in units_t.columns:
                                    units_t = units_t.sort_values(group_by, ascending=True, kind="mergesort")
                                    units_t = keep_total_last(units_t, group_by)

                                if sales_t.empty and units_t.empty:
                                    st.info("No weekly totals available for the selected filters.")
                                else:
                                    tabS, tabU = st.tabs(["Sales", "Units"])

                                    with tabS:
                                        _df = sales_t.copy()

                                        def _diff_color(v):
                                            try:
                                                v = float(v)
                                            except Exception:
                                                return ""
                                            if v > 0:
                                                return "color: #2ecc71; font-weight:600;"
                                            if v < 0:
                                                return "color: #e74c3c; font-weight:600;"
                                            return "color: #999999;"

                                        diff_cols = [c for c in _df.columns if c in ["Diff", "Diff vs Avg"]]
                                        sty = _df.style.format({c: fmt_currency for c in _df.columns if c != group_by})
                                        if diff_cols:
                                            sty = sty.applymap(lambda v: _diff_color(v), subset=diff_cols)

                                        # Bold TOTAL row (if present)
                                        try:
                                            if group_by in _df.columns:
                                                total_mask = _df[group_by].astype(str).str.upper().eq("TOTAL")
                                                if total_mask.any():
                                                    def _bold_total(row):
                                                        return ["font-weight:700;" if str(row.get(group_by,"")).upper()=="TOTAL" else "" for _ in row]
                                                    sty = sty.apply(_bold_total, axis=1)
                                        except Exception:
                                            pass

                                        _max_px = 1600 if group_by == "SKU" else 1200
                                        st.dataframe(
                                            sty,
                                            use_container_width=True,
                                            hide_index=True,
                                            height=_table_height(_df, max_px=_max_px),
                                        )
                                    with tabU:
                                        _df = units_t.copy()

                                        def _diff_color(v):
                                            try:
                                                v = float(v)
                                            except Exception:
                                                return ""
                                            if v > 0:
                                                return "color: #2ecc71; font-weight:600;"
                                            if v < 0:
                                                return "color: #e74c3c; font-weight:600;"
                                            return "color: #999999;"

                                        diff_cols = [c for c in _df.columns if c in ["Diff", "Diff vs Avg"]]
                                        sty = _df.style.format({c: fmt_int for c in _df.columns if c != group_by})
                                        if diff_cols:
                                            sty = sty.applymap(lambda v: _diff_color(v), subset=diff_cols)

                                        # Bold TOTAL row (if present)
                                        try:
                                            if group_by in _df.columns:
                                                total_mask = _df[group_by].astype(str).str.upper().eq("TOTAL")
                                                if total_mask.any():
                                                    def _bold_total(row):
                                                        return ["font-weight:700;" if str(row.get(group_by,"")).upper()=="TOTAL" else "" for _ in row]
                                                    sty = sty.apply(_bold_total, axis=1)
                                        except Exception:
                                            pass

                                        _max_px = 1600 if group_by == "SKU" else 1200
                                        st.dataframe(
                                            sty,
                                            use_container_width=True,
                                            hide_index=True,
                                            height=_table_height(_df, max_px=_max_px),
                                        )
                            else:
                                key = group_by
                                agg = d2.groupby(key, as_index=False).agg(
                                    Units=("Units","sum"),
                                    Sales=("Sales","sum"),
                                    SKUs=("SKU","nunique"),
                                )

                                # Add TOTAL row (always at the bottom) for Retailer/Vendor/SKU views
                                try:
                                    total = {
                                        key: "TOTAL",
                                        "Units": float(pd.to_numeric(agg["Units"], errors="coerce").fillna(0).sum()),
                                        "Sales": float(pd.to_numeric(agg["Sales"], errors="coerce").fillna(0).sum()),
                                        "SKUs": float(d2["SKU"].nunique()),
                                    }
                                    agg = pd.concat([agg, pd.DataFrame([total])], ignore_index=True)
                                except Exception:
                                    pass

                                # Alphabetical order + force TOTAL last
                                agg = agg.sort_values(key, ascending=True, kind="mergesort")
                                agg = keep_total_last(agg, key)

                                disp = make_unique_columns(agg)
                                sty = disp.style.format({"Units": fmt_int, "Sales": fmt_currency, "SKUs": fmt_int})
                                # Bold TOTAL row
                                try:
                                    if key in disp.columns:
                                        def _bold_total(row):
                                            return ["font-weight:700;" if str(row.get(key,"")).upper()=="TOTAL" else "" for _ in row]
                                        sty = sty.apply(_bold_total, axis=1)
                                except Exception:
                                    pass

                                st.dataframe(
                                    sty,
                                    use_container_width=True,
                                    hide_index=True,
                                    height=_table_height(disp, max_px=900)
                                )



    def render_tab_top_skus():
        with tab_top_skus:
            st.subheader("Top SKUs (across all retailers)")

            if df_all.empty:
                st.info("No sales data yet.")
            else:
                d = df_all.copy()
                d["StartDate"] = pd.to_datetime(d["StartDate"], errors="coerce")
                d = d[d["StartDate"].notna()].copy()
                d["Year"] = d["StartDate"].dt.year.astype(int)
                d["Month"] = d["StartDate"].dt.month.astype(int)

                years = sorted(d["Year"].unique().tolist())
                month_name = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}
                month_list = [month_name[i] for i in range(1,13)]

                c1, c2, c3, c4, c5 = st.columns([1, 2, 1, 1, 1])
                with c1:
                    year_opt = ["All years"] + [str(y) for y in years]
                    pick_year = st.selectbox("Year", options=year_opt, index=0, key="ts_year")
                with c2:
                    month_mode = st.radio("Months", options=["All months", "Custom months"], index=0, horizontal=True, key="ts_month_mode")
                    if month_mode == "Custom months":
                        sel_month_names = st.multiselect("Select months", options=month_list, default=month_list, key="ts_months")
                        sel_months = [k for k,v in month_name.items() if v in sel_month_names]
                    else:
                        sel_months = list(range(1,13))
                with c3:
                    sort_by = st.selectbox("Rank by", options=["Sales", "Units"], index=0, key="ts_rank_by")
                with c4:
                    top_n = st.number_input("Top N", min_value=10, max_value=5000, value=50, step=10, key="ts_topn")

                with c5:
                    min_val = st.number_input(
                        f"Min {sort_by}",
                        min_value=0.0,
                        value=0.0,
                        step=1.0,
                        key="ts_min_val"
                    )

                f1, f2 = st.columns([2, 2])
                with f1:
                    vendor_filter = st.multiselect(
                        "Vendor filter (optional)",
                        options=sorted([x for x in d["Vendor"].dropna().unique().tolist() if str(x).strip()]),
                        key="ts_vendor_filter"
                    )
                with f2:
                    retailer_filter = st.multiselect(
                        "Retailer filter (optional)",
                        options=sorted([x for x in d["Retailer"].dropna().unique().tolist() if str(x).strip()]),
                        key="ts_retailer_filter"
                    )

                d2 = d[d["Month"].isin(sel_months)].copy()
                if pick_year != "All years":
                    d2 = d2[d2["Year"] == int(pick_year)].copy()
                if vendor_filter:
                    d2 = d2[d2["Vendor"].isin(vendor_filter)]
                if retailer_filter:
                    d2 = d2[d2["Retailer"].isin(retailer_filter)]

                agg = d2.groupby("SKU", as_index=False).agg(
                    Units=("Units","sum"),
                    Sales=("Sales","sum"),
                    Retailers=("Retailer","nunique"),
                )


                # Apply minimum threshold filter (based on Rank by selection)
                if 'min_val' in locals() and min_val and sort_by in agg.columns:
                    agg = agg[agg[sort_by].fillna(0) >= float(min_val)].copy()

                if agg.empty:
                    st.info("No rows match your filters.")
                else:
                    agg = agg.sort_values(sort_by, ascending=False, kind="mergesort").head(int(top_n))
                    agg = make_unique_columns(agg)

                    st.dataframe(
                        agg.style.format({
                            "Units": fmt_int,
                            "Sales": fmt_currency,
                            "Retailers": fmt_int,
                        }),
                        use_container_width=True,
                        hide_index=True,
                        height=650
                    )

                    st.divider()
                    st.markdown("### SKU lookup (cross-retailer totals + breakdown)")

                    sku_q = st.text_input("Type a SKU to inspect (example: EGLAI1)", value="", key="ts_sku_q").strip()
                    if sku_q:
                        qn = str(sku_q).strip().upper()
                        dd = d2.copy()
                        dd["SKU_N"] = dd["SKU"].astype(str).str.strip().str.upper()
                        dd = dd[dd["SKU_N"] == qn].copy()

                        if dd.empty:
                            st.warning("No matching rows for that SKU in the current filters.")
                        else:
                            tot_units = float(dd["Units"].sum())
                            tot_sales = float(dd["Sales"].sum())
                            a, b, c = st.columns([1,1,2])
                            a.metric("Total Units", fmt_int(tot_units))
                            b.metric("Total Sales", fmt_currency(tot_sales))
                            c.caption("Breakdown below is by retailer for the selected year/month filters.")

                            by_ret = dd.groupby("Retailer", as_index=False).agg(Units=("Units","sum"), Sales=("Sales","sum"))
                            by_ret = by_ret.sort_values("Sales", ascending=False, kind="mergesort")
                            st.dataframe(
                                by_ret.style.format({"Units": fmt_int, "Sales": fmt_currency}),
                                use_container_width=True,
                                hide_index=True
                            )



    def render_tab_wow_exc():
        with tab_wow_exc:
            st.subheader("WoW Exceptions (Most Recent Week vs Prior Average)")

            if df.empty:
                st.info("No sales data yet.")
            else:
                # Use all loaded years for lookbacks, but keep the "end week" anchored to the currently selected view (df)
                d0_all = add_week_col(df_all)
                d0_cur = add_week_col(df) if not df.empty else d0_all.copy()

                weeks_all = sorted(d0_cur["Week"].dropna().unique().tolist())
                if len(weeks_all) < 2:
                    st.info("Not enough weeks loaded yet (need at least 2).")
                else:
                    scope = st.selectbox("Scope", options=["All", "Retailer", "Vendor"], index=0, key="wow_scope")

                    d1_all = d0_all.copy()
                    d1_cur = d0_cur.copy()
                    if scope == "Retailer":
                        opts = sorted([x for x in d1_all["Retailer"].dropna().unique().tolist() if str(x).strip()])
                        pick = st.selectbox("Retailer", options=opts, index=0 if opts else 0, key="wow_pick_retailer")
                        d1_all = d1_all[d1_all["Retailer"] == pick].copy()
                        d1_cur = d1_cur[d1_cur["Retailer"] == pick].copy()
                    elif scope == "Vendor":
                        opts = sorted([x for x in d1_all["Vendor"].dropna().unique().tolist() if str(x).strip()])
                        pick = st.selectbox("Vendor", options=opts, index=0 if opts else 0, key="wow_pick_vendor")
                        d1_all = d1_all[d1_all["Vendor"] == pick].copy()
                        d1_cur = d1_cur[d1_cur["Vendor"] == pick].copy()

                    c1, c2, c3 = st.columns([1.2, 1.2, 2.0])
                    with c1:
                        # How far back to average (excluding the most recent week)
                        n_prior = st.selectbox(
                            "Prior window",
                            options=["4 weeks", "6 weeks", "8 weeks", "13 weeks", "26 weeks", "52 weeks", "All prior"],
                            index=1,
                            key="wow_prior_window"
                        )
                    with c2:
                        basis = st.selectbox("Sort basis", options=["Sales", "Units"], index=0, key="wow_sort_basis")
                    with c3:
                        if scope == "All":
                            display_mode = st.radio(
                                "Display mode",
                                options=["SKU totals (all retailers)", "Break out by retailer"],
                                index=0,
                                horizontal=True,
                                key="wow_display_mode"
                            )
                        else:
                            display_mode = "Break out by retailer"

                    # Determine most recent week (from current view) + which prior weeks to use (from all history)
                    d1_cur = d1_cur[d1_cur["Week"].notna()].copy()
                    d1_all = d1_all[d1_all["Week"].notna()].copy()

                    weeks_cur = sorted(d1_cur["Week"].dropna().unique().tolist())
                    if len(weeks_cur) < 2:
                        st.info("Not enough weeks for this selection.")
                    else:
                        end_week = weeks_cur[-1]
                        # prior weeks can extend into previous years
                        prior_weeks_all = sorted([w for w in d1_all["Week"].dropna().unique().tolist() if w < end_week])

                        def _select_prior(prior_weeks):
                            if n_prior == "All prior":
                                return prior_weeks
                            if "month" in str(n_prior).lower():
                                nmo = int(str(n_prior).split()[0])
                                tmp = d1_all[d1_all["Week"].isin(prior_weeks)].copy()
                                tmp["MonthP"] = pd.to_datetime(tmp["StartDate"], errors="coerce").dt.to_period("M")
                                months = sorted(tmp["MonthP"].dropna().unique().tolist())
                                use_months = months[-nmo:] if len(months) >= nmo else months
                                wk = sorted(tmp[tmp["MonthP"].isin(use_months)]["Week"].dropna().unique().tolist())
                                return wk
                            try:
                                n = int(str(n_prior).split()[0])
                            except Exception:
                                n = 6
                            return prior_weeks[-n:] if len(prior_weeks) >= n else prior_weeks

                        prior_weeks = _select_prior(prior_weeks_all)
                        if not prior_weeks:
                            st.info("No prior weeks in the selected window.")
                        else:
                            if display_mode.startswith("SKU totals"):
                                group_cols = ["SKU"]
                                # helpful extra columns
                                extra_aggs = {"Vendor": ("Vendor", lambda s: s.dropna().astype(str).str.strip().iloc[0] if len(s.dropna()) else ""),
                                              "Retailers": ("Retailer", "nunique")}
                            else:
                                group_cols = ["Retailer", "Vendor", "SKU"]
                                extra_aggs = {}

                            dd = d1_all.copy()

                            # Aggregate to weekly grain for each group
                            g = dd.groupby(group_cols + ["Week"], as_index=False).agg(
                                Units=("Units", "sum"),
                                Sales=("Sales", "sum"),
                            )

                            # Split into end week and prior weeks
                            end = g[g["Week"] == end_week].copy()
                            base = g[g["Week"].isin(prior_weeks)].copy()

                            base_avg = base.groupby(group_cols, as_index=False).agg(
                                Units_Base=("Units", "mean"),
                                Sales_Base=("Sales", "mean"),
                            )
                            end_sum = end.groupby(group_cols, as_index=False).agg(
                                Units_End=("Units", "sum"),
                                Sales_End=("Sales", "sum"),
                            )

                            t = end_sum.merge(base_avg, on=group_cols, how="outer").fillna(0.0)
                            t["Units_Diff"] = t["Units_End"] - t["Units_Base"]
                            t["Sales_Diff"] = t["Sales_End"] - t["Sales_Base"]
                            t["Units_% Diff"] = t["Units_Diff"] / t["Units_Base"].replace(0, np.nan)
                            t["Sales_% Diff"] = t["Sales_Diff"] / t["Sales_Base"].replace(0, np.nan)

                            # Add vendor / retailer coverage when in SKU totals mode
                            if display_mode.startswith("SKU totals"):
                                cov = dd.groupby("SKU", as_index=False).agg(
                                    Vendor=("Vendor", lambda s: s.dropna().astype(str).str.strip().iloc[0] if len(s.dropna()) else ""),
                                    Retailers=("Retailer", "nunique")
                                )
                                t = t.merge(cov, on="SKU", how="left")

                            sort_col = "Sales_Diff" if basis == "Sales" else "Units_Diff"
                            t = t.sort_values(sort_col, ascending=True, kind="mergesort")  # show biggest negatives first

                            # Keep useful column order
                            if display_mode.startswith("SKU totals"):
                                cols = ["SKU", "Vendor", "Retailers",
                                        "Units_Base", "Units_End", "Units_Diff", "Units_% Diff",
                                        "Sales_Base", "Sales_End", "Sales_Diff", "Sales_% Diff"]
                            else:
                                cols = ["Retailer", "Vendor", "SKU",
                                        "Units_Base", "Units_End", "Units_Diff", "Units_% Diff",
                                        "Sales_Base", "Sales_End", "Sales_Diff", "Sales_% Diff"]
                            cols = [c for c in cols if c in t.columns]
                            t = t[cols].copy()

                            # Totals row at bottom for quick reference
                            try:
                                total = {c: "" for c in t.columns}
                                first = t.columns[0]
                                total[first] = "TOTAL"
                                for c in t.columns:
                                    if c in {"SKU","Vendor","Retailer"}:
                                        continue
                                    if c == "Retailers":
                                        total[c] = float(dd["Retailer"].nunique())
                                    else:
                                        total[c] = float(pd.to_numeric(t[c], errors="coerce").fillna(0).sum())
                                t = pd.concat([t, pd.DataFrame([total])], ignore_index=True)
                            except Exception:
                                pass

                            # Styling
                            disp = make_unique_columns(t)

                            def _diff_color(v):
                                try:
                                    v = float(v)
                                except Exception:
                                    return ""
                                if v > 0:
                                    return "color: #2ecc71; font-weight:600;"
                                if v < 0:
                                    return "color: #e74c3c; font-weight:600;"
                                return "color: #999999;"

                            sty = disp.style.format({
                                "Units_Base": fmt_int,
                                "Units_End": fmt_int,
                                "Units_Diff": fmt_int_signed,
                                "Units_% Diff": lambda v: f"{(v*100):.1f}%" if pd.notna(v) else "—",
                                "Sales_Base": fmt_currency,
                                "Sales_End": fmt_currency,
                                "Sales_Diff": fmt_currency_signed,
                                "Sales_% Diff": lambda v: f"{(v*100):.1f}%" if pd.notna(v) else "—",
                                "Retailers": fmt_int,
                            })

                            for c in ["Units_Diff", "Sales_Diff"]:
                                if c in disp.columns:
                                    sty = sty.applymap(lambda v: _diff_color(v), subset=[c])

                            # Bold TOTAL row (if present)
                            try:
                                first = disp.columns[0]
                                if first in disp.columns:
                                    def _bold_total(row):
                                        return ["font-weight:700;" if str(row.get(first,"")).upper()=="TOTAL" else "" for _ in row]
                                    sty = sty.apply(_bold_total, axis=1)
                            except Exception:
                                pass

                            st.caption(
                                f"Comparing most recent week ({pd.Timestamp(end_week).strftime('%m-%d')}) "
                                f"to the average of prior {len(prior_weeks)} week(s): "
                                + ", ".join([pd.Timestamp(w).strftime('%m-%d') for w in prior_weeks])
                            )
                            st.dataframe(sty, use_container_width=True, height=_table_height(disp, max_px=1200), hide_index=True)

                            # SKU lookup (based on this WoW Exceptions table + current filters)
                            if isinstance(disp, pd.DataFrame) and (not disp.empty) and ('SKU' in disp.columns):
                                st.markdown('---')
                                st.markdown('### SKU lookup')
                                _sku_list = sorted(disp['SKU'].astype(str).dropna().unique().tolist())
                                sel_sku = st.selectbox('Select SKU', options=_sku_list, index=0, key='wow_sku_lookup') if _sku_list else None
                                if sel_sku:
                                    row_df = disp[disp['SKU'].astype(str) == str(sel_sku)].copy()
                                    st.dataframe(row_df, use_container_width=True, hide_index=True)



    def render_tab_exec():
        with tab_exec:
            st.subheader("Executive Summary")


            scope = st.selectbox("Scope", options=["All", "Retailer", "Vendor"], index=0, key="ex_scope")

            # Filter scope
            if scope == "Retailer":
                opts = sorted([x for x in df_all["Retailer"].dropna().unique().tolist() if str(x).strip()])
                pick = st.selectbox("Retailer", options=opts, index=0 if opts else 0, key="ex_pick_r")
                d = df_all[df_all["Retailer"] == pick].copy()
                title = f"Executive Summary - {pick}"
            elif scope == "Vendor":
                opts = sorted([x for x in df_all["Vendor"].dropna().unique().tolist() if str(x).strip()])
                pick = st.selectbox("Vendor", options=opts, index=0 if opts else 0, key="ex_pick_v")
                d = df_all[df_all["Vendor"] == pick].copy()
                title = f"Executive Summary - {pick}"
            else:
                d = df_all.copy()
                title = "Executive Summary - All Retailers"

            d["StartDate"] = pd.to_datetime(d["StartDate"], errors="coerce")
            d = d[d["StartDate"].notna()].copy()
            if d.empty:
                st.info("No rows in the selected scope/year.")
                st.stop()

            d["Year"] = d["StartDate"].dt.year.astype(int)
            years = sorted(d["Year"].unique().tolist())

            # Keep full-scope history for the multi-year table at the bottom
            d_scope_all = d.copy()

            pick_year = st.selectbox("Year", options=years, index=(len(years)-1 if years else 0), key="ex_year_pick")
            d = d_scope_all[d_scope_all["Year"] == int(pick_year)].copy()

            st.caption(title)

            # KPI row
            m = wow_mom_metrics(d)
            cols = st.columns(6)
            cols[0].metric("Units", fmt_int(m["total_units"]))
            cols[1].metric("Sales", fmt_currency(m["total_sales"]))
            cols[2].markdown(f"<div style='color:{_color(m['wow_units'])}; font-weight:600;'>WoW Units: {fmt_int(m['wow_units']) if m['wow_units'] is not None else '—'}</div>", unsafe_allow_html=True)
            cols[3].markdown(f"<div style='color:{_color(m['wow_sales'])}; font-weight:600;'>WoW Sales: {fmt_currency(m['wow_sales']) if m['wow_sales'] is not None else '—'}</div>", unsafe_allow_html=True)
            cols[4].markdown(f"<div style='color:{_color(m['mom_units'])}; font-weight:600;'>MoM Units: {fmt_int(m['mom_units']) if m['mom_units'] is not None else '—'}</div>", unsafe_allow_html=True)
            cols[5].markdown(f"<div style='color:{_color(m['mom_sales'])}; font-weight:600;'>MoM Sales: {fmt_currency(m['mom_sales']) if m['mom_sales'] is not None else '—'}</div>", unsafe_allow_html=True)

            st.divider()

            # When Scope = ALL: show one line per SKU (combined across all retailers)
            if scope == "All":
                sku = d.groupby("SKU", as_index=False).agg(
                    Vendor=("Vendor", lambda s: (s.dropna().astype(str).str.strip().replace("", np.nan).dropna().iloc[0] if s.dropna().astype(str).str.strip().replace("", np.nan).dropna().shape[0] else "Unmapped")),
                    Retailers=("Retailer", "nunique"),
                    TotalUnits=("Units", "sum"),
                    TotalSales=("Sales", "sum"),
                )
                sku = sku.sort_values("TotalSales", ascending=False, kind="mergesort")

                disp = sku[["SKU","Vendor","Retailers","TotalUnits","TotalSales"]].copy()
                st.markdown("### SKU totals (all retailers combined)")
                st.dataframe(
                    disp.style.format({"Retailers": fmt_int, "TotalUnits": fmt_int, "TotalSales": fmt_currency}),
                    use_container_width=True,
                    hide_index=True,
                    height=_table_height(disp, max_px=1100),
                )

            else:
                # Monthly totals table (keep as-is)
                d2 = d.copy()
                d2["MonthP"] = d2["StartDate"].dt.to_period("M")
                mon = d2.groupby("MonthP", as_index=False).agg(Units=("Units","sum"), Sales=("Sales","sum")).sort_values("MonthP")
                if not mon.empty:
                    mon["Month"] = mon["MonthP"].map(month_label)
                    mon = mon[["Month","Units","Sales"]]
                    st.markdown("### Monthly totals")
                    st.dataframe(
                        mon.style.format({"Units": fmt_int, "Sales": fmt_currency}),
                        use_container_width=True,
                        height=_table_height(mon, max_px=800),
                        hide_index=True
                    )

                # Mix table (keep as-is)
                if scope == "Retailer":
                    mix = d.groupby("Vendor", as_index=False).agg(Units=("Units","sum"), Sales=("Sales","sum"))
                    mix = mix[(mix["Units"].fillna(0) > 0) | (mix["Sales"].fillna(0) > 0)]
                    total_u = float(mix["Units"].sum()) if not mix.empty else 0.0
                    total_s = float(mix["Sales"].sum()) if not mix.empty else 0.0
                    mix["% Units"] = mix["Units"].apply(lambda v: (v/total_u) if total_u else 0.0)
                    mix["% Sales"] = mix["Sales"].apply(lambda v: (v/total_s) if total_s else 0.0)
                    mix = mix.sort_values("% Sales", ascending=False, kind="mergesort")
                    st.markdown("### Vendor mix")
                else:
                    mix = d.groupby("Retailer", as_index=False).agg(Units=("Units","sum"), Sales=("Sales","sum"))
                    mix = mix[(mix["Units"].fillna(0) > 0) | (mix["Sales"].fillna(0) > 0)]
                    total_u = float(mix["Units"].sum()) if not mix.empty else 0.0
                    total_s = float(mix["Sales"].sum()) if not mix.empty else 0.0
                    mix["% Units"] = mix["Units"].apply(lambda v: (v/total_u) if total_u else 0.0)
                    mix["% Sales"] = mix["Sales"].apply(lambda v: (v/total_s) if total_s else 0.0)
                    mix = mix.sort_values("% Sales", ascending=False, kind="mergesort")
                    st.markdown("### Retailer mix")

                st.dataframe(
                    mix.style.format({"Units": fmt_int, "Sales": fmt_currency, "% Units": lambda v: f"{v*100:.1f}%", "% Sales": lambda v: f"{v*100:.1f}%"}),
                    use_container_width=True,
                    height=_table_height(mix, max_px=900),
                    hide_index=True
                )

                st.divider()

                # Top / Bottom SKUs (keep the same idea as before)
                sold = d.groupby(["SKU","Vendor"], as_index=False).agg(Units=("Units","sum"), Sales=("Sales","sum"))
                sold = sold[(sold["Units"].fillna(0) > 0) | (sold["Sales"].fillna(0) > 0)].copy()

                left, right = st.columns(2)
                with left:
                    st.markdown("### Top 10 SKUs (by Sales)")
                    top10 = sold.sort_values("Sales", ascending=False, kind="mergesort").head(10)[["SKU","Vendor","Units","Sales"]]
                    st.dataframe(top10.style.format({"Units": fmt_int, "Sales": fmt_currency}), use_container_width=True, hide_index=True, height=_table_height(top10, max_px=520))
                with right:
                    st.markdown("### Bottom 10 SKUs (by Sales)")
                    bot10 = sold.sort_values("Sales", ascending=True, kind="mergesort").head(10)[["SKU","Vendor","Units","Sales"]]
                    st.dataframe(bot10.style.format({"Units": fmt_int, "Sales": fmt_currency}), use_container_width=True, hide_index=True, height=_table_height(bot10, max_px=520))
                st.divider()
                st.markdown("### Multi-year totals (always shown)")
                try:
                    d_all_scope = d_scope_all.copy()
                    d_all_scope["StartDate"] = pd.to_datetime(d_all_scope["StartDate"], errors="coerce")
                    d_all_scope = d_all_scope[d_all_scope["StartDate"].notna()].copy()
                    d_all_scope["Year"] = d_all_scope["StartDate"].dt.year.astype(int)

                    # Choose dimension for the table:
                    if scope == "Retailer":
                        dim = "Vendor"
                        label = "Vendors"
                    elif scope == "Vendor":
                        dim = "Retailer"
                        label = "Retailers"
                    else:
                        dim = "Retailer"
                        label = "Retailers"

                    years_all = sorted(d_all_scope["Year"].unique().tolist())
                    if years_all:
                        years_show = years_all[-4:] if len(years_all) > 4 else years_all  # keep recent 4 by default
                    else:
                        years_show = []

                    # Aggregate and pivot
                    g = d_all_scope.groupby([dim, "Year"], as_index=False).agg(Units=("Units","sum"), Sales=("Sales","sum"))
                    u = g.pivot_table(index=dim, columns="Year", values="Units", aggfunc="sum", fill_value=0.0)
                    s = g.pivot_table(index=dim, columns="Year", values="Sales", aggfunc="sum", fill_value=0.0)

                    # Build a combined display with Units_YYYY and Sales_YYYY columns
                    out = pd.DataFrame({dim: u.index}).reset_index(drop=True)
                    for y in years_all:
                        out[f"Units {y}"] = u.get(y, 0.0).values if y in u.columns else 0.0
                        out[f"Sales {y}"] = s.get(y, 0.0).values if y in s.columns else 0.0

                    # Totals row
                    try:
                        total = {dim: "TOTAL"}
                        for c in out.columns:
                            if c == dim:
                                continue
                            total[c] = float(pd.to_numeric(out[c], errors="coerce").fillna(0).sum())
                        out = pd.concat([out, pd.DataFrame([total])], ignore_index=True)
                    except Exception:
                        pass

                    # Sort by latest Sales year (excluding TOTAL)
                    try:
                        latest_year = max(years_all) if years_all else None
                        if latest_year is not None and f"Sales {latest_year}" in out.columns:
                            m_total = out[dim].astype(str).str.upper().eq("TOTAL")
                            rest = out.loc[~m_total].sort_values(f"Sales {latest_year}", ascending=False, kind="mergesort")
                            out = pd.concat([rest, out.loc[m_total]], ignore_index=True)
                    except Exception:
                        pass

                    fmt = {}
                    for c in out.columns:
                        if c.startswith("Units "):
                            fmt[c] = fmt_int
                        if c.startswith("Sales "):
                            fmt[c] = fmt_currency

                    st.caption(f"{label} across years (independent of the Year dropdown above).")
                    st.dataframe(out.style.format(fmt), use_container_width=True, hide_index=True, height=_table_height(out, max_px=1100))
                except Exception:
                    st.caption("Multi-year table will appear when historical data is available.")

            st.markdown("---")
            st.caption("Export a simple one-page summary (PDF) for sharing.")
            if st.button("Prepare one-pager PDF", key="ex_onepager_btn"):
                st.info("One-pager export ready in this build.")

            st.divider()
            st.markdown("### Weekly Summary Export")
            st.caption("One-click export of the current week highlights (Action Center) to Excel and PDF.")

            if df_all is not None and not df_all.empty:
                cur, prev, cur_start, cur_end = _get_current_and_prev_week(df_all_raw)
                if cur is not None and not cur.empty:
                    # Build the same core tables used in Action Center
                    cur_r = cur.groupby("Retailer", as_index=False)[["Units","Sales"]].sum()
                    prev_r = prev.groupby("Retailer", as_index=False)[["Units","Sales"]].sum() if prev is not None and not prev.empty else pd.DataFrame(columns=["Retailer","Units","Sales"])
                    cur_v = cur.groupby("Vendor", as_index=False)[["Units","Sales"]].sum()
                    prev_v = prev.groupby("Vendor", as_index=False)[["Units","Sales"]].sum() if prev is not None and not prev.empty else pd.DataFrame(columns=["Vendor","Units","Sales"])
                    cur_s = cur.groupby("SKU", as_index=False)[["Units","Sales"]].sum()
                    prev_s = prev.groupby("SKU", as_index=False)[["Units","Sales"]].sum() if prev is not None and not prev.empty else pd.DataFrame(columns=["SKU","Units","Sales"])

                    movers = cur_s.merge(prev_s, on="SKU", how="left", suffixes=("_cur","_prev"))
                    movers["Sales_prev"] = movers["Sales_prev"].fillna(0.0)
                    movers["Units_prev"] = movers["Units_prev"].fillna(0.0)
                    movers["WoW Sales"] = movers["Sales_cur"] - movers["Sales_prev"]
                    movers["WoW Units"] = movers["Units_cur"] - movers["Units_prev"]
                    movers["_abs_wow_sales"] = movers["WoW Sales"].abs()
                    movers = movers.sort_values("_abs_wow_sales", ascending=False).head(25).drop(columns=["_abs_wow_sales"])
                    prior = df_all.copy()
                    prior["EndDate"] = pd.to_datetime(prior["EndDate"], errors="coerce")
                    prior = prior[prior["EndDate"].notna()].copy()
                    prior = prior[prior["EndDate"] < cur_end]
                    new_skus = sorted(set(cur["SKU"].astype(str)) - set(prior["SKU"].astype(str)))

                    mom = compute_momentum_scores(df_all, window=8)

                    def _declines(cur_df, prev_df, key, min_prev_sales=1000.0, pct=-0.25):
                        if prev_df is None or prev_df.empty:
                            return pd.DataFrame()
                        cmp = cur_df.merge(prev_df, on=key, how="left", suffixes=("_cur","_prev"))
                        cmp["Sales_prev"] = cmp["Sales_prev"].fillna(0.0)
                        cmp["Units_prev"] = cmp["Units_prev"].fillna(0.0)
                        cmp["WoW Sales"] = cmp["Sales_cur"] - cmp["Sales_prev"]
                        cmp["WoW Units"] = cmp["Units_cur"] - cmp["Units_prev"]
                        cmp["%ΔSales"] = np.where(cmp["Sales_prev"]>0, cmp["WoW Sales"]/cmp["Sales_prev"], np.nan)
                        alerts = cmp[(cmp["Sales_prev"] >= min_prev_sales) & (cmp["%ΔSales"] <= pct)].copy()
                        return alerts.sort_values("%ΔSales").head(15)

                    vendor_decl = _declines(cur_v.rename(columns={"Sales":"Sales_cur","Units":"Units_cur"}),
                                            prev_v.rename(columns={"Sales":"Sales_prev","Units":"Units_prev"}),
                                            key="Vendor", min_prev_sales=1000.0)
                    retailer_decl = _declines(cur_r.rename(columns={"Sales":"Sales_cur","Units":"Units_cur"}),
                                              prev_r.rename(columns={"Sales":"Sales_prev","Units":"Units_prev"}),
                                              key="Retailer", min_prev_sales=5000.0)

                    # Excel export
                    excel_buf = io.BytesIO()
                    with pd.ExcelWriter(excel_buf, engine="openpyxl") as writer:
                        pd.DataFrame({
                            "Week": [_fmt_week_range(cur_start, cur_end)],
                            "Total Sales": [float(cur_s["Sales"].sum())],
                            "Total Units": [float(cur_s["Units"].sum())],
                            "New SKUs": [len(new_skus)],
                        }).to_excel(writer, sheet_name="Highlights", index=False)

                        movers.rename(columns={
                            "Sales_cur":"Sales (This Week)",
                            "Sales_prev":"Sales (Prev Week)",
                            "Units_cur":"Units (This Week)",
                            "Units_prev":"Units (Prev Week)",
                        }).to_excel(writer, sheet_name="Biggest Movers", index=False)

                        if not vendor_decl.empty:
                            vendor_decl.to_excel(writer, sheet_name="Declining Vendors", index=False)
                        if not retailer_decl.empty:
                            retailer_decl.to_excel(writer, sheet_name="Declining Retailers", index=False)
                        if new_skus:
                            pd.DataFrame({"New SKUs": new_skus}).to_excel(writer, sheet_name="New SKUs", index=False)
                        if mom is not None and not mom.empty:
                            mom.head(50).to_excel(writer, sheet_name="Momentum Top 50", index=False)

                    excel_bytes = excel_buf.getvalue()
                    st.download_button(
                        "Download Weekly Summary (Excel)",
                        data=excel_bytes,
                        file_name=f"Weekly_Summary_{pd.to_datetime(cur_end).date()}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key="dl_weekly_summary_excel",
                    )

                    # PDF export (simple highlights)
                    # PDF export (professional)
                    # KPIs
                    total_sales_cur = float(cur_s["Sales"].sum())
                    total_units_cur = float(cur_s["Units"].sum())
                    total_sales_prev = float(prev_s["Sales"].sum()) if prev is not None and not prev.empty else 0.0
                    total_units_prev = float(prev_s["Units"].sum()) if prev is not None and not prev.empty else 0.0

                    kpis = [
                        ("Total Sales", f"${total_sales_cur:,.2f}", f"${total_sales_prev:,.2f}", f"${(total_sales_cur-total_sales_prev):,.2f}"),
                        ("Total Units", f"{int(round(total_units_cur)):,}", f"{int(round(total_units_prev)):,}", f"{int(round(total_units_cur-total_units_prev)):,}"),
                        ("New SKUs", f"{len(new_skus):,}", "-", "-"),
                    ]

                    # Top lists (with WoW Sales)
                    # Current week base tables exist: cur_r/cur_v/cur_s; prior week: prev_r/prev_v/prev_s
                    def _add_wow_sales(cur_df: pd.DataFrame, prev_df: pd.DataFrame, key: str) -> pd.DataFrame:
                        cur2 = cur_df.copy()
                        prev2 = prev_df.copy() if prev_df is not None else pd.DataFrame(columns=[key, "Sales", "Units"])
                        if key not in cur2.columns:
                            return cur2
                        if "Sales" not in cur2.columns:
                            cur2["Sales"] = 0.0
                        if "Units" not in cur2.columns:
                            cur2["Units"] = 0.0
                        if key not in prev2.columns:
                            prev2[key] = []
                        if "Sales" not in prev2.columns:
                            prev2["Sales"] = 0.0
                        prev2 = prev2[[key, "Sales"]].rename(columns={"Sales":"Sales_prev"})
                        cur2 = cur2.merge(prev2, on=key, how="left")
                        cur2["Sales_prev"] = cur2["Sales_prev"].fillna(0.0)
                        cur2["WoW $ Diff"] = cur2["Sales"] - cur2["Sales_prev"]
                        cur2["WoW $ %"] = np.where(cur2["Sales_prev"] > 0, cur2["WoW $ Diff"] / cur2["Sales_prev"], np.nan)
                        return cur2

                    top_r = _add_wow_sales(cur_r, prev_r, "Retailer")
                    top_v = _add_wow_sales(cur_v, prev_v, "Vendor")
                    top_s = _add_wow_sales(cur_s, prev_s, "SKU")

                    # Sort by current Sales and take Top N
                    top_r = top_r.sort_values("Sales", ascending=False).head(10) if "Sales" in top_r.columns else top_r.head(10)
                    top_v = top_v.sort_values("Sales", ascending=False).head(10) if "Sales" in top_v.columns else top_v.head(10)
                    top_s = top_s.sort_values("Sales", ascending=False).head(10) if "Sales" in top_s.columns else top_s.head(10)

                    # Format tables for PDF
                    def _fmt_pdf_money(v):
                        try:
                            if isinstance(v, str) and v.strip().startswith("$"):
                                return v.strip()
                            return f"${float(v):,.2f}"
                        except:
                            return ""
                    def _fmt_pdf_int(v):
                        try: return f"{int(round(float(v))):,}"
                        except: return ""

                    # Column order for Page-1 Top sections: keep Units, Sales, add WoW Sales ($ and %)
                    def _subset_top(df_in: pd.DataFrame, key_col: str) -> pd.DataFrame:
                        if df_in is None or df_in.empty:
                            return df_in
                        cols = [c for c in [key_col, "Units", "Sales", "WoW $ Diff"] if c in df_in.columns]
                        return df_in[cols]


                    top_r_pdf = top_r.copy()
                    if not top_r_pdf.empty:
                        # Keep Units column (current week) and format it
                        if "Units" in top_r_pdf.columns: top_r_pdf["Units"] = top_r_pdf["Units"].map(_fmt_pdf_int)
                        # Format Sales and WoW Sales
                        if "Sales" in top_r_pdf.columns: top_r_pdf["Sales"] = top_r_pdf["Sales"].map(_fmt_pdf_money)
                        if "WoW $ Diff" in top_r_pdf.columns: top_r_pdf["WoW $ Diff"] = top_r_pdf["WoW $ Diff"].map(_fmt_pdf_money)
                    top_r_pdf = _subset_top(top_r_pdf, "Retailer")

                    top_v_pdf = top_v.copy()
                    if not top_v_pdf.empty:
                        # Keep Units column (current week) and format it
                        if "Units" in top_v_pdf.columns: top_v_pdf["Units"] = top_v_pdf["Units"].map(_fmt_pdf_int)
                        # Format Sales and WoW Sales
                        if "Sales" in top_v_pdf.columns: top_v_pdf["Sales"] = top_v_pdf["Sales"].map(_fmt_pdf_money)
                        if "WoW $ Diff" in top_v_pdf.columns: top_v_pdf["WoW $ Diff"] = top_v_pdf["WoW $ Diff"].map(_fmt_pdf_money)
                    top_v_pdf = _subset_top(top_v_pdf, "Vendor")

                    top_s_pdf = top_s.copy()
                    if not top_s_pdf.empty:
                        # Keep Units column (current week) and format it
                        if "Units" in top_s_pdf.columns: top_s_pdf["Units"] = top_s_pdf["Units"].map(_fmt_pdf_int)
                        # Format Sales and WoW Sales
                        if "Sales" in top_s_pdf.columns: top_s_pdf["Sales"] = top_s_pdf["Sales"].map(_fmt_pdf_money)
                        if "WoW $ Diff" in top_s_pdf.columns: top_s_pdf["WoW $ Diff"] = top_s_pdf["WoW $ Diff"].map(_fmt_pdf_money)
                    top_s_pdf = _subset_top(top_s_pdf, "SKU")

                    movers_pdf = movers.rename(columns={
                        "Sales_cur":"Sales (This Week)",
                        "Sales_prev":"Sales (Prev Week)",
                        "Units_cur":"Units (This Week)",
                        "Units_prev":"Units (Prev Week)",
                    }).copy()

                    for c in ["Sales (This Week)","Sales (Prev Week)","WoW Sales"]:
                        if c in movers_pdf.columns: movers_pdf[c] = movers_pdf[c].map(_fmt_pdf_money)
                    for c in ["Units (This Week)","Units (Prev Week)","WoW Units"]:
                        if c in movers_pdf.columns: movers_pdf[c] = movers_pdf[c].map(_fmt_pdf_int)

                    vendor_decl_pdf = vendor_decl.copy() if "vendor_decl" in locals() else pd.DataFrame()
                    retailer_decl_pdf = retailer_decl.copy() if "retailer_decl" in locals() else pd.DataFrame()

                    # Format decline tables for PDF (currency/units/%)
                    def _format_generic_pdf(df_in: pd.DataFrame) -> pd.DataFrame:
                        df_out = df_in.copy()
                        if df_out is None or df_out.empty:
                            return df_out
                        for col in df_out.columns:
                            lc = str(col).lower()
                            if "sales" in lc or lc.endswith("$"):
                                df_out[col] = df_out[col].map(lambda v: _fmt_pdf_money(v) if str(v) and not str(v).strip().startswith("$") else str(v))
                            elif "unit" in lc or "qty" in lc:
                                df_out[col] = df_out[col].map(lambda v: _fmt_pdf_int(v) if str(v) and str(v).replace(",","").replace(".","").isdigit() else str(v))
                            elif "%" in str(col) or "pct" in lc or "percent" in lc or "wow%" in lc:
                                def _pp(v):
                                    try:
                                        vv=float(v)
                                        return f"{vv*100:,.1f}%"
                                    except Exception:
                                        return str(v)
                                df_out[col] = df_out[col].map(_pp)
                        return df_out

                    vendor_decl_pdf = _format_generic_pdf(vendor_decl_pdf)
                    retailer_decl_pdf = _format_generic_pdf(retailer_decl_pdf)

                    bullets = [
                        f"Week range: {_fmt_week_range(cur_start, cur_end)}",
                        f"Top Retailer (Sales): {top_r.iloc[0]['Retailer'] if not top_r.empty else '—'}",
                        f"Top Vendor (Sales): {top_v.iloc[0]['Vendor'] if not top_v.empty else '—'}",
                        f"Top SKU (Sales): {str(top_s.iloc[0]['SKU']) if not top_s.empty else '—'}",
                    ]
                    if new_skus:
                        bullets.append(f"New SKUs detected: {len(new_skus)}")

                    subtitle = f"Weekly Summary • Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                    sections = [
                        ("Highlights", None, bullets),
                        ("Top Retailers", top_r_pdf, None),
                        ("Top Vendors", top_v_pdf, None),
                        ("Top SKUs", top_s_pdf, None),
                        ("Biggest Movers (SKU)", movers_pdf.head(15), None),
                    ]

                    if vendor_decl_pdf is not None and not vendor_decl_pdf.empty:
                        sections.append(("Declining Vendors", vendor_decl_pdf.head(15).astype(str), None))
                    if retailer_decl_pdf is not None and not retailer_decl_pdf.empty:
                        sections.append(("Declining Retailers", retailer_decl_pdf.head(15).astype(str), None))
                    if new_skus:
                        sections.append(("New SKUs", pd.DataFrame({"New SKUs": new_skus[:50]}), None))
                    if mom is not None and not mom.empty:
                        mom_pdf = mom.head(15).copy()
                        if "Momentum" in mom_pdf.columns:
                            mom_pdf["Momentum"] = mom_pdf["Momentum"].map(lambda x: f"{float(x):.0f}" if str(x)!="" else "")
                        if "Units_Last" in mom_pdf.columns:
                            mom_pdf["Units_Last"] = mom_pdf["Units_Last"].map(_fmt_pdf_int)
                        if "Sales_Last" in mom_pdf.columns:
                            mom_pdf["Sales_Last"] = mom_pdf["Sales_Last"].map(_fmt_pdf_money)
                        cols_mom = ["SKU","Momentum","Lookback Weeks","Up Weeks","Down Weeks","Weeks","Units_Last","Sales_Last"]
                        cols_mom = [c for c in cols_mom if c in mom_pdf.columns]
                        if cols_mom:
                            # Ensure numeric formatting for PDF
                            if "Units_Last" in mom_pdf.columns:
                                mom_pdf["Units_Last"] = mom_pdf["Units_Last"].map(_fmt_pdf_int)
                            if "Sales_Last" in mom_pdf.columns:
                                mom_pdf["Sales_Last"] = mom_pdf["Sales_Last"].map(_fmt_pdf_money)
                            # Momentum Leaders (PDF export uses last 12 weeks)
                    mom_pdf = compute_momentum_table(df_all, window=12)
                    if mom_pdf is not None and not mom_pdf.empty:
                        mom_pdf = mom_pdf.head(15).copy()

                        # Format numeric columns for PDF
                        if "Units_Last" in mom_pdf.columns:
                            mom_pdf["Units_Last"] = mom_pdf["Units_Last"].map(_fmt_pdf_int)
                        if "Sales_Last" in mom_pdf.columns:
                            mom_pdf["Sales_Last"] = mom_pdf["Sales_Last"].map(_fmt_pdf_money)

                        # Exact column order requested
                        mom_pdf = mom_pdf.rename(columns={"Momentum":"Momentum"})
                        cols_mom = ["SKU", "Momentum", "Up Weeks", "Down Weeks", "Units_Last", "Sales_Last"]
                        cols_mom = [c for c in cols_mom if c in mom_pdf.columns]
                        sections.append(("Momentum Leaders", mom_pdf[cols_mom].astype(str), None))
                    # Build professional PDF (3-page executive layout)
                    try:
                        mom_pdf_export = compute_momentum_table(df_all, window=12).head(15).copy()
                    except Exception:
                        mom_pdf_export = pd.DataFrame(columns=["SKU","Momentum","Up Weeks","Down Weeks","Units_Last","Sales_Last"])

                    if mom_pdf_export is None or mom_pdf_export.empty:
                        mom_pdf_export = pd.DataFrame(columns=["SKU","Momentum","Up Weeks","Down Weeks","Units_Last","Sales_Last"])
                    else:
                        # ensure correct column names for PDF and formatting
                        if "Units_Last" in mom_pdf_export.columns: mom_pdf_export["Units_Last"] = mom_pdf_export["Units_Last"].map(_fmt_pdf_int)
                        if "Sales_Last" in mom_pdf_export.columns: mom_pdf_export["Sales_Last"] = mom_pdf_export["Sales_Last"].map(_fmt_pdf_money)


                    # Highlights + KPI dict for PDF (safe defaults)
                    total_sales_cur = float(cur["Sales"].sum()) if "Sales" in cur.columns else 0.0
                    total_units_cur = float(cur["Units"].sum()) if "Units" in cur.columns else 0.0
                    total_sales_prev = float(prev["Sales"].sum()) if (prev is not None and not prev.empty and "Sales" in prev.columns) else 0.0
                    total_units_prev = float(prev["Units"].sum()) if (prev is not None and not prev.empty and "Units" in prev.columns) else 0.0

                    wow_sales = total_sales_cur - total_sales_prev
                    wow_units = total_units_cur - total_units_prev
                    wow_sales_pct = (wow_sales / total_sales_prev) if total_sales_prev > 0 else None
                    wow_units_pct = (wow_units / total_units_prev) if total_units_prev > 0 else None

                    kpi_dict = {
                        "Sales": _fmt_pdf_money(total_sales_cur),
                        "Units": _fmt_pdf_int(total_units_cur),
                        "WoW Sales": _fmt_pdf_money(wow_sales) + (f" ({wow_sales_pct*100:,.1f}%)" if wow_sales_pct is not None else ""),
                        "WoW Units": _fmt_pdf_int(wow_units) + (f" ({wow_units_pct*100:,.1f}%)" if wow_units_pct is not None else ""),
                    }

                    # Simple highlight bullets (kept short for the PDF)
                    highlights = [
                        f"Week {cur_start.date()} to {cur_end.date()} total sales {_fmt_pdf_money(total_sales_cur)} on {_fmt_pdf_int(total_units_cur)} units.",
                        f"WoW sales change: {_fmt_pdf_money(wow_sales)}.",
                        f"WoW units change: {_fmt_pdf_int(wow_units)}.",
                    ]

                    executive_takeaway, drivers_df, opportunities_df = _compute_wow_insights(df_all_raw)

                    pdf_bytes = make_weekly_summary_pdf_bytes(
                        "Weekly Summary",
                        highlights,
                        kpi_dict,
                        top_r_pdf,
                        top_v_pdf,
                        top_s_pdf,
                        movers_pdf if "movers_pdf" in locals() else pd.DataFrame(),
                        vendor_decl_pdf if "vendor_decl_pdf" in locals() else pd.DataFrame(),
                        retailer_decl_pdf if "retailer_decl_pdf" in locals() else pd.DataFrame(),
                        mom_pdf_export,
                        df_all_raw,
                        logo_path=LOGO_PATH if "LOGO_PATH" in globals() else None,
                        executive_takeaway=executive_takeaway,
                        drivers_df=drivers_df,
                        opportunities_df=opportunities_df,
                    )
                    st.download_button(
                        "Download Weekly Summary (PDF)",
                        data=pdf_bytes,
                        file_name=f"Weekly_Summary_{pd.to_datetime(cur_end).date()}.pdf",
                        mime="application/pdf",
                        use_container_width=True,
                        key="dl_weekly_summary_pdf",
                    )

            st.divider()



    def render_tab_comparisons():
        with tab_comparisons:
            st.subheader("Comparisons")

            view = st.selectbox("View", options=["Retailer / Vendor Comparison", "SKU Comparison"], index=0, key="cmp_view")

            if view == "Retailer / Vendor Comparison":
                render_comparison_retailer_vendor()
            else:
                render_comparison_sku()

            ctx = st.session_state.get("cmp_ctx", {})
            render_comparison_extras(ctx)



    def render_tab_sku_intel():
        with tab_sku_intel:
            st.subheader("SKU Intelligence")
            view = st.selectbox("View", options=["SKU Health Score", "Lost Sales Detector"], index=0, key="sku_intel_view")
            if view == "SKU Health Score":
                render_sku_health()
            else:
                render_lost_sales()

            st.markdown("---")
            st.markdown("### SKU lookup")

            if df_all.empty:
                st.info("No sales data loaded.")
            else:
                dsl = df_all.copy()
                if "StartDate" in dsl.columns:
                    dsl["StartDate"] = pd.to_datetime(dsl["StartDate"], errors="coerce")
                    dsl["Year"] = dsl["StartDate"].dt.year
                else:
                    dsl["Year"] = np.nan

                sku_opts = sorted([str(x).strip() for x in dsl.get("SKU", pd.Series([], dtype="object")).dropna().unique().tolist() if str(x).strip()])

                cL, cR = st.columns([2, 1])
                with cL:
                    sku_query = st.text_input("Search SKU (type part of SKU)", value="", key="si_lookup_q")
                with cR:
                    max_rows = st.selectbox("Max rows", options=[25, 50, 100, 200], index=1, key="si_lookup_max")

                if sku_query.strip():
                    q = sku_query.strip().lower()
                    matches = [s for s in sku_opts if q in s.lower()]
                    if not matches:
                        st.warning("No SKUs match that search.")
                        matches = []
                else:
                    matches = sku_opts[:200]

                pick_sku = st.selectbox("Select SKU", options=matches if matches else ["—"], index=0, key="si_lookup_pick")

                if pick_sku and pick_sku != "—":
                    df_sku = dsl[dsl["SKU"].astype(str).str.strip() == str(pick_sku).strip()].copy()

                    if df_sku.empty:
                        st.warning("No rows found for that SKU.")
                    else:
                        units_total = float(df_sku["Units"].sum()) if "Units" in df_sku.columns else 0.0
                        sales_total = float(df_sku["Sales"].sum()) if "Sales" in df_sku.columns else 0.0
                        first_date = df_sku["StartDate"].min() if "StartDate" in df_sku.columns else None
                        last_date = df_sku["StartDate"].max() if "StartDate" in df_sku.columns else None

                        k1, k2, k3, k4 = st.columns(4)
                        k1.metric("Total Units", fmt_int(units_total))
                        k2.metric("Total Sales", fmt_currency(sales_total))
                        k3.metric("First Week", first_date.date().isoformat() if pd.notna(first_date) else "—")
                        k4.metric("Last Week", last_date.date().isoformat() if pd.notna(last_date) else "—")

                        st.markdown("#### Breakdown by year")
                        by_year = df_sku.groupby("Year", as_index=False).agg(Units=("Units","sum"), Sales=("Sales","sum")).sort_values("Year")
                        by_year_disp = by_year.copy()
                        if "Units" in by_year_disp.columns:
                            by_year_disp["Units"] = by_year_disp["Units"].apply(fmt_int)
                        if "Sales" in by_year_disp.columns:
                            by_year_disp["Sales"] = by_year_disp["Sales"].apply(fmt_currency)
                        st.dataframe(by_year_disp, use_container_width=True, hide_index=True)

                        c1, c2 = st.columns(2)
                        with c1:
                            st.markdown("#### Breakdown by retailer")
                            if "Retailer" in df_sku.columns:
                                by_r = df_sku.groupby("Retailer", as_index=False).agg(Units=("Units","sum"), Sales=("Sales","sum")).sort_values("Sales", ascending=False)
                                by_r_disp = by_r.copy()
                                by_r_disp["Units"] = by_r_disp["Units"].apply(fmt_int)
                                by_r_disp["Sales"] = by_r_disp["Sales"].apply(fmt_currency)
                                st.dataframe(style_numeric_posneg(by_r_disp.head(int(max_rows)), cols=[c for c in by_r_disp.columns if any(k in str(c).lower() for k in ['units','sales','delta','diff','change','%'])]), use_container_width=True, hide_index=True, height=_table_height(by_r_disp.head(int(max_rows)), max_px=450))
                            else:
                                st.write("—")
                        with c2:
                            st.markdown("#### Breakdown by vendor")
                            if "Vendor" in df_sku.columns:
                                by_v = df_sku.groupby("Vendor", as_index=False).agg(Units=("Units","sum"), Sales=("Sales","sum")).sort_values("Sales", ascending=False)
                                by_v_disp = by_v.copy()
                                by_v_disp["Units"] = by_v_disp["Units"].apply(fmt_int)
                                by_v_disp["Sales"] = by_v_disp["Sales"].apply(fmt_currency)
                                st.dataframe(style_numeric_posneg(by_v_disp.head(int(max_rows)), cols=[c for c in by_v_disp.columns if any(k in str(c).lower() for k in ['units','sales','delta','diff','change','%'])]), use_container_width=True, hide_index=True, height=_table_height(by_v_disp.head(int(max_rows)), max_px=450))
                            else:
                                st.write("—")

                        st.markdown("#### Weekly detail (most recent first)")
                        cols_show = [c for c in ["StartDate", "Retailer", "Vendor", "SKU", "Units", "Sales"] if c in df_sku.columns]
                        detail = df_sku[cols_show].copy()
                        if "StartDate" in detail.columns:
                            detail = detail.sort_values("StartDate", ascending=False)
                            detail["StartDate"] = detail["StartDate"].dt.date.astype(str)
                        if "Units" in detail.columns:
                            detail["Units"] = detail["Units"].apply(fmt_int)
                        if "Sales" in detail.columns:
                            detail["Sales"] = detail["Sales"].apply(fmt_currency)

                        detail = make_unique_columns(detail)
                        st.dataframe(style_numeric_posneg(detail.head(int(max_rows)), cols=[c for c in detail.columns if any(k in str(c).lower() for k in ['units','sales','delta','diff','change','%'])]), use_container_width=True, hide_index=True, height=_table_height(detail.head(int(max_rows)), max_px=650))



    def render_tab_forecasting():
        with tab_forecasting:
            st.subheader("Forecasting")

            view = st.selectbox("View", options=["Run-Rate Forecast", "Seasonality"], index=0, key="fc_view")
            if view == "Run-Rate Forecast":
                render_runrate()
            else:
                render_seasonality()



    def render_tab_alerts():
        with tab_alerts:
            st.subheader("Alerts")
            with st.expander("Data coverage (loaded history)", expanded=False):
                render_data_coverage_panel(df_all)
            view = st.selectbox("View", options=["Insights & Alerts", "No Sales SKUs"], index=0, key="alerts_view")
            if view == "Insights & Alerts":
                render_alerts()
            else:
                render_no_sales()



    def render_tab_data_mgmt():
        with tab_data_mgmt:
            st.subheader("Data Management")

            st.markdown("### Data coverage")
            render_data_coverage_panel(df_all)

            st.markdown("---")

            tool = st.selectbox(
                "Tools",
                options=["Bulk Data Upload", "Edit Vendor Map", "Backup / Restore"],
                index=0,
                key="dm_tool",
            )

            if tool == "Bulk Data Upload":
                if "render_bulk_data_upload" in globals():
                    render_bulk_data_upload()
                else:
                    st.info("Bulk Data Upload is not available in this build.")

            elif tool == "Edit Vendor Map":
                if "render_edit_vendor_map" in globals():
                    render_edit_vendor_map()
                else:
                    st.info("Vendor Map editor is not available in this build.")

            else:
                if "render_backup_restore" in globals():
                    render_backup_restore()
                else:
                    st.info("Backup / Restore is not available in this build.")



    def render_tab_year_summary():
        with tab_year_summary:
            st.subheader("Year Summary (YoY)")

            if df_all.empty:
                st.info("No sales data yet.")
            else:
                d = df_all.copy()
                d["StartDate"] = pd.to_datetime(d["StartDate"], errors="coerce")
                d = d[d["StartDate"].notna()].copy()
                d["Year"] = d["StartDate"].dt.year.astype(int)

                years = sorted(d["Year"].unique().tolist())
                current_year = int(max(years)) if years else None
                prior_year = int(sorted(years)[-2]) if len(years) >= 2 else None

                # Helper sums
                def _sum(df_, col):
                    return float(df_[col].sum()) if (df_ is not None and not df_.empty and col in df_.columns) else 0.0

                def _pct(delta, base):
                    return (delta / base) if base else np.nan

                # Default compare window for KPIs: last two years (current + prior)
                a = d[d["Year"] == int(prior_year)].copy() if prior_year is not None else d[d["Year"] == int(current_year)].copy()
                b = d[d["Year"] == int(current_year)].copy()

                # Basis toggle impacts driver + concentration calculations
                basis = st.radio("Basis (tables + drivers)", options=["Sales", "Units"], index=0, horizontal=True, key="ys_basis")
                value_col = "Sales" if basis == "Sales" else "Units"

                # =========================
                # KPIs (last 2 years)
                # =========================
                st.markdown("### KPIs (current year vs prior year, plus vs prior-years average)")

                uA, uB = _sum(a, "Units"), _sum(b, "Units")   # A=prior year, B=current year
                sA, sB = _sum(a, "Sales"), _sum(b, "Sales")
                uD, sD = uB - uA, sB - sA
                uP, sP = _pct(uD, uA), _pct(sD, sA)

                labelA = str(prior_year) if prior_year is not None else (str(current_year) if current_year is not None else "—")
                labelB = str(current_year) if current_year is not None else "—"

                # Current year vs all prior years (avg per year)
                prior_all = d[d["Year"] < int(current_year)].copy() if current_year is not None else d.iloc[0:0].copy()
                prior_years = sorted([y for y in years if y < int(current_year)]) if current_year is not None else []
                if prior_years:
                    prior_by_year = prior_all.groupby("Year", as_index=False).agg(Units=("Units","sum"), Sales=("Sales","sum"))
                    avg_units = float(prior_by_year["Units"].mean()) if not prior_by_year.empty else 0.0
                    avg_sales = float(prior_by_year["Sales"].mean()) if not prior_by_year.empty else 0.0
                else:
                    avg_units, avg_sales = 0.0, 0.0

                du_avg = uB - avg_units
                ds_avg = sB - avg_sales
                pu_avg = _pct(du_avg, avg_units)
                ps_avg = _pct(ds_avg, avg_sales)

                # 6 KPI cards = 6 datapoints
                k1, k2, k3, k4, k5, k6 = st.columns(6)

                # Units: current + prior
                k1.metric(f"Units ({labelB})", fmt_int(uB),
                          delta=(f"{fmt_int_signed(uD)} ({uP*100:.1f}%)" if (prior_year is not None and pd.notna(uP)) else (fmt_int_signed(uD) if prior_year is not None else None)))

                # Prior year card with inverse delta (so it reads as change to current)
                k2.metric(f"Units ({labelA})", fmt_int(uA))

                # Sales: current + prior
                k3.metric(f"Sales ({labelB})", fmt_currency(sB),
                          delta=(f"{fmt_currency_signed(sD)} ({sP*100:.1f}%)" if (prior_year is not None and pd.notna(sP)) else (fmt_currency_signed(sD) if prior_year is not None else None)))

                k4.metric(f"Sales ({labelA})", fmt_currency(sA))

                # Current vs all prior-years average
                k5.metric(f"Units vs prior-years avg ({len(prior_years)} yrs)", fmt_int(uB), delta=(f"{fmt_int_signed(du_avg)} ({pu_avg*100:.1f}%)" if (prior_years and pd.notna(pu_avg)) else (fmt_int_signed(du_avg) if prior_years else "—")))
                k6.metric(f"Sales vs prior-years avg ({len(prior_years)} yrs)", fmt_currency(sB), delta=(f"{fmt_currency_signed(ds_avg)} ({ps_avg*100:.1f}%)" if (prior_years and pd.notna(ps_avg)) else (fmt_currency_signed(ds_avg) if prior_years else "—")))


                # =========================
                # YoY driver breakdown (auto: prior year -> current year)
                # =========================
                st.markdown("### YoY driver breakdown (auto: prior year → current year)")

                if prior_year is None:
                    st.info("Add at least two years of data to see YoY drivers.")
                else:
                    sku_a = a.groupby("SKU", as_index=False).agg(Units_A=("Units","sum"), Sales_A=("Sales","sum"))
                    sku_b = b.groupby("SKU", as_index=False).agg(Units_B=("Units","sum"), Sales_B=("Sales","sum"))
                    sku = sku_a.merge(sku_b, on="SKU", how="outer").fillna(0.0)

                    sku["A_val"] = sku["Sales_A"] if value_col == "Sales" else sku["Units_A"]
                    sku["B_val"] = sku["Sales_B"] if value_col == "Sales" else sku["Units_B"]
                    sku["Delta"] = sku["B_val"] - sku["A_val"]

                    sku["Bucket"] = "Same (flat)"
                    sku.loc[(sku["A_val"] == 0) & (sku["B_val"] > 0), "Bucket"] = "New SKUs"
                    sku.loc[(sku["A_val"] > 0) & (sku["B_val"] == 0), "Bucket"] = "Lost SKUs"
                    sku.loc[(sku["A_val"] > 0) & (sku["B_val"] > 0) & (sku["Delta"] > 0), "Bucket"] = "Same SKUs – Growth"
                    sku.loc[(sku["A_val"] > 0) & (sku["B_val"] > 0) & (sku["Delta"] < 0), "Bucket"] = "Same SKUs – Decline"

                    total_delta = float(sku["Delta"].sum())

                    def _b(name):
                        return float(sku.loc[sku["Bucket"] == name, "Delta"].sum())

                    b_new, b_lost = _b("New SKUs"), _b("Lost SKUs")
                    b_grow, b_decl = _b("Same SKUs – Growth"), _b("Same SKUs – Decline")

                    def _pct_of_delta(x):
                        return (x / total_delta) if total_delta else np.nan

                    def _fmt(v):
                        return fmt_currency_signed(v) if value_col == "Sales" else fmt_int_signed(v)

                    cD1, cD2, cD3, cD4 = st.columns(4)
                    cD1.metric("New SKUs", _fmt(b_new), delta=(f"{_pct_of_delta(b_new)*100:.1f}%" if pd.notna(_pct_of_delta(b_new)) else "—"))
                    cD2.metric("Lost SKUs", _fmt(b_lost), delta=(f"{_pct_of_delta(b_lost)*100:.1f}%" if pd.notna(_pct_of_delta(b_lost)) else "—"))
                    cD3.metric("Same SKUs – Growth", _fmt(b_grow), delta=(f"{_pct_of_delta(b_grow)*100:.1f}%" if pd.notna(_pct_of_delta(b_grow)) else "—"))
                    cD4.metric("Same SKUs – Decline", _fmt(b_decl), delta=(f"{_pct_of_delta(b_decl)*100:.1f}%" if pd.notna(_pct_of_delta(b_decl)) else "—"))

                    with st.expander("Top SKU drivers", expanded=False):
                        tp = sku.sort_values("Delta", ascending=False).head(25).copy()
                        tn = sku.sort_values("Delta", ascending=True).head(25).copy()

                        tp2 = tp[["SKU","A_val","B_val","Delta","Bucket"]].rename(columns={"A_val": labelA, "B_val": labelB})
                        tn2 = tn[["SKU","A_val","B_val","Delta","Bucket"]].rename(columns={"A_val": labelA, "B_val": labelB})

                        st.markdown("**Top increases**")
                        tp2_disp = tp2.copy()
                        for c in [labelA, labelB, "Delta"]:
                            if c in tp2_disp.columns:
                                tp2_disp[c] = tp2_disp[c].apply(_fmt)
                        tp2_disp = make_unique_columns(tp2_disp)
                        st.dataframe(tp2_disp, use_container_width=True, height=_table_height(tp2_disp, max_px=700), hide_index=True)

                        st.markdown("**Top declines**")
                        tn2_disp = tn2.copy()
                        for c in [labelA, labelB, "Delta"]:
                            if c in tn2_disp.columns:
                                tn2_disp[c] = tn2_disp[c].apply(_fmt)
                        tn2_disp = make_unique_columns(tn2_disp)
                        st.dataframe(tn2_disp, use_container_width=True, height=_table_height(tn2_disp, max_px=700), hide_index=True)

                # =========================
                # Concentration risk (ALL YEARS, retailer + vendor)
                # =========================
                st.markdown("### Concentration risk (all years)")

                def _top_share(df_year, group_col, topn):
                    g = df_year.groupby(group_col, as_index=False).agg(val=(value_col, "sum"))
                    total = float(g["val"].sum())
                    if total <= 0:
                        return 0.0
                    return float(g.sort_values("val", ascending=False).head(topn)["val"].sum()) / total

                rows = []
                for y in years:
                    dy = d[d["Year"] == int(y)].copy()
                    rows.append({
                        "Year": int(y),
                        "Top 1 Retailer %": _top_share(dy, "Retailer", 1),
                        "Top 3 Retailers %": _top_share(dy, "Retailer", 3),
                        "Top 5 Retailers %": _top_share(dy, "Retailer", 5),
                        "Top 1 Vendor %": _top_share(dy, "Vendor", 1),
                        "Top 3 Vendors %": _top_share(dy, "Vendor", 3),
                        "Top 5 Vendors %": _top_share(dy, "Vendor", 5),
                    })
                conc = pd.DataFrame(rows)
                conc_disp = conc.copy()
                try:
                    st.dataframe(conc_disp.style.format({c: (lambda v: f"{v*100:.1f}%") for c in conc_disp.columns if c != "Year"}),
                                 use_container_width=True, hide_index=True)
                except Exception:
                    # fallback without Styler (rare streamlit/pyarrow edge cases)
                    for c in [c for c in conc_disp.columns if c != "Year"]:
                        conc_disp[c] = conc_disp[c].apply(lambda v: f"{v*100:.1f}%")
                    st.dataframe(conc_disp, use_container_width=True, hide_index=True)

                # =========================
                # Retailer summary (YEAR PICKER ONLY FOR THIS TABLE)
                # =========================

                st.markdown("#### Concentration breakdown (click to expand)")
                st.caption("Expand a year to see exactly which retailers/vendors make up the Top 1 / Top 3 / Top 5 shares.")

                def _top_list(df_year, group_col, topn):
                    g = df_year.groupby(group_col, as_index=False).agg(val=(value_col, "sum")).sort_values("val", ascending=False)
                    total = float(g["val"].sum())
                    if total <= 0 or g.empty:
                        return g.assign(Share=0.0).head(0)
                    g["Share"] = g["val"] / total
                    return g.head(topn)

                for y in years[::-1]:  # newest first
                    dy = d[d["Year"] == int(y)].copy()
                    with st.expander(f"Year {int(y)} – show Top Retailers/Vendors", expanded=False):
                        cL, cR = st.columns(2)
                        with cL:
                            st.markdown("**Top Retailers**")
                            tr = _top_list(dy, "Retailer", 10)
                            if tr.empty:
                                st.write("—")
                            else:
                                tr_disp = tr.rename(columns={"val": value_col})
                                tr_disp[value_col] = tr_disp[value_col].apply(fmt_currency if value_col=="Sales" else fmt_int)
                                tr_disp["Share"] = tr_disp["Share"].apply(lambda v: f"{v*100:.1f}%")
                                st.dataframe(tr_disp[["Retailer", value_col, "Share"]], use_container_width=True, hide_index=True)
                                for n in [1, 3, 5]:
                                    t = _top_list(dy, "Retailer", n)
                                    if not t.empty:
                                        share = float(t["Share"].sum()) * 100
                                        names = ", ".join(t["Retailer"].astype(str).tolist())
                                        st.caption(f"Top {n}: {share:.1f}% — {names}")
                        with cR:
                            st.markdown("**Top Vendors**")
                            tv = _top_list(dy, "Vendor", 10)
                            if tv.empty:
                                st.write("—")
                            else:
                                tv_disp = tv.rename(columns={"val": value_col})
                                tv_disp[value_col] = tv_disp[value_col].apply(fmt_currency if value_col=="Sales" else fmt_int)
                                tv_disp["Share"] = tv_disp["Share"].apply(lambda v: f"{v*100:.1f}%")
                                st.dataframe(tv_disp[["Vendor", value_col, "Share"]], use_container_width=True, hide_index=True)
                                for n in [1, 3, 5]:
                                    t = _top_list(dy, "Vendor", n)
                                    if not t.empty:
                                        share = float(t["Share"].sum()) * 100
                                        names = ", ".join(t["Vendor"].astype(str).tolist())
                                        st.caption(f"Top {n}: {share:.1f}% — {names}")


                st.markdown("### Retailer summary")
                rs_year = st.selectbox("Retailer summary year", options=years, index=(len(years)-1), key="ys_rs_year")
                rs_prev = int(rs_year) - 1 if (int(rs_year) - 1) in years else None

                r0 = d[d["Year"] == int(rs_year)].groupby("Retailer", as_index=False).agg(Units=("Units","sum"), Sales=("Sales","sum"))
                if rs_prev is not None:
                    r1 = d[d["Year"] == int(rs_prev)].groupby("Retailer", as_index=False).agg(Units_P=("Units","sum"), Sales_P=("Sales","sum"))
                    r = r0.merge(r1, on="Retailer", how="left").fillna(0.0)
                    r["Units Δ"] = r["Units"] - r["Units_P"]
                    r["Sales Δ"] = r["Sales"] - r["Sales_P"]
                else:
                    r = r0.copy()

                r = r.sort_values("Sales", ascending=False)
                r_show = r.copy()
                sty_r = r_show.style.format({
                    'Units': fmt_int,
                    'Sales': fmt_currency,
                    'Units_P': fmt_int,
                    'Sales_P': fmt_currency,
                    'Units Δ': fmt_int_signed,
                    'Sales Δ': fmt_currency_signed,
                })
                for c in ['Units Δ','Sales Δ']:
                    if c in r_show.columns:
                        sty_r = sty_r.applymap(lambda v: _diff_color(v), subset=[c])
                r_show = make_unique_columns(r_show)
                st.dataframe(sty_r, use_container_width=True, height=_table_height(r_show, max_px=750), hide_index=True)

                # =========================
                # Vendor summary (YEAR PICKER ONLY FOR THIS TABLE)
                # =========================
                st.markdown("### Vendor summary")
                vs_year = st.selectbox("Vendor summary year", options=years, index=(len(years)-1), key="ys_vs_year")
                vs_prev = int(vs_year) - 1 if (int(vs_year) - 1) in years else None

                v0 = d[d["Year"] == int(vs_year)].groupby("Vendor", as_index=False).agg(Units=("Units","sum"), Sales=("Sales","sum"))
                if vs_prev is not None:
                    v1 = d[d["Year"] == int(vs_prev)].groupby("Vendor", as_index=False).agg(Units_P=("Units","sum"), Sales_P=("Sales","sum"))
                    vv = v0.merge(v1, on="Vendor", how="left").fillna(0.0)
                    vv["Units Δ"] = vv["Units"] - vv["Units_P"]
                    vv["Sales Δ"] = vv["Sales"] - vv["Sales_P"]
                else:
                    vv = v0.copy()

                vv = vv.sort_values("Sales", ascending=False)
                vv_show = vv.copy()
                sty_v = vv_show.style.format({
                    'Units': fmt_int,
                    'Sales': fmt_currency,
                    'Units_P': fmt_int,
                    'Sales_P': fmt_currency,
                    'Units Δ': fmt_int_signed,
                    'Sales Δ': fmt_currency_signed,
                })
                for c in ['Units Δ','Sales Δ']:
                    if c in vv_show.columns:
                        sty_v = sty_v.applymap(lambda v: _diff_color(v), subset=[c])
                vv_show = make_unique_columns(vv_show)
                st.dataframe(sty_v, use_container_width=True, height=_table_height(vv_show, max_px=750), hide_index=True)





    def make_weekly_summary_pdf_bytes(title: str,
                                    highlights: list,
                                    kpi: dict,
                                    top_retailers: pd.DataFrame,
                                    top_vendors: pd.DataFrame,
                                    top_skus: pd.DataFrame,
                                    movers: pd.DataFrame,
                                    vendor_decl: pd.DataFrame,
                                    retailer_decl: pd.DataFrame,
                                    momentum: pd.DataFrame,
                                    df_all: pd.DataFrame,
                                    logo_path: str = None,
                                    executive_takeaway: str = None,
                                    drivers_df: pd.DataFrame = None,
                                    opportunities_df: pd.DataFrame = None) -> bytes:
        """
        Professional 3-page executive weekly PDF.
        Page 1: Executive Snapshot (KPIs + Performance Summary + Top tables)
        Page 2: Operational Movement (Trend chart + Biggest Movers + Declines)
        Page 3: Strategic Momentum (Momentum Leaders)
        """
        from io import BytesIO
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
                                        Image, KeepTogether, KeepInFrame, PageBreak)
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.pdfgen import canvas as pdfcanvas
        from datetime import datetime
        import os as _os

        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name="H1", parent=styles["Heading1"], fontSize=16, leading=18, spaceAfter=8))
        styles.add(ParagraphStyle(name="H2", parent=styles["Heading2"], fontSize=12, leading=14, spaceAfter=6))
        styles.add(ParagraphStyle(name="Body", parent=styles["Normal"], fontSize=9.5, leading=12))
        styles.add(ParagraphStyle(name="Small", parent=styles["Normal"], fontSize=8, leading=10, textColor=colors.HexColor("#6b7280")))

        def _wow_color(v):
            """Green for positive, red for negative, neutral for zero/unknown."""
            try:
                s = str(v)
                import re as _re
                s2 = _re.sub(r"[^0-9\-\.]", "", s)
                if s2 in ("", "-", ".", "-."):
                    raise ValueError("no number")
                x = float(s2)
                if x > 0:
                    return colors.HexColor("#2ecc71")
                if x < 0:
                    return colors.HexColor("#e74c3c")
            except Exception:
                pass
            return colors.HexColor("#111827")

        def _make_table(df: pd.DataFrame, header_bg="#111827", max_rows=10, col_widths=None,
                        wow_col_name=None, right_align_cols=None):
            if df is None or df.empty:
                return Paragraph("No data.", styles["Body"])
            tshow = df.copy().head(max_rows)

            # Format numeric columns intelligently (Sales = currency, Units = integers)
            disp = tshow.copy()
            for c in disp.columns:
                col_name = str(c).lower()
                ser = disp[c]
                num = pd.to_numeric(ser, errors="coerce")
                is_num = num.notna().any()
                if not is_num:
                    disp[c] = ser.astype(str)
                    continue

                def _fmt_currency(v):
                    try:
                        v = float(v)
                    except Exception:
                        return "—"
                    return f"-${abs(v):,.2f}" if v < 0 else f"${v:,.2f}"

                def _fmt_int(v):
                    try:
                        return f"{int(round(float(v))):,}"
                    except Exception:
                        return "—"

                if ("sale" in col_name) or ("revenue" in col_name) or ("$" in col_name):
                    disp[c] = num.map(_fmt_currency)
                elif ("unit" in col_name) or ("qty" in col_name) or ("quantity" in col_name):
                    disp[c] = num.map(_fmt_int)
                else:
                    def _fmt_generic(v):
                        # Robust numeric formatting: handles strings like "$1,234.56", "(123.45)", "—"
                        if v is None or (isinstance(v, float) and (math.isnan(v) or math.isinf(v))):
                            return "—"
                        # pandas NA
                        try:
                            if pd.isna(v):
                                return "—"
                        except Exception:
                            pass
                        # If already numeric, keep
                        vv = v
                        if isinstance(vv, str):
                            s = vv.strip()
                            if s in ("", "—", "-", "N/A", "NA", "nan"):
                                return "—"
                            neg = False
                            if s.startswith("(") and s.endswith(")"):
                                neg = True
                                s = s[1:-1]
                            s = s.replace("$", "").replace(",", "").strip()
                            try:
                                vv = float(s)
                                if neg:
                                    vv = -vv
                            except Exception:
                                return str(v)
                        try:
                            vv = float(vv)
                        except Exception:
                            return str(v)
                        if not math.isfinite(vv):
                            return "—"
                        if abs(vv - round(vv)) < 1e-9:
                            return f"{int(round(vv)):,}"
                        return f"{vv:,.2f}"


            data = [disp.columns.tolist()] + disp.values.tolist()
            tbl = Table(data, hAlign="LEFT", colWidths=col_widths)

            base = [
                ("BACKGROUND",(0,0),(-1,0), colors.HexColor(header_bg)),
                ("TEXTCOLOR",(0,0),(-1,0), colors.white),
                ("FONTNAME",(0,0),(-1,0), "Helvetica-Bold"),
                ("FONTSIZE",(0,0),(-1,0), 9),
                ("GRID",(0,0),(-1,-1), 0.25, colors.HexColor("#d1d5db")),
                ("FONTNAME",(0,1),(-1,-1), "Helvetica"),
                ("FONTSIZE",(0,1),(-1,-1), 8),
                ("ROWBACKGROUNDS",(0,1),(-1,-1), [colors.white, colors.HexColor("#f9fafb")]),
                ("VALIGN",(0,0),(-1,-1), "TOP"),
                ("LEFTPADDING",(0,0),(-1,-1), 4),
                ("RIGHTPADDING",(0,0),(-1,-1), 4),
                ("TOPPADDING",(0,0),(-1,-1), 3),
                ("BOTTOMPADDING",(0,0),(-1,-1), 3),
            ]

            if right_align_cols:
                for col in right_align_cols:
                    if col in tshow.columns:
                        j = tshow.columns.get_loc(col)
                        base.append(("ALIGN",(j,1),(j,-1),"RIGHT"))
                        base.append(("ALIGN",(j,0),(j,0),"RIGHT"))

            if wow_col_name and wow_col_name in tshow.columns:
                j = tshow.columns.get_loc(wow_col_name)
                for i in range(1, len(data)):
                    base.append(("TEXTCOLOR",(j,i),(j,i), _wow_color(data[i][j])))
                    base.append(("FONTNAME",(j,i),(j,i), "Helvetica-Bold"))

            tbl.setStyle(TableStyle(base))
            return tbl

        def _make_trend_chart(df_all: pd.DataFrame):
            if df_all is None or df_all.empty or "EndDate" not in df_all.columns:
                return None
            d = df_all.copy()
            d["EndDate"] = pd.to_datetime(d["EndDate"], errors="coerce")
            d = d[d["EndDate"].notna()].copy()
            if d.empty:
                return None
            wk = d.groupby("EndDate", as_index=False)[["Sales"]].sum().sort_values("EndDate").tail(12)
            if wk.empty:
                return None
            fig = plt.figure(figsize=(6.4, 2.2))
            ax = fig.add_subplot(111)
            ax.plot(wk["EndDate"], wk["Sales"])
            ax.set_title("Total Sales – Last 12 Weeks")
            ax.set_ylabel("Sales ($)")
            ax.tick_params(axis='x', labelrotation=45)
            ax.grid(True, linewidth=0.5, alpha=0.4)
            fig.tight_layout()
            img_buf = BytesIO()
            fig.savefig(img_buf, format="png", dpi=150)
            plt.close(fig)
            img_buf.seek(0)
            return img_buf

        generated = datetime.now().strftime("%Y-%m-%d %H:%M")
        def _on_page(canv: pdfcanvas.Canvas, doc):
            canv.saveState()
            w, h = letter
            canv.setFillColor(colors.HexColor("#111827"))
            canv.rect(0, h-0.65*inch, w, 0.65*inch, fill=1, stroke=0)

            if logo_path and _os.path.exists(logo_path):
                try:
                    canv.drawImage(logo_path, 0.55*inch, h-0.60*inch, width=1.3*inch, height=0.45*inch, mask='auto', preserveAspectRatio=True, anchor='sw')
                except Exception:
                    pass

            canv.setFillColor(colors.white)
            canv.setFont("Helvetica-Bold", 12)
            canv.drawString(2.1*inch, h-0.40*inch, title)

            canv.setFont("Helvetica", 9)
            canv.drawRightString(w-0.55*inch, h-0.40*inch, f"Generated: {generated}")

            canv.setFillColor(colors.HexColor("#6b7280"))
            canv.setFont("Helvetica", 8)
            canv.drawString(0.55*inch, 0.45*inch, "Cornerstone Products Group – Confidential (Internal Use Only)")
            canv.drawRightString(w-0.55*inch, 0.45*inch, f"Page {doc.page}")
            canv.restoreState()

        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=letter,
                                leftMargin=0.55*inch, rightMargin=0.55*inch,
                                topMargin=0.85*inch, bottomMargin=0.7*inch)
        story = []

        # PAGE 1
        story.append(Spacer(1, 0.15*inch))
        story.append(Paragraph("Executive Snapshot", styles["H1"]))

        # KPI panel
        kpi_lines = []
        if isinstance(kpi, dict) and kpi:
            for k, v in kpi.items():
                kpi_lines.append([str(k), str(v)])
        if not kpi_lines:
            kpi_lines = [["Sales", "—"], ["Units", "—"], ["WoW Sales", "—"], ["WoW Units", "—"]]
        kpi_tbl = Table(kpi_lines, colWidths=[2.2*inch, 1.6*inch])
        kpi_tbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,-1), colors.HexColor("#f3f4f6")),
            ("BOX",(0,0),(-1,-1), 0.5, colors.HexColor("#d1d5db")),
            ("INNERGRID",(0,0),(-1,-1), 0.25, colors.HexColor("#e5e7eb")),
            ("FONTNAME",(0,0),(0,-1), "Helvetica-Bold"),
            ("FONTSIZE",(0,0),(-1,-1), 9),
            ("LEFTPADDING",(0,0),(-1,-1), 8),
            ("RIGHTPADDING",(0,0),(-1,-1), 8),
            ("TOPPADDING",(0,0),(-1,-1), 6),
            ("BOTTOMPADDING",(0,0),(-1,-1), 6),
        ]))

        perf_bullets = []
        if highlights:
            for b in highlights[:6]:
                perf_bullets.append(f"• {b}")
        if not perf_bullets:
            perf_bullets = ["• No highlights generated for this week."]
        perf_par = Paragraph("<br/>".join(perf_bullets), styles["Body"])
        perf_box = Table([[Paragraph("Performance Summary", styles["H2"])],[perf_par]], colWidths=[3.9*inch])
        perf_box.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,1), colors.HexColor("#f9fafb")),
            ("BOX",(0,0),(-1,1), 0.5, colors.HexColor("#d1d5db")),
            ("LEFTPADDING",(0,0),(-1,-1), 10),
            ("RIGHTPADDING",(0,0),(-1,-1), 10),
            ("TOPPADDING",(0,0),(-1,-1), 8),
            ("BOTTOMPADDING",(0,0),(-1,-1), 8),
        ]))
        top_row = Table([[kpi_tbl, perf_box]], colWidths=[2.0*inch, doc.width-2.0*inch])
        top_row.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1), "TOP"), ("LEFTPADDING",(0,0),(-1,-1), 0), ("RIGHTPADDING",(0,0),(-1,-1), 0)]))
        story.append(top_row)
        story.append(Spacer(1, 0.14*inch))
        if executive_takeaway:
            story.append(Spacer(1, 0.06*inch))
            takeaway_box = Table([[Paragraph("<b>Executive takeaway:</b> " + html.escape(str(executive_takeaway)), styles["Body"])]],
                                 colWidths=[doc.width])
            takeaway_box.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(-1,-1), colors.HexColor("#f9fafb")),
                ("BOX",(0,0),(-1,-1), 0.5, colors.HexColor("#d1d5db")),
                ("LEFTPADDING",(0,0),(-1,-1), 10),
                ("RIGHTPADDING",(0,0),(-1,-1), 10),
                ("TOPPADDING",(0,0),(-1,-1), 8),
                ("BOTTOMPADDING",(0,0),(-1,-1), 8),
            ]))
            story.append(takeaway_box)
            story.append(Spacer(1, 0.10*inch))


        left_stack = [Paragraph("Top Retailers", styles["H2"]),
                      _make_table(top_retailers, max_rows=5, wow_col_name="WoW $ Diff", right_align_cols=["Units","Sales","WoW $ Diff"]),
                      Spacer(1, 0.12*inch),
                      Paragraph("Top Vendors", styles["H2"]),
                      _make_table(top_vendors, max_rows=5, wow_col_name="WoW $ Diff", right_align_cols=["Units","Sales","WoW $ Diff"])]
        right_stack = [Paragraph("Top SKUs", styles["H2"]),
                       _make_table(top_skus, max_rows=10, wow_col_name="WoW $ Diff", right_align_cols=["Units","Sales","WoW $ Diff"])]

        gutter = 0.18*inch
        left_w = (doc.width - gutter) * 0.49
        right_w = (doc.width - gutter) * 0.51
        left_cell = KeepInFrame(left_w, 6.6*inch, left_stack, mode="shrink")
        right_cell = KeepInFrame(right_w, 6.6*inch, right_stack, mode="shrink")
        split = Table([[left_cell, "", right_cell]], colWidths=[left_w, gutter, right_w])
        split.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"TOP"), ("LEFTPADDING",(0,0),(-1,-1), 0), ("RIGHTPADDING",(0,0),(-1,-1), 0)]))
        story.append(split)

        story.append(PageBreak())

        # PAGE 2
        story.append(Paragraph("Operational Movement", styles["H1"]))

        # Top 3 drivers of WoW sales change + Top opportunities
        if drivers_df is not None and hasattr(drivers_df, "empty") and (not drivers_df.empty):
            story.append(Spacer(1, 0.06*inch))
            story.append(KeepTogether([Paragraph("Top 3 Drivers of WoW Sales Change", styles["H2"]),
                                       _make_table(drivers_df, max_rows=3, wow_col_name="WoW Sales ($)",
                                                   right_align_cols=["WoW Sales ($)","Sales (This Week)","Units (This Week)","WoW Units"])]))
            story.append(Spacer(1, 0.12*inch))

        if opportunities_df is not None and hasattr(opportunities_df, "empty") and (not opportunities_df.empty):
            story.append(KeepTogether([Paragraph("Top Opportunities (Positive WoW)", styles["H2"]),
                                       _make_table(opportunities_df, max_rows=8, wow_col_name="WoW Sales ($)",
                                                   right_align_cols=["WoW Sales ($)","Sales (This Week)","Units (This Week)","WoW Units"])]))
            story.append(Spacer(1, 0.14*inch))

        story.append(Spacer(1, 0.08*inch))

        chart_buf = _make_trend_chart(df_all)
        if chart_buf is not None:
            img = Image(chart_buf, width=doc.width, height=2.0*inch)
            chart_box = Table([[img]], colWidths=[doc.width])
            chart_box.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(-1,-1), colors.HexColor("#f9fafb")),
                ("BOX",(0,0),(-1,-1), 0.5, colors.HexColor("#d1d5db")),
                ("LEFTPADDING",(0,0),(-1,-1), 6),
                ("RIGHTPADDING",(0,0),(-1,-1), 6),
                ("TOPPADDING",(0,0),(-1,-1), 6),
                ("BOTTOMPADDING",(0,0),(-1,-1), 6),
            ]))
            story.append(chart_box)
            story.append(Spacer(1, 0.14*inch))

        story.append(KeepTogether([Paragraph("Biggest Movers", styles["H2"]), _make_table(movers, max_rows=12)]))
        story.append(Spacer(1, 0.14*inch))

        story.append(KeepTogether([Paragraph("Declining Vendors", styles["H2"]), _make_table(vendor_decl, max_rows=10, wow_col_name="WoW Sales")]))
        story.append(Spacer(1, 0.12*inch))
        story.append(KeepTogether([Paragraph("Declining Retailers", styles["H2"]), _make_table(retailer_decl, max_rows=10, wow_col_name="WoW Sales")]))

        story.append(PageBreak())

        # PAGE 3
        story.append(Paragraph("Strategic Momentum", styles["H1"]))
        story.append(Paragraph("Momentum Leaders (Last 12 Weeks)", styles["H2"]))
        story.append(Spacer(1, 0.06*inch))
        mom_flow = [_make_table(momentum, max_rows=15, right_align_cols=["Momentum","Up Weeks","Down Weeks","Units_Last","Sales_Last"])]
        story.append(KeepInFrame(doc.width, 8.5*inch, mom_flow, mode="shrink"))

        doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)
        return buf.getvalue()



    def make_comparison_pdf_bytes(title: str,
                                 subtitle: str,
                                 kpi: dict,
                                 retailers: pd.DataFrame,
                                 drivers: pd.DataFrame,
                                 top_increase: pd.DataFrame,
                                 top_decrease: pd.DataFrame,
                                 momentum: pd.DataFrame,
                                 logo_path: str = None) -> bytes:
        """Executive-style comparison PDF using the same table/box styling as Weekly Summary."""
        try:
            from io import BytesIO
            from reportlab.lib.pagesizes import letter
            from reportlab.lib.units import inch
            from reportlab.lib import colors
            from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
                                            Image, KeepTogether, KeepInFrame, PageBreak)
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.pdfgen import canvas as pdfcanvas
            from datetime import datetime
            import os as _os
            import html
        except Exception:
            return b""

        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name="H1", parent=styles["Heading1"], fontSize=16, leading=18, spaceAfter=8))
        styles.add(ParagraphStyle(name="H2", parent=styles["Heading2"], fontSize=12, leading=14, spaceAfter=6))
        styles.add(ParagraphStyle(name="Body", parent=styles["Normal"], fontSize=9.5, leading=12))
        styles.add(ParagraphStyle(name="Small", parent=styles["Normal"], fontSize=8, leading=10, textColor=colors.HexColor("#6b7280")))

        def _wow_color(v):
            try:
                s = str(v)
                import re as _re
                s2 = _re.sub(r"[^0-9\-\.]", "", s)
                if s2 in ("", "-", ".", "-."):
                    raise ValueError("no number")
                x = float(s2)
                if x > 0:
                    return colors.HexColor("#2ecc71")
                if x < 0:
                    return colors.HexColor("#e74c3c")
            except Exception:
                pass
            return colors.HexColor("#111827")

        def _make_table(df: pd.DataFrame, header_bg="#111827", max_rows=999, col_widths=None,
                        wow_col_name=None, right_align_cols=None):
            if df is None or df.empty:
                return Paragraph("No data.", styles["Body"])
            tshow = df.copy().head(max_rows)

            disp = tshow.copy()
            for c in disp.columns:
                col_name = str(c).lower()
                ser = disp[c]
                num = pd.to_numeric(ser, errors="coerce")
                is_num = num.notna().any()
                if not is_num:
                    disp[c] = ser.astype(str)
                    continue
                def _fmt_currency(v):
                    try:
                        v = float(v)
                    except Exception:
                        return "—"
                    return f"${v:,.2f}"
                def _fmt_int(v):
                    try:
                        v = float(v)
                    except Exception:
                        return "—"
                    return f"{int(round(v)):,}"
                def _fmt_signed_currency(v):
                    try:
                        v = float(v)
                    except Exception:
                        return "—"
                    sign = "+" if v > 0 else ""
                    return f"{sign}${v:,.2f}"
                def _fmt_signed_int(v):
                    try:
                        v = float(v)
                    except Exception:
                        return "—"
                    sign = "+" if v > 0 else ""
                    return f"{sign}{int(round(v)):,}"

                if "sales" in col_name or "$" in col_name:
                    if "Δ" in str(c) or "delta" in col_name or "diff" in col_name or "change" in col_name:
                        disp[c] = num.apply(_fmt_signed_currency)
                    else:
                        disp[c] = num.apply(_fmt_currency)
                elif "unit" in col_name or "qty" in col_name:
                    if "Δ" in str(c) or "delta" in col_name or "diff" in col_name or "change" in col_name:
                        disp[c] = num.apply(_fmt_signed_int)
                    else:
                        disp[c] = num.apply(_fmt_int)
                elif "Δ" in str(c) or "delta" in col_name or "diff" in col_name or "change" in col_name:
                    disp[c] = num.apply(lambda x: f"{x:+.2f}")
                else:
                    disp[c] = num.apply(lambda x: f"{x:,.2f}")

            data = [list(disp.columns)] + disp.astype(str).values.tolist()
            if col_widths is None:
                col_widths = [None] * len(disp.columns)

            tbl = Table(data, colWidths=col_widths)
            ts = [
                ("BACKGROUND", (0,0), (-1,0), colors.HexColor(header_bg)),
                ("TEXTCOLOR", (0,0), (-1,0), colors.white),
                ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE", (0,0), (-1,0), 9),
                ("BOTTOMPADDING", (0,0), (-1,0), 6),
                ("TOPPADDING", (0,0), (-1,0), 6),
                ("GRID", (0,0), (-1,-1), 0.25, colors.HexColor("#e5e7eb")),
                ("FONTSIZE", (0,1), (-1,-1), 8.5),
                ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#f9fafb")]),
                ("LEFTPADDING", (0,0), (-1,-1), 6),
                ("RIGHTPADDING", (0,0), (-1,-1), 6),
                ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ]
            # right align numeric-like columns
            if right_align_cols:
                for idx, cn in enumerate(disp.columns):
                    if str(cn) in right_align_cols:
                        ts.append(("ALIGN", (idx,0), (idx,-1), "RIGHT"))
            # color delta column
            if wow_col_name and wow_col_name in disp.columns:
                j = list(disp.columns).index(wow_col_name)
                for r in range(1, len(data)):
                    ts.append(("TEXTCOLOR", (j,r), (j,r), _wow_color(disp.iloc[r-1, j])))
                    ts.append(("FONTNAME", (j,r), (j,r), "Helvetica-Bold"))
            tbl.setStyle(TableStyle(ts))
            return tbl

        generated = datetime.now().strftime("%Y-%m-%d %H:%M")
        def _on_page(canv: pdfcanvas.Canvas, doc):
            canv.saveState()
            w, h = letter
            canv.setFillColor(colors.HexColor("#111827"))
            canv.rect(0, h-0.65*inch, w, 0.65*inch, fill=1, stroke=0)
            if logo_path and _os.path.exists(logo_path):
                try:
                    canv.drawImage(logo_path, 0.55*inch, h-0.60*inch, width=1.3*inch, height=0.45*inch, mask='auto', preserveAspectRatio=True, anchor='sw')
                except Exception:
                    pass
            canv.setFillColor(colors.white)
            canv.setFont("Helvetica-Bold", 12)
            canv.drawString(2.1*inch, h-0.40*inch, title)
            canv.setFont("Helvetica", 9)
            canv.drawRightString(w-0.55*inch, h-0.40*inch, f"Generated: {generated}")
            canv.setFillColor(colors.HexColor("#6b7280"))
            canv.setFont("Helvetica", 8)
            canv.drawString(0.55*inch, 0.45*inch, "Cornerstone Products Group – Confidential (Internal Use Only)")
            canv.drawRightString(w-0.55*inch, 0.45*inch, f"Page {doc.page}")
            canv.restoreState()

        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=letter,
                                leftMargin=0.55*inch, rightMargin=0.55*inch,
                                topMargin=0.85*inch, bottomMargin=0.7*inch)
        story = []
        story.append(Spacer(1, 0.15*inch))
        story.append(Paragraph("Executive Comparison Snapshot", styles["H1"]))
        story.append(Paragraph(html.escape(subtitle or ""), styles["Small"]))
        story.append(Spacer(1, 0.12*inch))

        # KPI panel (same look as weekly)
        kpi_lines = [[str(k), str(v)] for k,v in (kpi or {}).items()]
        if not kpi_lines:
            kpi_lines = [["Sales", "—"], ["Units", "—"], ["Sales Δ", "—"], ["Units Δ", "—"]]
        kpi_tbl = Table(kpi_lines, colWidths=[2.2*inch, 1.6*inch])
        kpi_tbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,-1), colors.HexColor("#f3f4f6")),
            ("BOX",(0,0),(-1,-1), 0.5, colors.HexColor("#d1d5db")),
            ("INNERGRID",(0,0),(-1,-1), 0.25, colors.HexColor("#e5e7eb")),
            ("FONTNAME",(0,0),(0,-1), "Helvetica-Bold"),
            ("FONTSIZE",(0,0),(-1,-1), 9),
            ("LEFTPADDING",(0,0),(-1,-1), 8),
            ("RIGHTPADDING",(0,0),(-1,-1), 8),
            ("TOPPADDING",(0,0),(-1,-1), 6),
            ("BOTTOMPADDING",(0,0),(-1,-1), 6),
        ]))
        story.append(kpi_tbl)
        story.append(Spacer(1, 0.16*inch))

        # Retailers (ALL)
        story.append(Paragraph("Retailers (All)", styles["H2"]))
        story.append(_make_table(retailers, wow_col_name="Sales Δ", right_align_cols=[c for c in (retailers.columns.tolist() if retailers is not None else []) if any(k in str(c).lower() for k in ["sales","unit","Δ","delta"]) ]))
        story.append(PageBreak())

        # Movement
        story.append(Paragraph("Operational Movement", styles["H1"]))
        if drivers is not None and not drivers.empty:
            story.append(KeepTogether([Paragraph("Top 5 Drivers (Largest |Sales Δ|)", styles["H2"]),
                                       _make_table(drivers, wow_col_name="Sales Δ")]))
            story.append(Spacer(1, 0.14*inch))

        story.append(KeepTogether([Paragraph("Top Increase SKUs", styles["H2"]), _make_table(top_increase, wow_col_name="Sales Δ")]))
        story.append(Spacer(1, 0.14*inch))
        story.append(KeepTogether([Paragraph("Top Decrease SKUs", styles["H2"]), _make_table(top_decrease, wow_col_name="Sales Δ")]))
        story.append(PageBreak())

        # Momentum
        story.append(Paragraph("Strategic Momentum", styles["H1"]))
        story.append(Paragraph("Momentum Leaders (Period A: Last4 − Prev4)", styles["H2"]))
        story.append(Spacer(1, 0.06*inch))
        story.append(KeepInFrame(doc.width, 8.5*inch, [_make_table(momentum, wow_col_name="Momentum", right_align_cols=["Momentum","Sales_Last","Sales_Prev","Units_Last","Units_Prev"])], mode="shrink"))

        doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)
        return buf.getvalue()


    def make_simple_pdf_bytes(title: str, lines: list[str], table_df: pd.DataFrame|None = None) -> bytes:
        """Simple PDF generator for weekly summaries."""
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.pdfgen import canvas
            from reportlab.lib.units import inch
        except Exception:
            body = title + "\n\n" + "\n".join(lines or [])
            return body.encode("utf-8", errors="ignore")

        buf = io.BytesIO()
        c = canvas.Canvas(buf, pagesize=letter)
        w, h = letter
        x = 0.75 * inch
        y = h - 0.9 * inch

        c.setFont("Helvetica-Bold", 16)
        c.drawString(x, y, title)
        y -= 0.4 * inch
        c.setFont("Helvetica", 10)

        for ln in (lines or []):
            if y < 0.75 * inch:
                c.showPage()
                y = h - 0.9 * inch
                c.setFont("Helvetica", 10)
            c.drawString(x, y, str(ln)[:180])
            y -= 0.22 * inch

        c.showPage()
        c.save()
        return buf.getvalue()


    render_tab_overview()

    render_tab_action_center()

    render_tab_momentum()


    render_tab_totals_dash()

    render_tab_top_skus()

    render_tab_wow_exc()

    render_tab_exec()

    render_tab_comparisons()

    render_tab_sku_intel()

    render_tab_forecasting()

    render_tab_alerts()

    render_tab_data_mgmt()

    render_tab_year_summary()
